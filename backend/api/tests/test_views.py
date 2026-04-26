from PyPDF2 import PdfReader, PdfWriter
from django.conf import settings
from django.core.exceptions import SuspiciousFileOperation
from django.utils._os import safe_join
from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from types import SimpleNamespace
from unittest.mock import patch, MagicMock
from PIL import Image

from io import BytesIO
from datetime import timedelta
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from django.utils import timezone

from rest_framework import status
from rest_framework.test import APIClient, APIRequestFactory, APISimpleTestCase
from rest_framework.test import force_authenticate
from rest_framework.response import Response

from api import views
from api.models import GroupMember
from api.views import (
    _delete_history_artifact_file,
    _resolve_download_filename,
    _sanitize_download_filename,
    upload,
)
from artifact_history.models import ArtifactHistory, HistoryExportArtifact
from artifact_history.serializers import HISTORY_CUSTOM_NAME_MAX_LENGTH
from authentication.models import User
from file_processing.services.export_service import (
    OutputCSVDownloadLookupError,
    OutputCSVGenerationError,
    OutputCSVMappingError,
    OutputExcelDownloadLookupError,
    OutputExcelDownloadStorageError,
    OutputExcelGenerationError,
    OutputLLMValidationError,
)

from file_processing.services.upload_service import MAX_FILE_SIZE


class BaseApiViewTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.factory = APIRequestFactory()


class HealthCheckViewTest(BaseApiViewTest):
    def test_health_endpoint_returns_200(self):
        response = self.client.get("/health/")
        self.assertEqual(response.status_code, 200)

    def test_health_endpoint_returns_correct_data(self):
        response = self.client.get("/health/")
        self.assertEqual(response.data["status"], "ok")
        self.assertEqual(response.data["message"], "Backend is running!")

    def test_health_endpoint_rejects_post(self):
        response = self.client.post("/health/")
        self.assertEqual(response.status_code, 405)


class AboutViewTest(BaseApiViewTest):
    def test_about_endpoint_returns_200(self):
        response = self.client.get("/about/")
        self.assertEqual(response.status_code, 200)

    def test_about_endpoint_returns_correct_data(self):
        response = self.client.get("/about/")
        self.assertEqual(response.data["team"], "PPL C - Equitek")
        self.assertEqual(response.data["project"], "Excel Generator")

    def test_about_endpoint_rejects_post(self):
        response = self.client.post("/about/")
        self.assertEqual(response.status_code, 405)


class HistoryListViewTest(BaseApiViewTest):
    def setUp(self):
        super().setUp()
        self.verified_user = User.objects.create_user(
            email="verified-history@example.com",
            name="Verified History",
            password="secret",
            status="verified",
        )
        self.unverified_user = User.objects.create_user(
            email="unverified-history@example.com",
            name="Unverified History",
            password="secret",
            status="unverified",
        )
        self.other_user = User.objects.create_user(
            email="other-history@example.com",
            name="Other History",
            password="secret",
            status="verified",
        )

    def _create_history(self, owner, original_name, created_at):
        return ArtifactHistory.objects.create(
            owner=owner,
            original_name=original_name,
            custom_name=None,
            output_json={
                "document_info": {"filename": original_name},
                "summary": {"table_count": 1},
                "content_data": [
                    {"table_name": "Sheet1", "headers": ["A"], "rows": [["1"]]}
                ],
            },
            status_processing="completed",
            created_at=created_at,
        )

    def test_history_list_returns_401_for_anonymous_user(self):
        response = self.client.get("/history/")

        self.assertEqual(response.status_code, 401)

    def test_history_list_returns_403_for_authenticated_unverified_user(self):
        self.client.force_authenticate(user=self.unverified_user)

        response = self.client.get("/history/")

        self.assertEqual(response.status_code, 403)

    def test_history_list_returns_only_owned_records_in_newest_first_order(self):
        from django.utils import timezone
        from datetime import timedelta

        older = self._create_history(
            owner=self.verified_user,
            original_name="older.pdf",
            created_at=timezone.now() - timedelta(hours=2),
        )
        newer = self._create_history(
            owner=self.verified_user,
            original_name="newer.pdf",
            created_at=timezone.now() - timedelta(hours=1),
        )
        self._create_history(
            owner=self.other_user,
            original_name="other.pdf",
            created_at=timezone.now(),
        )
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get("/history/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 2)
        self.assertEqual(response.data["limit"], 10)
        self.assertEqual(response.data["offset"], 0)
        self.assertEqual(
            [item["id"] for item in response.data["results"]],
            [str(newer.id), str(older.id)],
        )
        self.assertEqual(
            response.data["results"][0]["original_name"],
            "newer.pdf",
        )
        self.assertIn("created_at", response.data["results"][0])
        self.assertEqual(
            set(response.data["results"][0].keys()),
            {"id", "original_name", "custom_name", "status_processing", "created_at"},
        )

    def test_history_list_returns_empty_results_when_user_has_no_history(self):
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get("/history/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 0)
        self.assertEqual(response.data["results"], [])

    def test_history_list_rejects_invalid_limit(self):
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get("/history/?limit=0")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(response.data["message"], "Invalid history pagination request.")

    def test_history_list_rejects_invalid_offset(self):
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get("/history/?offset=-1")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(response.data["message"], "Invalid history pagination request.")

    def test_history_list_accepts_custom_limit_and_offset(self):
        self._create_history(
            owner=self.verified_user,
            original_name="older.pdf",
            created_at=timezone.now() - timedelta(hours=3),
        )
        middle = self._create_history(
            owner=self.verified_user,
            original_name="middle.pdf",
            created_at=timezone.now() - timedelta(hours=2),
        )
        self._create_history(
            owner=self.verified_user,
            original_name="newest.pdf",
            created_at=timezone.now() - timedelta(hours=1),
        )
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get("/history/?limit=1&offset=1")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 3)
        self.assertEqual(response.data["limit"], 1)
        self.assertEqual(response.data["offset"], 1)
        self.assertEqual(
            [item["id"] for item in response.data["results"]],
            [str(middle.id)],
        )


class HistoryDetailViewTest(BaseApiViewTest):
    def setUp(self):
        super().setUp()
        self.verified_user = User.objects.create_user(
            email="verified-history-detail@example.com",
            name="Verified History Detail",
            password="secret",
            status="verified",
        )
        self.unverified_user = User.objects.create_user(
            email="unverified-history-detail@example.com",
            name="Unverified History Detail",
            password="secret",
            status="unverified",
        )
        self.other_user = User.objects.create_user(
            email="other-history-detail@example.com",
            name="Other History Detail",
            password="secret",
            status="verified",
        )
        self.history = ArtifactHistory.objects.create(
            owner=self.verified_user,
            original_name="invoice.pdf",
            custom_name="April Invoice",
            output_json={
                "document_info": {"source_type": "PDF", "filename": "invoice.pdf"},
                "summary": {"table_count": 1},
                "content_data": [
                    {"table_name": "Sheet1", "headers": ["A"], "rows": [{"A": "1"}]}
                ],
            },
            status_processing="completed",
            created_at=timezone.now() - timedelta(minutes=2),
        )

    def test_history_detail_patch_returns_401_for_anonymous_user(self):
        response = self.client.patch(
            f"/history/{self.history.id}/rename/",
            {"custom_name": "Renamed Invoice"},
            format="json",
        )

        self.assertEqual(response.status_code, 401)

    def test_history_detail_patch_returns_403_for_unverified_user(self):
        self.client.force_authenticate(user=self.unverified_user)

        response = self.client.patch(
            f"/history/{self.history.id}/rename/",
            {"custom_name": "Renamed Invoice"},
            format="json",
        )

        self.assertEqual(response.status_code, 403)

    def test_history_detail_patch_updates_custom_name_for_owner(self):
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.patch(
            f"/history/{self.history.id}/rename/",
            {"custom_name": "Renamed Invoice"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.history.refresh_from_db()
        self.assertEqual(self.history.custom_name, "Renamed Invoice")
        self.assertEqual(response.data["id"], str(self.history.id))
        self.assertEqual(response.data["custom_name"], "Renamed Invoice")
        self.assertEqual(response.data["original_name"], "invoice.pdf")

    def test_history_detail_patch_allows_blank_custom_name(self):
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.patch(
            f"/history/{self.history.id}/rename/",
            {"custom_name": "   "},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.history.refresh_from_db()
        self.assertEqual(self.history.custom_name, "")

    def test_history_detail_patch_returns_404_for_non_owner(self):
        self.client.force_authenticate(user=self.other_user)

        response = self.client.patch(
            f"/history/{self.history.id}/rename/",
            {"custom_name": "Renamed Invoice"},
            format="json",
        )

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.data["message"], "History item not found.")

    def test_history_detail_patch_returns_400_for_missing_custom_name(self):
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.patch(
            f"/history/{self.history.id}/rename/",
            {},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("custom_name", response.data)

    def test_history_detail_patch_returns_400_for_custom_name_above_limit(self):
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.patch(
            f"/history/{self.history.id}/rename/",
            {"custom_name": "A" * (HISTORY_CUSTOM_NAME_MAX_LENGTH + 1)},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("custom_name", response.data)
        self.assertIn(
            str(HISTORY_CUSTOM_NAME_MAX_LENGTH),
            str(response.data["custom_name"][0]),
        )

    def test_history_detail_delete_returns_401_for_anonymous_user(self):
        response = self.client.delete(f"/history/{self.history.id}/delete/")

        self.assertEqual(response.status_code, 401)

    def test_history_detail_delete_returns_403_for_unverified_user(self):
        self.client.force_authenticate(user=self.unverified_user)

        response = self.client.delete(f"/history/{self.history.id}/delete/")

        self.assertEqual(response.status_code, 403)

    @patch("api.views.os.remove")
    def test_history_detail_delete_removes_record_and_cached_artifacts_for_owner(
        self,
        mock_remove_file,
    ):
        HistoryExportArtifact.objects.create(
            history=self.history,
            owner=self.verified_user,
            requested_format="csv",
            artifact_type="csv",
            file_id="csv_token",
            file_name="cached.csv",
            created_at=timezone.now() - timedelta(minutes=1),
        )
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.delete(f"/history/{self.history.id}/delete/")

        self.assertEqual(response.status_code, 204)
        self.assertFalse(ArtifactHistory.objects.filter(id=self.history.id).exists())
        self.assertFalse(
            HistoryExportArtifact.objects.filter(history_id=self.history.id).exists()
        )
        mock_remove_file.assert_called_once_with(
            safe_join(settings.CSV_EXPORT_DIR, "cached.csv")
        )

    def test_history_detail_delete_returns_404_for_non_owner(self):
        self.client.force_authenticate(user=self.other_user)

        response = self.client.delete(f"/history/{self.history.id}/delete/")

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.data["message"], "History item not found.")

    @patch("api.views.delete_artifact_history", side_effect=RuntimeError("boom"))
    @patch("api.views._delete_history_cached_artifacts")
    def test_history_detail_delete_returns_500_when_deletion_fails(
        self,
        mock_delete_cached_artifacts,
        mock_delete_history,
    ):
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.delete(f"/history/{self.history.id}/delete/")

        self.assertEqual(response.status_code, 500)
        self.assertEqual(
            response.data["message"],
            "Failed to delete history item due to internal error.",
        )
        mock_delete_cached_artifacts.assert_called_once_with(self.history)
        mock_delete_history.assert_called_once_with(self.history)

class HistoryDownloadViewTest(BaseApiViewTest):
    def setUp(self):
        super().setUp()
        self.verified_user = User.objects.create_user(
            email="verified-download@example.com",
            name="Verified Download",
            password="secret",
            status="verified",
        )
        self.unverified_user = User.objects.create_user(
            email="unverified-download@example.com",
            name="Unverified Download",
            password="secret",
            status="unverified",
        )
        self.other_user = User.objects.create_user(
            email="other-download@example.com",
            name="Other Download",
            password="secret",
            status="verified",
        )
        self.history = ArtifactHistory.objects.create(
            owner=self.verified_user,
            original_name="invoice.pdf",
            custom_name=None,
            output_json={
                "document_info": {"source_type": "PDF", "filename": "invoice.pdf"},
                "summary": {"table_count": 1},
                "content_data": [
                    {"table_name": "Sheet1", "headers": ["A"], "rows": [{"A": "1"}]}
                ],
            },
            status_processing="completed",
            created_at=timezone.now() - timedelta(minutes=1),
        )

    def test_history_download_returns_401_for_anonymous_user(self):
        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=csv"
        )

        self.assertEqual(response.status_code, 401)

    def test_history_download_returns_403_for_authenticated_unverified_user(self):
        self.client.force_authenticate(user=self.unverified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=csv"
        )

        self.assertEqual(response.status_code, 403)

    @patch("api.views.open")
    @patch("api.views.export_csv_to_filesystem")
    def test_history_download_returns_csv_file_for_owner(
        self, mock_export_csv, mock_open_file
    ):
        mock_export_csv.return_value = {
            "file_id": "csv_token",
            "file_name": "export_token.csv",
            "artifact_type": "csv",
            "size_bytes": 12,
            "created_at": "2026-04-08T10:00:00Z",
        }
        mock_open_file.return_value = BytesIO(b"col\n1\n")
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=csv"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertIn(
            'attachment; filename="export_token.csv"',
            response["Content-Disposition"],
        )
        mock_export_csv.assert_called_once()
        self.assertEqual(
            mock_export_csv.call_args.kwargs["output_json"],
            self.history.output_json,
        )

    @patch("api.views.open")
    @patch("api.views.os.remove")
    @patch("api.views.create_history_export_artifact")
    @patch("api.views.get_history_export_artifact")
    @patch("api.views.export_excel_to_filesystem")
    def test_history_download_creates_cache_record_after_excel_cache_miss(
        self,
        mock_export_excel,
        mock_get_cached_artifact,
        mock_create_cached_artifact,
        mock_remove_file,
        mock_open_file,
    ):
        mock_get_cached_artifact.return_value = None
        mock_export_excel.return_value = {
            "file_id": "xlsx_token",
            "file_name": "export_token.xlsx",
            "artifact_type": "xlsx",
            "size_bytes": 12,
            "created_at": "2026-04-08T10:00:00Z",
        }
        mock_create_cached_artifact.return_value = HistoryExportArtifact(
            history=self.history,
            owner=self.verified_user,
            requested_format="xlsx",
            artifact_type="xlsx",
            file_id="xlsx_token",
            file_name="export_token.xlsx",
            created_at=timezone.now(),
        )
        mock_open_file.return_value = BytesIO(b"xlsx")
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=xlsx"
        )

        self.assertEqual(response.status_code, 200)
        mock_get_cached_artifact.assert_called_once_with(
            history=self.history,
            owner=self.verified_user,
            requested_format="xlsx",
        )
        mock_export_excel.assert_called_once()
        mock_create_cached_artifact.assert_called_once_with(
            history=self.history,
            owner=self.verified_user,
            requested_format="xlsx",
            artifact_type="xlsx",
            file_id="xlsx_token",
            file_name="export_token.xlsx",
            created_at="2026-04-08T10:00:00Z",
        )
        mock_remove_file.assert_not_called()

    @patch("api.views.open")
    @patch("api.views.os.remove")
    @patch("api.views.create_history_export_artifact")
    @patch("api.views.get_history_export_artifact")
    @patch("api.views.export_excel_to_filesystem")
    def test_history_download_uses_race_winner_cached_artifact_and_deletes_loser_file(
        self,
        mock_export_excel,
        mock_get_cached_artifact,
        mock_create_cached_artifact,
        mock_remove_file,
        mock_open_file,
    ):
        mock_get_cached_artifact.return_value = None
        mock_export_excel.return_value = {
            "file_id": "xlsx_loser",
            "file_name": "export_loser.xlsx",
            "artifact_type": "xlsx",
            "size_bytes": 12,
            "created_at": "2026-04-08T10:00:00Z",
        }
        mock_create_cached_artifact.return_value = HistoryExportArtifact(
            history=self.history,
            owner=self.verified_user,
            requested_format="xlsx",
            artifact_type="xlsx",
            file_id="xlsx_winner",
            file_name="export_winner.xlsx",
            created_at=timezone.now(),
        )
        mock_open_file.return_value = BytesIO(b"xlsx")
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=xlsx"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'attachment; filename="export_winner.xlsx"',
            response["Content-Disposition"],
        )
        mock_open_file.assert_called_once_with(
            safe_join(settings.EXCEL_EXPORT_DIR, "export_winner.xlsx"),
            "rb",
        )
        mock_remove_file.assert_called_once_with(
            safe_join(settings.EXCEL_EXPORT_DIR, "export_loser.xlsx")
        )

    @patch("api.views.logger.warning")
    @patch("api.views.open")
    @patch("api.views.os.remove")
    @patch("api.views.create_history_export_artifact")
    @patch("api.views.get_history_export_artifact")
    @patch("api.views.export_excel_to_filesystem")
    def test_history_download_still_serves_race_winner_when_orphan_cleanup_fails(
        self,
        mock_export_excel,
        mock_get_cached_artifact,
        mock_create_cached_artifact,
        mock_remove_file,
        mock_open_file,
        mock_log_warning,
    ):
        mock_get_cached_artifact.return_value = None
        mock_export_excel.return_value = {
            "file_id": "xlsx_loser",
            "file_name": "export_loser.xlsx",
            "artifact_type": "xlsx",
            "size_bytes": 12,
            "created_at": "2026-04-08T10:00:00Z",
        }
        mock_create_cached_artifact.return_value = HistoryExportArtifact(
            history=self.history,
            owner=self.verified_user,
            requested_format="xlsx",
            artifact_type="xlsx",
            file_id="xlsx_winner",
            file_name="export_winner.xlsx",
            created_at=timezone.now(),
        )
        mock_remove_file.side_effect = OSError("cannot delete orphan")
        mock_open_file.return_value = BytesIO(b"xlsx")
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=xlsx"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'attachment; filename="export_winner.xlsx"',
            response["Content-Disposition"],
        )
        mock_log_warning.assert_called_once()

    @patch("api.views.open")
    @patch("api.views.get_history_export_artifact")
    @patch("api.views.export_csv_to_filesystem")
    def test_history_download_reuses_cached_csv_artifact_without_regenerating(
        self,
        mock_export_csv,
        mock_get_cached_artifact,
        mock_open_file,
    ):
        cached_artifact = HistoryExportArtifact(
            history=self.history,
            owner=self.verified_user,
            requested_format="csv",
            artifact_type="zip",
            file_id="csv_token",
            file_name="export_token.zip",
            created_at=timezone.now(),
        )
        mock_get_cached_artifact.return_value = cached_artifact
        mock_open_file.return_value = BytesIO(b"zip-bytes")
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=csv"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        self.assertIn(
            'attachment; filename="export_token.zip"',
            response["Content-Disposition"],
        )
        mock_get_cached_artifact.assert_called_once_with(
            history=self.history,
            owner=self.verified_user,
            requested_format="csv",
        )
        mock_export_csv.assert_not_called()

    @patch("api.views.open")
    @patch("api.views.get_history_export_artifact")
    @patch("api.views.export_excel_to_filesystem")
    def test_history_download_reuses_cached_xlsx_artifact_without_regenerating(
        self,
        mock_export_excel,
        mock_get_cached_artifact,
        mock_open_file,
    ):
        cached_artifact = HistoryExportArtifact(
            history=self.history,
            owner=self.verified_user,
            requested_format="xlsx",
            artifact_type="xlsx",
            file_id="xlsx_token",
            file_name="export_token.xlsx",
            created_at=timezone.now(),
        )
        mock_get_cached_artifact.return_value = cached_artifact
        mock_open_file.return_value = BytesIO(b"xlsx-bytes")
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=xlsx"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            'attachment; filename="export_token.xlsx"',
            response["Content-Disposition"],
        )
        mock_get_cached_artifact.assert_called_once_with(
            history=self.history,
            owner=self.verified_user,
            requested_format="xlsx",
        )
        mock_export_excel.assert_not_called()

    @patch("api.views.open")
    @patch("api.views.create_history_export_artifact")
    @patch("api.views.get_history_export_artifact")
    @patch("api.views.export_csv_to_filesystem")
    def test_history_download_regenerates_when_cached_csv_artifact_file_is_missing(
        self,
        mock_export_csv,
        mock_get_cached_artifact,
        mock_create_cached_artifact,
        mock_open_file,
    ):
        stale_cached_artifact = HistoryExportArtifact(
            history=self.history,
            owner=self.verified_user,
            requested_format="csv",
            artifact_type="csv",
            file_id="csv_old",
            file_name="stale.csv",
            created_at=timezone.now(),
        )
        mock_get_cached_artifact.return_value = stale_cached_artifact
        mock_export_csv.return_value = {
            "file_id": "csv_new",
            "file_name": "export_new.csv",
            "artifact_type": "csv",
            "size_bytes": 12,
            "created_at": "2026-04-08T10:00:00Z",
        }
        mock_create_cached_artifact.return_value = HistoryExportArtifact(
            history=self.history,
            owner=self.verified_user,
            requested_format="csv",
            artifact_type="csv",
            file_id="csv_new",
            file_name="export_new.csv",
            created_at=timezone.now(),
        )
        mock_open_file.side_effect = [
            OSError("stale file missing"),
            BytesIO(b"col\n1\n"),
        ]
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=csv"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertIn(
            'attachment; filename="export_new.csv"',
            response["Content-Disposition"],
        )
        mock_export_csv.assert_called_once()
        mock_create_cached_artifact.assert_called_once_with(
            history=self.history,
            owner=self.verified_user,
            requested_format="csv",
            artifact_type="csv",
            file_id="csv_new",
            file_name="export_new.csv",
            created_at="2026-04-08T10:00:00Z",
        )

    @patch("api.views.open")
    @patch("api.views.get_history_export_artifact")
    @patch("api.views.export_excel_to_filesystem")
    def test_history_download_returns_500_when_stale_cached_xlsx_regeneration_fails(
        self,
        mock_export_excel,
        mock_get_cached_artifact,
        mock_open_file,
    ):
        stale_cached_artifact = HistoryExportArtifact(
            history=self.history,
            owner=self.verified_user,
            requested_format="xlsx",
            artifact_type="xlsx",
            file_id="xlsx_old",
            file_name="stale.xlsx",
            created_at=timezone.now(),
        )
        mock_get_cached_artifact.return_value = stale_cached_artifact
        mock_open_file.side_effect = OSError("stale file missing")
        mock_export_excel.side_effect = OutputExcelGenerationError("failed")
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=xlsx"
        )

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(
            response.data["message"],
            "Failed to download history file due to internal error.",
        )
        mock_export_excel.assert_called_once()

    @patch("api.views.open")
    @patch("api.views.export_excel_to_filesystem")
    def test_history_download_returns_excel_file_for_owner(
        self, mock_export_excel, mock_open_file
    ):
        mock_export_excel.return_value = {
            "file_id": "xlsx_token",
            "file_name": "export_token.xlsx",
            "artifact_type": "xlsx",
            "size_bytes": 12,
            "created_at": "2026-04-08T10:00:00Z",
        }
        mock_open_file.return_value = BytesIO(b"xlsx")
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=xlsx"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            'attachment; filename="export_token.xlsx"',
            response["Content-Disposition"],
        )
        mock_export_excel.assert_called_once()
        self.assertEqual(
            mock_export_excel.call_args.kwargs["output_json"],
            self.history.output_json,
        )

    def test_history_download_returns_404_for_non_owner(self):
        self.client.force_authenticate(user=self.other_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=csv"
        )

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(response.data["message"], "History item not found.")

    def test_history_download_returns_404_for_missing_history_item(self):
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            "/history/00000000-0000-0000-0000-000000000000/download/?file_format=csv"
        )

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(response.data["message"], "History item not found.")

    def test_history_download_rejects_invalid_file_format(self):
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=pdf"
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(response.data["message"], "Invalid history download format.")

    @patch("api.views.export_csv_to_filesystem")
    def test_history_download_returns_500_when_csv_generation_fails(self, mock_export_csv):
        mock_export_csv.side_effect = OutputCSVGenerationError("failed")
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=csv"
        )

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(
            response.data["message"],
            "Failed to download history file due to internal error.",
        )

    @patch("api.views.open", side_effect=OSError("fresh file missing"))
    @patch("api.views.get_history_export_artifact")
    @patch("api.views.export_csv_to_filesystem")
    def test_history_download_returns_500_when_newly_generated_csv_file_cannot_be_opened(
        self,
        mock_export_csv,
        mock_get_cached_artifact,
        _mock_open_file,
    ):
        mock_get_cached_artifact.return_value = None
        mock_export_csv.return_value = {
            "file_id": "csv_token",
            "file_name": "export_token.csv",
            "artifact_type": "csv",
            "size_bytes": 12,
            "created_at": "2026-04-08T10:00:00Z",
        }
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=csv"
        )

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(
            response.data["message"],
            "Failed to download history file due to internal error.",
        )
        mock_export_csv.assert_called_once()

    @patch("api.views.export_csv_to_filesystem")
    def test_history_download_returns_500_for_invalid_stored_output(self, mock_export_csv):
        mock_export_csv.side_effect = OutputLLMValidationError("invalid stored output")
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=csv"
        )

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(
            response.data["message"],
            "Failed to download history file due to internal error.",
        )

    @patch("api.views.open")
    @patch("api.views.export_excel_to_filesystem")
    def test_history_download_returns_500_for_unexpected_error(
        self, mock_export_excel, mock_open_file
    ):
        mock_export_excel.return_value = {
            "file_id": "xlsx_token",
            "file_name": "export_token.xlsx",
            "artifact_type": "xlsx",
            "size_bytes": 12,
            "created_at": "2026-04-08T10:00:00Z",
        }
        mock_open_file.side_effect = RuntimeError("unexpected read failure")
        self.client.force_authenticate(user=self.verified_user)

        response = self.client.get(
            f"/history/{self.history.id}/download/?file_format=xlsx"
        )

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(
            response.data["message"],
            "Failed to download history file due to internal error.",
        )


class MembersViewTest(BaseApiViewTest):
    @classmethod
    def setUpTestData(cls):
        GroupMember.objects.create(npm="2306152260", name="Steven Setiawan")
        GroupMember.objects.create(npm="2306152172", name="Siti Shofi Nadhifa")

    def test_members_endpoint_returns_200(self):
        response = self.client.get("/members/")
        self.assertEqual(response.status_code, 200)

    def test_members_endpoint_returns_group_and_members(self):
        response = self.client.get("/members/")
        self.assertEqual(response.data["group"], "Kelompok 7")
        self.assertEqual(len(response.data["members"]), 2)
        self.assertEqual(response.data["members"][0]["npm"], "2306152172")
        self.assertEqual(response.data["members"][0]["name"], "Siti Shofi Nadhifa")

    def test_members_endpoint_rejects_post(self):
        response = self.client.post("/members/")
        self.assertEqual(response.status_code, 405)


class UploadEndpointTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.factory = APIRequestFactory()

    def _post_file(self, name, content, content_type):
        f = SimpleUploadedFile(name, content, content_type=content_type)
        return self.client.post("/upload/", {"file": f}, format="multipart")

    def generate_valid_pdf_bytes(self):
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        p.drawString(100, 750, "Hello PDF")
        p.save()
        buffer.seek(0)
        return buffer.read()

    def generate_private_pdf_bytes(self, password="secret"):
        valid_pdf_bytes = self.generate_valid_pdf_bytes()

        input_buffer = BytesIO(valid_pdf_bytes)
        output_buffer = BytesIO()

        reader = PdfReader(input_buffer)
        writer = PdfWriter()

        for page in reader.pages:
            writer.add_page(page)

        writer.encrypt(password)
        writer.write(output_buffer)

        output_buffer.seek(0)
        return output_buffer.read()

    def generate_valid_xlsx_bytes(self):
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Hello"
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def generate_valid_xls_bytes(self):
        import xlwt

        buffer = BytesIO()
        wb = xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        ws.write(0, 0, "Hello XLS")
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def generate_valid_png_bytes(self):
        buffer = BytesIO()
        img = Image.new("RGB", (20, 20), color="red")
        img.save(buffer, format="PNG")
        buffer.seek(0)
        return buffer.read()

    def test_upload_with_invalid_content_length_header_returns_no_file_error(self):
        request = self.factory.post("/upload/", data={}, format="multipart")
        request.META["CONTENT_LENGTH"] = "not-a-number"

        response = upload(request)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["message"], "No file provided")

    def test_upload_without_content_length_header_returns_no_file_error(self):
        request = self.factory.post("/upload/", data={}, format="multipart")
        request.META.pop("CONTENT_LENGTH", None)

        response = upload(request)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["message"], "No file provided")

    def test_upload_non_multipart_content_type_uses_raw_max_size_guard(self):
        request = self.factory.post(
            "/upload/",
            data="{}",
            content_type="application/json",
        )
        request.META["CONTENT_LENGTH"] = str(MAX_FILE_SIZE + 1)

        response = upload(request)

        self.assertEqual(response.status_code, 413)
        self.assertEqual(response.data["status"], "error")

    def test_upload_pdf_success(self):
        pdf_doc = self.generate_valid_pdf_bytes()
        resp = self._post_file("doc.pdf", pdf_doc, "application/pdf")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["status"], "success")
        self.assertEqual(resp.data["filename"], "doc.pdf")

    @patch("file_processing.services.upload_service.NonOCRPDFService.extract_non_ocr_pdf_to_json")
    @patch("file_processing.services.upload_service.OCRService.process_pdf")
    def test_upload_pdf_scanned_fallback(self, mock_process_pdf, mock_extract):
        mock_extract.return_value = {"content": []}
        mock_process_pdf.return_value = {"content": [{"page": 1, "text": ["Scanned Content"]}]}

        pdf_doc = self.generate_valid_pdf_bytes()
        resp = self._post_file("scanned.pdf", pdf_doc, "application/pdf")

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["extracted"]["content"][0]["text"], ["Scanned Content"])
        mock_process_pdf.assert_called_once()

    @patch("file_processing.services.upload_service.NonOCRPDFService.extract_non_ocr_pdf_to_json")
    @patch("file_processing.services.upload_service.OCRService.process_pdf_pages")
    def test_upload_pdf_mixed_pages(self, mock_process_pages, mock_extract):
        # Page 1 has text, Page 2 does not.
        mock_extract.return_value = {
            "content": [
                {"page": 1, "text": ["Native Text"]},
                {"page": 2, "text": []}
            ]
        }
        mock_process_pages.return_value = {
            "content": [{"page": 2, "text": ["OCR Text"]}]
        }

        pdf_doc = self.generate_valid_pdf_bytes()
        resp = self._post_file("mixed.pdf", pdf_doc, "application/pdf")

        self.assertEqual(resp.status_code, 200)
        content = resp.data["extracted"]["content"]
        self.assertEqual(content[0]["text"], ["Native Text"])
        self.assertEqual(content[1]["text"], ["OCR Text"])
        mock_process_pages.assert_called_once()

    @patch("file_processing.services.upload_service.NonOCRPDFService.extract_non_ocr_pdf_to_json")
    @patch("file_processing.services.upload_service.OCRService.process_pdf")
    def test_upload_pdf_exception_fallback(self, mock_process_pdf, mock_extract):
        mock_extract.side_effect = Exception("Native extract failed")
        mock_process_pdf.return_value = {
            "content": [{"page": 1, "text": ["Fallback OCR Content"]}]
        }

        pdf_doc = self.generate_valid_pdf_bytes()
        resp = self._post_file("fail.pdf", pdf_doc, "application/pdf")

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["extracted"]["content"][0]["text"], ["Fallback OCR Content"])
        mock_process_pdf.assert_called_once()

    @patch("file_processing.services.upload_service.NonOCRPDFService.extract_non_ocr_pdf_to_json")
    @patch("file_processing.services.upload_service.OCRService.process_pdf")
    def test_upload_pdf_empty_extracted_data(self, mock_process_pdf, mock_extract):
        mock_extract.return_value = {}  # Missing 'content' key, hitting line 53
        mock_process_pdf.return_value = {
            "content": [{"page": 1, "text": ["OCR Triggered"]}]
        }

        pdf_doc = self.generate_valid_pdf_bytes()
        resp = self._post_file("empty_struct.pdf", pdf_doc, "application/pdf")

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["extracted"]["content"][0]["text"], ["OCR Triggered"])
        mock_process_pdf.assert_called_once()

    @patch("api.views.process_upload")
    def test_upload_xls_success(self, mock_process):
        mock_process.return_value = (True, None, "/tmp/file.xls", None)

        resp = self._post_file(
            "sheet.xls",
            b"dummy",
            "application/vnd.ms-excel",
        )

        self.assertEqual(resp.status_code, 200)

    def test_upload_xlsx_success(self):
        xlsx_content = self.generate_valid_xlsx_bytes()

        resp = self._post_file(
            "sheet.xlsx",
            xlsx_content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.assertEqual(resp.status_code, 200)

    @patch("file_processing.services.upload_service._process_image")
    def test_upload_png_success_with_extracted_payload(self, mock_process_image):
        mock_process_image.return_value = (
            True,
            None,
            {"content": [{"page": 1, "text": ["OCR image text"]}]},
        )

        png_content = self.generate_valid_png_bytes()

        resp = self._post_file(
            "photo.png",
            png_content,
            "image/png",
        )

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["status"], "success")
        self.assertEqual(resp.data["filename"], "photo.png")
        self.assertIn("extracted", resp.data)
        self.assertEqual(resp.data["extracted"]["content"][0]["text"], ["OCR image text"])

    def test_upload_unsupported_type(self):
        resp = self._post_file("note.html", b"<html></html>", "text/html")
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.data["status"], "error")
        self.assertIn("message", resp.data)

    def test_upload_no_file(self):
        resp = self.client.post("/upload/", {}, format="multipart")
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.data["status"], "error")
        self.assertIn("message", resp.data)

    def test_upload_response_does_not_expose_path(self):
        pdf_doc = self.generate_valid_pdf_bytes()

        resp = self._post_file("doc.pdf", pdf_doc, "application/pdf")

        self.assertEqual(resp.status_code, 200)
        self.assertNotIn("path", resp.data)

    def test_upload_internal_server_error(self):
        pdf_doc = self.generate_valid_pdf_bytes()

        with patch("api.views.process_upload") as mock_process_upload:
            mock_process_upload.side_effect = Exception("Unexpected failure")

            resp = self._post_file("doc.pdf", pdf_doc, "application/pdf")

            self.assertEqual(resp.status_code, 500)
            self.assertEqual(resp.data["status"], "error")
            self.assertIn("message", resp.data)

    def test_invalid_file_path_detection(self):
        pdf_doc = self.generate_valid_pdf_bytes()

        with patch(
            "file_processing.services.upload_service.os.path.abspath"
        ) as mock_abspath:

            def fake_abspath(path):
                if "doc.pdf" in path:
                    return "/evil/path/file.pdf"
                return "/safe/base"

            mock_abspath.side_effect = fake_abspath

            with self.assertRaises(ValueError):
                from file_processing.services.upload_service import save_temp_file

                f = SimpleUploadedFile(
                    "doc.pdf", pdf_doc, content_type="application/pdf"
                )
                save_temp_file(f)

    def test_save_temp_file_success(self):
        from file_processing.services.upload_service import save_temp_file

        pdf_doc = self.generate_valid_pdf_bytes()

        f = SimpleUploadedFile(
            "doc.pdf",
            pdf_doc,
            content_type="application/pdf",
        )

        path = save_temp_file(f)

        self.assertTrue(path.endswith(".pdf"))

    def test_upload_pdf_uppercase_extension(self):
        pdf_doc = self.generate_valid_pdf_bytes()

        resp = self._post_file(
            "DOC.PDF",
            pdf_doc,
            "application/pdf",
        )

        self.assertEqual(resp.status_code, 200)

    def test_file_header_not_pdf_with_extension_pdf(self):
        resp = self._post_file("doc.pdf", b"data", "application/pdf")
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.data["status"], "error")
        self.assertIn("message", resp.data)

    def test_file_is_corrupt_pdf(self):
        pdf_doc = self.generate_valid_pdf_bytes()
        corrupt_pdf = pdf_doc[:20]

        resp = self._post_file("corrupt.pdf", corrupt_pdf, "application/pdf")
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.data["status"], "error")
        self.assertIn("message", resp.data)
        self.assertIn("corrupt", resp.data["message"].lower())

    def test_upload_file_too_large(self):
        big_content = b"a" * (11 * 1024 * 1024)
        resp = self._post_file("big.pdf", big_content, "application/pdf")

        self.assertEqual(resp.status_code, 413)
        self.assertEqual(resp.data["status"], "error")
        self.assertIn("message", resp.data)
        self.assertIn("10mb", resp.data["message"].lower())

    @patch("api.views.process_upload")
    def test_upload_content_length_too_large_short_circuits_processing(self, mock_process):
        pdf_doc = self.generate_valid_pdf_bytes()
        file_obj = SimpleUploadedFile(
            "doc.pdf",
            pdf_doc,
            content_type="application/pdf",
        )

        response = self.client.post(
            "/upload/",
            {"file": file_obj},
            format="multipart",
            CONTENT_LENGTH=str(11 * 1024 * 1024),
        )

        self.assertEqual(response.status_code, 413)
        self.assertEqual(response.data["status"], "error")
        self.assertIn("10mb", response.data["message"].lower())
        mock_process.assert_not_called()

    def test_upload_file_exact_10mb_allowed(self):
        valid_pdf = self.generate_valid_pdf_bytes()
        remaining_size = (10 * 1024 * 1024) - len(valid_pdf)

        padding = b"\0" * remaining_size
        exact_content = valid_pdf + padding
        resp = self._post_file("exact.pdf", exact_content, "application/pdf")
        self.assertEqual(resp.status_code, 200)

    def test_upload_file_less_than_10mb_allowed(self):
        valid_pdf = self.generate_valid_pdf_bytes()
        padding = b"\0" * (5 * 1024 * 1024)
        less_content = valid_pdf + padding
        resp = self._post_file("small.pdf", less_content, "application/pdf")
        self.assertEqual(resp.status_code, 200)

    def test_file_is_private_pdf(self):
        private_pdf = self.generate_private_pdf_bytes(password="1234")

        resp = self._post_file("private.pdf", private_pdf, "application/pdf")

        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.data["status"], "error")
        self.assertIn("message", resp.data)
        self.assertIn("password-protected", resp.data["message"].lower())
        self.assertNotIn("corrupt", resp.data["message"].lower())

    def test_xls_extension_but_invalid_mime(self):
        resp = self._post_file(
            "fake.xls",
            b"not an excel file",
            "application/vnd.ms-excel",
        )

        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.data["status"], "error")

    def test_xlsx_extension_but_invalid_mime(self):
        resp = self._post_file(
            "fake.xlsx",
            b"this is not an excel file",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.data["status"], "error")
        self.assertEqual(resp.data["message"], "File content does not match its extension.")

    @patch("file_processing.services.upload_service.magic.from_buffer")
    def test_xlsx_octet_stream_without_zip_signature_returns_extension_mismatch(
        self, mock_magic
    ):
        mock_magic.return_value = "application/octet-stream"

        resp = self._post_file(
            "fake.xlsx",
            b"plain text masquerading as xlsx",
            "application/octet-stream",
        )

        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.data["status"], "error")
        self.assertEqual(resp.data["message"], "File content does not match its extension.")

    @patch("file_processing.services.upload_service._is_legacy_xls_content")
    @patch("file_processing.services.upload_service.magic.from_buffer")
    def test_password_protected_xlsx_returns_specific_error(
        self,
        mock_magic,
        mock_is_legacy_xls,
    ):
        mock_magic.return_value = "application/vnd.ms-excel"
        mock_is_legacy_xls.return_value = False
        ole_payload = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1" + b"X" * 2048

        resp = self._post_file(
            "protected.xlsx",
            ole_payload,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.data["status"], "error")
        self.assertIn("password-protected", resp.data["message"].lower())
        self.assertIn("excel", resp.data["message"].lower())

    @patch("file_processing.services.upload_service._is_legacy_xls_content")
    @patch("file_processing.services.upload_service.magic.from_buffer")
    def test_legacy_xls_renamed_to_xlsx_is_accepted(
        self,
        mock_magic,
        mock_is_legacy_xls,
    ):
        mock_magic.return_value = "application/vnd.ms-excel"
        mock_is_legacy_xls.return_value = True
        xls_content = self.generate_valid_xls_bytes()

        resp = self._post_file(
            "renamed.xlsx",
            xls_content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["status"], "success")

    def test_mime_detection_exception(self):
        with patch(
            "file_processing.services.upload_service.magic.from_buffer"
        ) as mock_magic:
            mock_magic.side_effect = Exception("libmagic failure")

            resp = self._post_file(
                "sheet.xlsx",
                b"dummy content",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            self.assertEqual(resp.status_code, 400)
            self.assertEqual(resp.data["status"], "error")
            self.assertEqual(resp.data["message"], "Unable to determine file type.")

    def test_upload_pdf_too_many_pages(self):
        pdf_doc = self.generate_valid_pdf_bytes()

        with patch(
            "file_processing.services.upload_service.PdfReader"
        ) as mock_reader_cls:
            mock_instance = MagicMock()
            mock_instance.is_encrypted = False
            mock_instance.pages = [MagicMock()] * 101
            mock_reader_cls.return_value = mock_instance

            resp = self._post_file("doc.pdf", pdf_doc, "application/pdf")
            self.assertEqual(resp.status_code, 400)
            self.assertEqual(resp.data["status"], "error")
            self.assertIn("100", resp.data["message"])

    @patch("api.views.process_upload")
    def test_upload_returns_extracted_text(self, mock_process):
        mock_process.return_value = (
            True,
            None,
            "/tmp/file.pdf",
            {"content": [{"page": 1, "text": ["hello"]}]},
        )

        resp = self.client.post(
            "/upload/",
            {
                "file": SimpleUploadedFile(
                    "doc.pdf", b"%PDF-1.4", content_type="application/pdf"
                )
            },
            format="multipart",
        )

        self.assertEqual(resp.status_code, 200)
        self.assertIn("extracted", resp.data)

    @patch("file_processing.services.upload_service.OCRService.process_pdf")
    @patch("file_processing.services.upload_service.NonOCRPDFService.extract_non_ocr_pdf_to_json")
    def test_ocr_failure_returns_internal_server_error(
        self, mock_non_ocr, mock_ocr
    ):

        mock_non_ocr.return_value = {
            "content": [{"page": 1, "text": []}]
        }

        mock_ocr.side_effect = Exception("OCR crash")

        pdf_doc = self.generate_valid_pdf_bytes()

        resp = self._post_file(
            "doc.pdf",
            pdf_doc,
            "application/pdf",
        )

        self.assertEqual(resp.status_code, 500)
        self.assertEqual(resp.data["status"], "error")

    @patch("file_processing.services.upload_service.os.remove")
    def test_cleanup_failure_logged(self, mock_remove):
        mock_remove.side_effect = Exception("delete failed")

        pdf_doc = self.generate_valid_pdf_bytes()

        resp = self._post_file(
            "doc.pdf",
            pdf_doc,
            "application/pdf",
        )

        self.assertEqual(resp.status_code, 200)

    @patch("file_processing.services.upload_service.os.path.exists")
    def test_cleanup_when_temp_file_not_exists(self, mock_exists):
        mock_exists.return_value = False

        pdf_doc = self.generate_valid_pdf_bytes()

        resp = self._post_file(
            "doc.pdf",
            pdf_doc,
            "application/pdf",
        )

        self.assertEqual(resp.status_code, 200)

    @patch("file_processing.services.upload_service.validate_mime_type")
    def test_upload_endpoint_mime_validation_failure(self, mock_validate):
        mock_validate.return_value = (False, "Invalid MIME")

        resp = self._post_file(
            "file.pdf",
            b"%PDF-1.4",
            "application/pdf",
        )

        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.data["status"], "error")

class ExportCSVViewTest(APISimpleTestCase):
    def _valid_output_json(self):
        return {
            "document_info": {
                "source_type": "Excel",
                "filename": "laporan_tahunan.xlsx",
            },
            "summary": {
                "grand_total": 1500000,
                "period": "2026",
            },
            "content_data": [
                {
                    "table_name": "Sheet1_Januari",
                    "headers": ["item_name", "quantity", "price"],
                    "rows": [
                        {"item_name": "Kertas", "quantity": 10, "price": 50000},
                    ],
                }
            ],
        }

    def _verified_user(self):
        return SimpleNamespace(
            id="verified-user-id",
            email="verified@example.com",
            is_authenticated=True,
            status="verified",
        )

    def _unverified_user(self):
        return SimpleNamespace(
            id="unverified-user-id",
            email="unverified@example.com",
            is_authenticated=True,
            status="unverified",
        )

    def test_export_csv_endpoint_rejects_get_method(self):
        response = self.client.get("/export/csv")
        self.assertEqual(response.status_code, 405)

    def test_export_csv_endpoint_returns_400_if_output_json_missing(self):
        self.client.force_authenticate(user=self._verified_user())
        response = self.client.post("/export/csv", data={}, format="json")

        self.assertEqual(response.status_code, 400)
        self.assertIn("output_json", response.data)

    @patch("api.views.export_csv_to_filesystem")
    def test_export_csv_endpoint_returns_401_for_unauthenticated_user(
        self,
        mocked_export,
    ):
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/csv", data=payload, format="json")

        self.assertEqual(response.status_code, 401)
        mocked_export.assert_not_called()

    @patch("api.views.export_csv_to_filesystem")
    def test_export_csv_endpoint_returns_403_for_authenticated_unverified_user(
        self,
        mocked_export,
    ):
        self.client.force_authenticate(user=self._unverified_user())
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/csv", data=payload, format="json")

        self.assertEqual(response.status_code, 403)
        mocked_export.assert_not_called()

    @patch("api.views.export_csv_to_filesystem")
    def test_export_csv_endpoint_returns_200_with_metadata(self, mocked_export):
        self.client.force_authenticate(user=self._verified_user())
        mocked_export.return_value = {
            "file_id": "csv_abc123",
            "file_name": "export_abc123.csv",
            "artifact_type": "csv",
            "size_bytes": 128,
            "created_at": "2026-03-07T10:00:00Z",
        }
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/csv", data=payload, format="json")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["file_id"], "csv_abc123")
        self.assertEqual(response.data["file_name"], "export_abc123.csv")
        mocked_export.assert_called_once()

    @patch("api.views.export_csv_to_filesystem")
    def test_export_csv_endpoint_returns_400_when_service_validation_fails(
        self,
        mocked_export,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_export.side_effect = OutputLLMValidationError("invalid schema")
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/csv", data=payload, format="json")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(response.data["message"], "Invalid CSV export request.")

    @patch("api.views.export_csv_to_filesystem")
    def test_export_csv_endpoint_returns_500_on_internal_error(self, mocked_export):
        self.client.force_authenticate(user=self._verified_user())
        mocked_export.side_effect = RuntimeError("disk full")
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/csv", data=payload, format="json")

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertIn("Failed to generate CSV", response.data["message"])

    @patch("api.views.export_csv_to_filesystem")
    def test_export_csv_endpoint_returns_500_on_generation_error(self, mocked_export):
        self.client.force_authenticate(user=self._verified_user())
        mocked_export.side_effect = OutputCSVGenerationError("storage failure")
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/csv", data=payload, format="json")

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(
            response.data["message"],
            "Failed to generate CSV due to internal error.",
        )

    @patch("api.views.export_csv_to_filesystem")
    def test_export_csv_endpoint_returns_500_when_response_metadata_invalid(
        self,
        mocked_export,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_export.return_value = {
            "file_id": "csv_abc123",
            "file_name": "../unsafe.csv",
            "artifact_type": "csv",
            "size_bytes": 128,
            "created_at": "2026-03-07T10:00:00Z",
        }
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/csv", data=payload, format="json")

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertIn("invalid response metadata", response.data["message"])


class ExportExcelViewTest(APISimpleTestCase):
    def _valid_output_json(self):
        return {
            "document_info": {
                "source_type": "Excel",
                "filename": "laporan_tahunan.xlsx",
            },
            "summary": {
                "grand_total": 1500000,
                "period": "2026",
            },
            "content_data": [
                {
                    "table_name": "Sheet1_Januari",
                    "headers": ["item_name", "quantity", "price"],
                    "rows": [
                        {"item_name": "Kertas", "quantity": 10, "price": 50000},
                    ],
                }
            ],
        }

    def _verified_user(self):
        return SimpleNamespace(
            id="verified-user-id",
            email="verified@example.com",
            is_authenticated=True,
            status="verified",
        )

    def _unverified_user(self):
        return SimpleNamespace(
            id="unverified-user-id",
            email="unverified@example.com",
            is_authenticated=True,
            status="unverified",
        )

    def test_export_excel_endpoint_rejects_get_method(self):
        response = self.client.get("/export/excel")
        self.assertEqual(response.status_code, 405)

    def test_export_excel_endpoint_returns_400_if_output_json_missing(self):
        self.client.force_authenticate(user=self._verified_user())
        response = self.client.post("/export/excel", data={}, format="json")

        self.assertEqual(response.status_code, 400)
        self.assertIn("output_json", response.data)

    @patch("api.views.export_excel_to_filesystem")
    def test_export_excel_endpoint_returns_401_for_unauthenticated_user(
        self,
        mocked_export,
    ):
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/excel", data=payload, format="json")

        self.assertEqual(response.status_code, 401)
        mocked_export.assert_not_called()

    @patch("api.views.export_excel_to_filesystem")
    def test_export_excel_endpoint_returns_403_for_authenticated_unverified_user(
        self,
        mocked_export,
    ):
        self.client.force_authenticate(user=self._unverified_user())
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/excel", data=payload, format="json")

        self.assertEqual(response.status_code, 403)
        mocked_export.assert_not_called()

    @patch("api.views.export_excel_to_filesystem")
    def test_export_excel_endpoint_returns_200_with_metadata(self, mocked_export):
        self.client.force_authenticate(user=self._verified_user())
        mocked_export.return_value = {
            "file_id": "xlsx_abc123",
            "file_name": "export_abc123.xlsx",
            "artifact_type": "xlsx",
            "size_bytes": 512,
            "created_at": "2026-03-29T10:00:00Z",
        }
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/excel", data=payload, format="json")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["file_id"], "xlsx_abc123")
        self.assertEqual(response.data["file_name"], "export_abc123.xlsx")
        self.assertEqual(response.data["artifact_type"], "xlsx")
        mocked_export.assert_called_once()

    @patch("api.views.export_excel_to_filesystem")
    def test_export_excel_endpoint_returns_400_when_service_validation_fails(
        self,
        mocked_export,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_export.side_effect = OutputLLMValidationError("invalid schema")
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/excel", data=payload, format="json")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(response.data["message"], "Invalid Excel export request.")

    @patch("api.views.export_excel_to_filesystem")
    def test_export_excel_endpoint_returns_400_when_service_mapping_fails(
        self,
        mocked_export,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_export.side_effect = OutputCSVMappingError("invalid mapping")
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/excel", data=payload, format="json")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(response.data["message"], "Invalid Excel export request.")

    @patch("api.views.export_excel_to_filesystem")
    def test_export_excel_endpoint_returns_500_on_generation_error(self, mocked_export):
        self.client.force_authenticate(user=self._verified_user())
        mocked_export.side_effect = OutputExcelGenerationError("storage failure")
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/excel", data=payload, format="json")

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(
            response.data["message"],
            "Failed to generate Excel due to internal error.",
        )

    @patch("api.views.export_excel_to_filesystem")
    def test_export_excel_endpoint_returns_500_on_unexpected_error(self, mocked_export):
        self.client.force_authenticate(user=self._verified_user())
        mocked_export.side_effect = RuntimeError("disk full")
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/excel", data=payload, format="json")

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(
            response.data["message"],
            "Failed to generate Excel due to internal error.",
        )

    @patch("api.views.export_excel_to_filesystem")
    def test_export_excel_endpoint_returns_500_when_response_metadata_invalid(
        self,
        mocked_export,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_export.return_value = {
            "file_id": "xlsx_abc123",
            "file_name": "../unsafe.xlsx",
            "artifact_type": "xlsx",
            "size_bytes": 512,
            "created_at": "2026-03-29T10:00:00Z",
        }
        payload = {"output_json": self._valid_output_json()}

        response = self.client.post("/export/excel", data=payload, format="json")

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertIn("invalid response metadata", response.data["message"])


class DownloadCSVViewTest(APISimpleTestCase):
    def _response_data(self, response):
        if hasattr(response, "data"):
            return response.data
        return {}

    def _verified_user(self):
        return SimpleNamespace(
            id="verified-user-id",
            email="verified@example.com",
            is_authenticated=True,
            status="verified",
        )

    def _unverified_user(self):
        return SimpleNamespace(
            id="unverified-user-id",
            email="unverified@example.com",
            is_authenticated=True,
            status="unverified",
        )

    @patch("api.views.resolve_csv_download_artifact", create=True)
    def test_download_csv_endpoint_returns_401_for_unauthenticated_user(
        self,
        mocked_resolver,
    ):
        response = self.client.get("/export/csv/csv_abc123/download")

        self.assertEqual(response.status_code, 401)
        mocked_resolver.assert_not_called()

    @patch("api.views.resolve_csv_download_artifact", create=True)
    def test_download_csv_endpoint_returns_403_for_authenticated_unverified_user(
        self,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._unverified_user())

        response = self.client.get("/export/csv/csv_abc123/download")

        self.assertEqual(response.status_code, 403)
        mocked_resolver.assert_not_called()

    @patch("api.views.resolve_csv_download_artifact", create=True)
    @patch("api.views.open", create=True)
    def test_download_csv_endpoint_returns_200_with_attachment_headers(
        self,
        mocked_open,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.return_value = {
            "file_name": "export_abc123.csv",
            "file_path": "/safe/storage/export_abc123.csv",
            "artifact_type": "csv",
            "content_type": "text/csv",
        }
        mocked_open.return_value.__enter__.return_value = b"name,age\r\nZufar,21\r\n"

        response = self.client.get("/export/csv/csv_abc123/download")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertIn(
            'attachment; filename="export_abc123.csv"',
            response["Content-Disposition"],
        )

    @patch("api.views.resolve_csv_download_artifact", create=True)
    @patch("api.views.open", create=True)
    def test_download_csv_endpoint_uses_custom_filename_from_query(
        self,
        mocked_open,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.return_value = {
            "file_name": "export_abc123.csv",
            "file_path": "/safe/storage/export_abc123.csv",
            "artifact_type": "csv",
            "content_type": "text/csv",
        }
        mocked_open.return_value.__enter__.return_value = b"name,age\r\nZufar,21\r\n"

        response = self.client.get(
            "/export/csv/csv_abc123/download?filename=laporan_tahunan"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'attachment; filename="laporan_tahunan.csv"',
            response["Content-Disposition"],
        )

    @patch("api.views.resolve_csv_download_artifact", create=True)
    @patch("api.views.open", create=True)
    def test_download_csv_endpoint_uses_custom_filename_with_zip_artifact(
        self,
        mocked_open,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.return_value = {
            "file_name": "export_abc123.zip",
            "file_path": "/safe/storage/export_abc123.zip",
            "artifact_type": "zip",
            "content_type": "application/zip",
        }
        mocked_open.return_value.__enter__.return_value = b"fake zip content"

        response = self.client.get(
            "/export/csv/zip_abc123/download?filename=arsip_laporan"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'attachment; filename="arsip_laporan.zip"',
            response["Content-Disposition"],
        )

    @patch("api.views.resolve_csv_download_artifact", create=True)
    @patch("api.views.open", create=True)
    def test_download_csv_endpoint_falls_back_to_default_filename_when_query_is_unsafe(
        self,
        mocked_open,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.return_value = {
            "file_name": "export_abc123.csv",
            "file_path": "/safe/storage/export_abc123.csv",
            "artifact_type": "csv",
            "content_type": "text/csv",
        }
        mocked_open.return_value.__enter__.return_value = b"name,age\r\nZufar,21\r\n"

        response = self.client.get(
            "/export/csv/csv_abc123/download?filename=..%2Fevil.csv"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'attachment; filename="export_abc123.csv"',
            response["Content-Disposition"],
        )

    @patch("api.views.resolve_csv_download_artifact", create=True)
    def test_download_csv_endpoint_returns_404_for_invalid_file_id(
        self,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.side_effect = OutputCSVDownloadLookupError("invalid file id")

        response = self.client.get("/export/csv/csv_bad-token/download")
        response_data = self._response_data(response)

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response_data.get("status"), "error")
        self.assertEqual(response_data.get("message"), "CSV file not found.")

    @patch("api.views.resolve_csv_download_artifact", create=True)
    def test_download_csv_endpoint_returns_404_for_missing_file(
        self,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.side_effect = OutputCSVDownloadLookupError("missing file")

        response = self.client.get("/export/csv/csv_deadbeef/download")
        response_data = self._response_data(response)

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response_data.get("status"), "error")
        self.assertEqual(response_data.get("message"), "CSV file not found.")

    @patch("api.views.resolve_csv_download_artifact", create=True)
    @patch("api.views.open", create=True)
    def test_download_csv_endpoint_returns_404_for_unsafe_artifact_filename(
        self,
        mocked_open,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.return_value = {
            "file_name": "../evil.csv",
            "artifact_type": "csv",
            "content_type": "text/csv",
        }

        response = self.client.get("/export/csv/csv_abc123/download")
        response_data = self._response_data(response)

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response_data.get("status"), "error")
        self.assertEqual(response_data.get("message"), "CSV file not found.")
        mocked_open.assert_not_called()

    @patch("api.views.resolve_csv_download_artifact", create=True)
    @patch("api.views.open", side_effect=OSError("disk read failed"), create=True)
    def test_download_csv_endpoint_returns_500_when_reading_file_fails(
        self,
        _mocked_open,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.return_value = {
            "file_name": "export_abc123.csv",
            "file_path": "/safe/storage/export_abc123.csv",
            "artifact_type": "csv",
            "content_type": "text/csv",
        }

        response = self.client.get("/export/csv/csv_abc123/download")

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(
            response.data["message"],
            "Failed to download CSV due to internal error.",
        )

    @patch("api.views.resolve_csv_download_artifact", create=True)
    @patch("api.views.open", side_effect=RuntimeError("unexpected read failure"), create=True)
    def test_download_csv_endpoint_returns_500_for_unexpected_error_when_opening_file(
        self,
        _mocked_open,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.return_value = {
            "file_name": "export_abc123.csv",
            "file_path": "/safe/storage/export_abc123.csv",
            "artifact_type": "csv",
            "content_type": "text/csv",
        }

        response = self.client.get("/export/csv/csv_abc123/download")

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(
            response.data["message"],
            "Failed to download CSV due to internal error.",
        )

    def test_download_csv_endpoint_rejects_post_method(self):
        response = self.client.post("/export/csv/csv_abc123/download")

        self.assertEqual(response.status_code, 405)


class DownloadExcelViewTest(APISimpleTestCase):
    def _response_data(self, response):
        if hasattr(response, "data"):
            return response.data
        return {}

    def _verified_user(self):
        return SimpleNamespace(
            id="verified-user-id",
            email="verified@example.com",
            is_authenticated=True,
            status="verified",
        )

    def _unverified_user(self):
        return SimpleNamespace(
            id="unverified-user-id",
            email="unverified@example.com",
            is_authenticated=True,
            status="unverified",
        )

    @patch("api.views.resolve_excel_download_artifact", create=True)
    def test_download_excel_endpoint_returns_401_for_unauthenticated_user(
        self,
        mocked_resolver,
    ):
        response = self.client.get("/export/excel/xlsx_abc123/download")

        self.assertEqual(response.status_code, 401)
        mocked_resolver.assert_not_called()

    @patch("api.views.resolve_excel_download_artifact", create=True)
    def test_download_excel_endpoint_returns_403_for_authenticated_unverified_user(
        self,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._unverified_user())

        response = self.client.get("/export/excel/xlsx_abc123/download")

        self.assertEqual(response.status_code, 403)
        mocked_resolver.assert_not_called()

    @patch("api.views.resolve_excel_download_artifact", create=True)
    @patch("api.views.open", create=True)
    def test_download_excel_endpoint_returns_200_with_attachment_headers(
        self,
        mocked_open,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.return_value = {
            "file_name": "export_abc123.xlsx",
            "file_path": "/safe/storage/export_abc123.xlsx",
            "artifact_type": "xlsx",
            "content_type": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        }
        mocked_open.return_value = BytesIO(b"fake xlsx bytes")

        response = self.client.get("/export/excel/xlsx_abc123/download")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            'attachment; filename="export_abc123.xlsx"',
            response["Content-Disposition"],
        )
        mocked_resolver.assert_called_once()

    @patch("api.views.resolve_excel_download_artifact", create=True)
    @patch("api.views.open", create=True)
    def test_download_excel_endpoint_uses_custom_filename_from_query(
        self,
        mocked_open,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.return_value = {
            "file_name": "export_abc123.xlsx",
            "file_path": "/safe/storage/export_abc123.xlsx",
            "artifact_type": "xlsx",
            "content_type": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        }
        mocked_open.return_value = BytesIO(b"fake xlsx bytes")

        response = self.client.get(
            "/export/excel/xlsx_abc123/download?filename=laporan_tahunan"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'attachment; filename="laporan_tahunan.xlsx"',
            response["Content-Disposition"],
        )

    @patch("api.views.resolve_excel_download_artifact", create=True)
    @patch("api.views.open", create=True)
    def test_download_excel_endpoint_falls_back_to_default_filename_when_query_is_unsafe(
        self,
        mocked_open,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.return_value = {
            "file_name": "export_abc123.xlsx",
            "file_path": "/safe/storage/export_abc123.xlsx",
            "artifact_type": "xlsx",
            "content_type": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        }
        mocked_open.return_value = BytesIO(b"fake xlsx bytes")

        response = self.client.get(
            "/export/excel/xlsx_abc123/download?filename=..%2Fevil.xlsx"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'attachment; filename="export_abc123.xlsx"',
            response["Content-Disposition"],
        )

    @patch("api.views.resolve_excel_download_artifact", create=True)
    def test_download_excel_endpoint_returns_400_for_invalid_export_id(
        self,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.side_effect = OutputExcelDownloadLookupError(
            "export_id format is invalid."
        )

        response = self.client.get("/export/excel/xlsx_bad-token/download")
        response_data = self._response_data(response)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response_data.get("status"), "error")
        self.assertEqual(response_data.get("message"), "Invalid Excel export id.")
        mocked_resolver.assert_called_once()

    @patch("api.views.resolve_excel_download_artifact", create=True)
    def test_download_excel_endpoint_returns_404_for_missing_file(
        self,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.side_effect = OutputExcelDownloadLookupError(
            "Excel artifact not found for given export_id."
        )

        response = self.client.get("/export/excel/xlsx_abc123/download")
        response_data = self._response_data(response)

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response_data.get("status"), "error")
        self.assertEqual(response_data.get("message"), "Excel file not found.")
        mocked_resolver.assert_called_once()

    @patch("api.views.safe_join", side_effect=SuspiciousFileOperation("path traversal attempt"), create=True)
    @patch("api.views.resolve_excel_download_artifact", create=True)
    def test_download_excel_endpoint_returns_404_for_unsafe_artifact_filename(
        self,
        mocked_resolver,
        _mocked_safe_join,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.return_value = {
            "file_name": "../evil.xlsx",
            "file_path": "/safe/storage/../evil.xlsx",
            "artifact_type": "xlsx",
            "content_type": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        }

        response = self.client.get("/export/excel/xlsx_abc123/download")
        response_data = self._response_data(response)

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response_data.get("status"), "error")
        self.assertEqual(response_data.get("message"), "Excel file not found.")

    @patch("api.views.resolve_excel_download_artifact", create=True)
    @patch("api.views.open", side_effect=OSError("disk read failed"), create=True)
    def test_download_excel_endpoint_returns_500_when_reading_file_fails(
        self,
        _mocked_open,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.return_value = {
            "file_name": "export_abc123.xlsx",
            "file_path": "/safe/storage/export_abc123.xlsx",
            "artifact_type": "xlsx",
            "content_type": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        }

        response = self.client.get("/export/excel/xlsx_abc123/download")

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(
            response.data["message"],
            "Failed to download Excel due to internal error.",
        )

    @patch("api.views.resolve_excel_download_artifact", create=True)
    @patch("api.views.open", side_effect=RuntimeError("unexpected read failure"), create=True)
    def test_download_excel_endpoint_returns_500_for_unexpected_error_when_opening_file(
        self,
        _mocked_open,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.return_value = {
            "file_name": "export_abc123.xlsx",
            "file_path": "/safe/storage/export_abc123.xlsx",
            "artifact_type": "xlsx",
            "content_type": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        }

        response = self.client.get("/export/excel/xlsx_abc123/download")

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(
            response.data["message"],
            "Failed to download Excel due to internal error.",
        )

    @patch("api.views.resolve_excel_download_artifact", create=True)
    def test_download_excel_endpoint_returns_500_for_unexpected_service_error(
        self,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.side_effect = RuntimeError("unexpected resolver failure")

        response = self.client.get("/export/excel/xlsx_abc123/download")

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.data["status"], "error")
        self.assertEqual(
            response.data["message"],
            "Failed to download Excel due to internal error.",
        )

    @patch("api.views.resolve_excel_download_artifact", create=True)
    def test_download_excel_endpoint_returns_500_when_storage_is_unavailable(
        self,
        mocked_resolver,
    ):
        self.client.force_authenticate(user=self._verified_user())
        mocked_resolver.side_effect = OutputExcelDownloadStorageError(
            "Excel artifact storage is unavailable."
        )

        response = self.client.get("/export/excel/xlsx_abc123/download")
        response_data = self._response_data(response)

        self.assertEqual(response_data.get("status"), "error")
        self.assertEqual(
            response_data.get("message"),
            "Failed to download Excel due to internal error.",
        )
        self.assertEqual(response.status_code, 500)

    def test_download_excel_endpoint_rejects_post_method(self):
        response = self.client.post("/export/excel/xlsx_abc123/download")

        self.assertEqual(response.status_code, 405)


class DownloadFilenameHelperTest(APISimpleTestCase):
    def test_sanitize_download_filename_rejects_empty_after_trim(self):
        self.assertIsNone(_sanitize_download_filename("   \r\n   "))

    def test_sanitize_download_filename_rejects_null_byte(self):
        self.assertIsNone(_sanitize_download_filename("report\x00.csv"))

    def test_sanitize_download_filename_rejects_dot_and_dotdot(self):
        self.assertIsNone(_sanitize_download_filename("."))
        self.assertIsNone(_sanitize_download_filename(".."))

    def test_resolve_download_filename_rewrites_wrong_extension(self):
        resolved = _resolve_download_filename(
            requested_name="laporan.txt",
            default_name="export_abc123.csv",
            artifact_type="csv",
        )
        self.assertEqual(resolved, "laporan.csv")

    def test_resolve_download_filename_keeps_matching_extension_case_insensitive(self):
        resolved = _resolve_download_filename(
            requested_name="laporan.CSV",
            default_name="export_abc123.csv",
            artifact_type="csv",
        )
        self.assertEqual(resolved, "laporan.CSV")


class HistoryArtifactCleanupHelperTest(APISimpleTestCase):
    @patch("api.views.safe_join", side_effect=ValueError("unsafe path"))
    def test_delete_history_artifact_file_skips_unsafe_cached_path(
        self,
        _mock_safe_join,
    ):
        artifact = SimpleNamespace(artifact_type="csv", file_name="cached.csv")

        with self.assertLogs("api.views", level="WARNING") as log:
            _delete_history_artifact_file(artifact)

        self.assertIn(
            "History artifact cleanup skipped because the cached path is unsafe.",
            "\n".join(log.output),
        )

    @patch("api.views.os.remove", side_effect=FileNotFoundError())
    def test_delete_history_artifact_file_ignores_missing_cached_file(
        self,
        mock_remove,
    ):
        artifact = SimpleNamespace(artifact_type="csv", file_name="cached.csv")

        _delete_history_artifact_file(artifact)

        mock_remove.assert_called_once()

    @patch("api.views.os.remove", side_effect=OSError("disk busy"))
    def test_delete_history_artifact_file_logs_warning_when_cleanup_fails(
        self,
        mock_remove,
    ):
        artifact = SimpleNamespace(artifact_type="csv", file_name="cached.csv")

        with self.assertLogs("api.views", level="WARNING") as log:
            _delete_history_artifact_file(artifact)

        mock_remove.assert_called_once()
        self.assertIn(
            "Failed to delete cached history artifact file during history removal.",
            "\n".join(log.output),
        )


class SessionEndpointTests(TestCase):
    def setUp(self):
        self.factory = APIRequestFactory()
        self.session_resource_view = views.SessionResourceView.as_view()
        self.user = User.objects.create_user(
            email="owner@example.com",
            name="Owner",
            password="Test12345",
            status="verified",
        )
        self.unverified_user = User.objects.create_user(
            email="pending@example.com",
            name="Pending",
            password="Test12345",
            status="unverified",
        )
        self.session_id = "123e4567-e89b-12d3-a456-426614174000"
        self.output_id = "123e4567-e89b-12d3-a456-426614174001"

    @patch("api.views.list_sessions_for_user")
    def test_session_list_requires_authentication(self, mock_list_sessions):
        request = self.factory.get("/sessions/")

        response = views.session_list(request)

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)
        mock_list_sessions.assert_not_called()

    @patch("api.views.list_sessions_for_user")
    def test_session_list_requires_verified_user(self, mock_list_sessions):
        request = self.factory.get("/sessions/")
        force_authenticate(request, user=self.unverified_user)

        response = views.session_list(request)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        mock_list_sessions.assert_not_called()

    @patch("api.views.list_sessions_for_user")
    def test_session_list_returns_serialized_owned_sessions(self, mock_list_sessions):
        stub_session = SimpleNamespace(
            id="123e4567-e89b-12d3-a456-426614174000",
            title="April report",
            created_at="2026-04-21T10:00:00Z",
            updated_at="2026-04-21T10:05:00Z",
            last_message_at=None,
            last_output_at=None,
        )
        mock_list_sessions.return_value = [stub_session]
        request = self.factory.get("/sessions/")
        force_authenticate(request, user=self.user)

        response = views.session_list(request)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["title"], "April report")
        self.assertEqual(response.data["limit"], 10)
        self.assertEqual(response.data["offset"], 0)
        mock_list_sessions.assert_called_once_with(self.user, limit=10, offset=0)

    @patch("api.views.list_sessions_for_user")
    def test_session_list_applies_default_pagination_contract(self, mock_list_sessions):
        stub_session = SimpleNamespace(
            id="123e4567-e89b-12d3-a456-426614174000",
            title="April report",
            created_at="2026-04-21T10:00:00Z",
            updated_at="2026-04-21T10:05:00Z",
            last_message_at=None,
            last_output_at=None,
        )
        mock_list_sessions.return_value = [stub_session]
        request = self.factory.get("/sessions/")
        force_authenticate(request, user=self.user)

        response = views.session_list(request)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["limit"], 10)
        self.assertEqual(response.data["offset"], 0)
        mock_list_sessions.assert_called_once_with(self.user, limit=10, offset=0)

    @patch("api.views.list_sessions_for_user")
    def test_session_list_allows_explicit_limit_and_offset(self, mock_list_sessions):
        mock_list_sessions.return_value = []
        request = self.factory.get("/sessions/?limit=5&offset=10")
        force_authenticate(request, user=self.user)

        response = views.session_list(request)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["limit"], 5)
        self.assertEqual(response.data["offset"], 10)
        mock_list_sessions.assert_called_once_with(self.user, limit=5, offset=10)

    @patch("api.views.list_sessions_for_user")
    def test_session_list_rejects_non_numeric_limit(self, mock_list_sessions):
        request = self.factory.get("/sessions/?limit=abc")
        force_authenticate(request, user=self.user)

        response = views.session_list(request)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data, {"detail": "limit must be an integer."})
        mock_list_sessions.assert_not_called()

    @patch("api.views.list_sessions_for_user")
    def test_session_list_rejects_negative_offset(self, mock_list_sessions):
        request = self.factory.get("/sessions/?offset=-1")
        force_authenticate(request, user=self.user)

        response = views.session_list(request)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.data,
            {"detail": "offset must be greater than or equal to 0."},
        )
        mock_list_sessions.assert_not_called()

    @patch("api.views.list_sessions_for_user")
    def test_session_list_rejects_non_numeric_offset(self, mock_list_sessions):
        request = self.factory.get("/sessions/?offset=abc")
        force_authenticate(request, user=self.user)

        response = views.session_list(request)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data, {"detail": "offset must be an integer."})
        mock_list_sessions.assert_not_called()

    @patch("api.views.list_sessions_for_user")
    def test_session_list_returns_whitelisted_message_for_service_validation_error(self, mock_list_sessions):
        mock_list_sessions.side_effect = ValueError("limit must be less than or equal to 50.")
        request = self.factory.get("/sessions/?limit=51")
        force_authenticate(request, user=self.user)

        response = views.session_list(request)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.data,
            {"detail": "limit must be less than or equal to 50."},
        )
        mock_list_sessions.assert_called_once_with(self.user, limit=51, offset=0)

    @patch("api.views.list_sessions_for_user")
    def test_session_list_hides_non_whitelisted_service_validation_error(self, mock_list_sessions):
        mock_list_sessions.side_effect = ValueError("database exploded")
        request = self.factory.get("/sessions/?limit=10")
        force_authenticate(request, user=self.user)

        response = views.session_list(request)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.data,
            {"detail": "Invalid session list pagination."},
        )
        mock_list_sessions.assert_called_once_with(self.user, limit=10, offset=0)

    @patch("api.views.get_paginated_session_detail_for_user")
    def test_session_detail_returns_not_found_when_missing(self, mock_get_session_detail):
        mock_get_session_detail.return_value = None
        request = self.factory.get("/sessions/session-1/")
        force_authenticate(request, user=self.user)

        response = views.session_detail(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
        mock_get_session_detail.assert_called_once_with(
            self.user,
            "session-1",
            messages_limit=20,
            messages_offset=0,
            outputs_limit=10,
            outputs_offset=0,
        )

    @patch("api.views.get_paginated_session_detail_for_user")
    def test_session_detail_returns_serialized_session_when_found(self, mock_get_session_detail):
        mock_get_session_detail.return_value = SimpleNamespace(
            id="123e4567-e89b-12d3-a456-426614174000",
            title="April report",
            created_at="2026-04-21T10:00:00Z",
            updated_at="2026-04-21T10:05:00Z",
            last_message_at=None,
            last_output_at=None,
            messages={"count": 0, "limit": 20, "offset": 0, "results": []},
            generated_outputs={"count": 0, "limit": 10, "offset": 0, "results": []},
        )
        request = self.factory.get("/sessions/session-1/")
        force_authenticate(request, user=self.user)

        response = views.session_detail(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["title"], "April report")
        self.assertEqual(response.data["messages"]["limit"], 20)
        self.assertEqual(response.data["generated_outputs"]["limit"], 10)
        mock_get_session_detail.assert_called_once_with(
            self.user,
            "session-1",
            messages_limit=20,
            messages_offset=0,
            outputs_limit=10,
            outputs_offset=0,
        )

    @patch("api.views.get_paginated_session_detail_for_user")
    def test_session_detail_allows_explicit_messages_and_outputs_pagination(
        self,
        mock_get_session_detail,
    ):
        mock_get_session_detail.return_value = SimpleNamespace(
            id="123e4567-e89b-12d3-a456-426614174000",
            title="April report",
            created_at="2026-04-21T10:00:00Z",
            updated_at="2026-04-21T10:05:00Z",
            last_message_at=None,
            last_output_at=None,
            messages={"count": 5, "limit": 2, "offset": 1, "results": []},
            generated_outputs={"count": 3, "limit": 1, "offset": 2, "results": []},
        )
        request = self.factory.get(
            "/sessions/session-1/?messages_limit=2&messages_offset=1&outputs_limit=1&outputs_offset=2"
        )
        force_authenticate(request, user=self.user)

        response = views.session_detail(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["messages"]["limit"], 2)
        self.assertEqual(response.data["messages"]["offset"], 1)
        self.assertEqual(response.data["generated_outputs"]["limit"], 1)
        self.assertEqual(response.data["generated_outputs"]["offset"], 2)
        mock_get_session_detail.assert_called_once_with(
            self.user,
            "session-1",
            messages_limit=2,
            messages_offset=1,
            outputs_limit=1,
            outputs_offset=2,
        )

    @patch("api.views.get_paginated_session_detail_for_user")
    def test_session_detail_rejects_non_numeric_messages_limit(self, mock_get_session_detail):
        request = self.factory.get("/sessions/session-1/?messages_limit=abc")
        force_authenticate(request, user=self.user)

        response = views.session_detail(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data, {"detail": "messages_limit must be an integer."})
        mock_get_session_detail.assert_not_called()

    @patch("api.views.get_paginated_session_detail_for_user")
    def test_session_detail_rejects_negative_outputs_offset(self, mock_get_session_detail):
        request = self.factory.get("/sessions/session-1/?outputs_offset=-1")
        force_authenticate(request, user=self.user)

        response = views.session_detail(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.data,
            {"detail": "outputs_offset must be greater than or equal to 0."},
        )
        mock_get_session_detail.assert_not_called()

    @patch("api.views.get_paginated_session_detail_for_user")
    def test_session_detail_rejects_zero_messages_limit(self, mock_get_session_detail):
        request = self.factory.get("/sessions/session-1/?messages_limit=0")
        force_authenticate(request, user=self.user)

        response = views.session_detail(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data, {"detail": "messages_limit must be greater than 0."})
        mock_get_session_detail.assert_not_called()

    @patch("api.views.get_paginated_session_detail_for_user")
    def test_session_detail_rejects_outputs_limit_above_maximum(self, mock_get_session_detail):
        request = self.factory.get("/sessions/session-1/?outputs_limit=51")
        force_authenticate(request, user=self.user)

        response = views.session_detail(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data, {"detail": "outputs_limit must be less than or equal to 50."})
        mock_get_session_detail.assert_not_called()

    @patch("api.views.get_paginated_session_detail_for_user")
    def test_session_detail_hides_non_whitelisted_service_validation_error(
        self,
        mock_get_session_detail,
    ):
        mock_get_session_detail.side_effect = ValueError("internal query detail")
        request = self.factory.get("/sessions/session-1/?messages_limit=10")
        force_authenticate(request, user=self.user)

        response = views.session_detail(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data, {"detail": "Invalid session detail pagination."})

    @patch("api.views.update_session_title")
    @patch("api.views.get_session_for_user")
    def test_session_update_rejects_blank_title(self, mock_get_session, mock_update_session_title):
        mock_get_session.return_value = SimpleNamespace(id="session-1", title="Old title")
        request = self.factory.patch(
            "/sessions/session-1/",
            {"title": "   "},
            format="json",
        )
        force_authenticate(request, user=self.user)

        response = views.session_update(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        mock_get_session.assert_called_once_with(self.user, "session-1")
        mock_update_session_title.assert_not_called()

    @patch("api.views.get_session_for_user")
    def test_session_update_returns_not_found_when_missing(self, mock_get_session):
        mock_get_session.return_value = None
        request = self.factory.patch(
            "/sessions/session-1/",
            {"title": "Renamed"},
            format="json",
        )
        force_authenticate(request, user=self.user)

        response = views.session_update(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
        mock_get_session.assert_called_once_with(self.user, "session-1")

    @patch("api.views.update_session_title")
    @patch("api.views.get_session_for_user")
    def test_session_update_returns_serialized_updated_session(
        self,
        mock_get_session,
        mock_update_session_title,
    ):
        original_session = SimpleNamespace(id="session-1", title="Old title")
        updated_session = SimpleNamespace(
            id="123e4567-e89b-12d3-a456-426614174000",
            title="Renamed",
            created_at="2026-04-21T10:00:00Z",
            updated_at="2026-04-21T10:05:00Z",
            last_message_at=None,
            last_output_at=None,
        )
        mock_get_session.return_value = original_session
        mock_update_session_title.return_value = updated_session
        request = self.factory.patch(
            "/sessions/session-1/",
            {"title": "  Renamed  "},
            format="json",
        )
        force_authenticate(request, user=self.user)

        response = views.session_update(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["title"], "Renamed")
        mock_get_session.assert_called_once_with(self.user, "session-1")
        mock_update_session_title.assert_called_once_with(original_session, "Renamed")

    @patch("api.views.delete_session")
    @patch("api.views.get_session_for_user")
    def test_session_delete_returns_no_content_for_owned_session(self, mock_get_session, mock_delete_session):
        stub_session = SimpleNamespace(id="session-1")
        mock_get_session.return_value = stub_session
        request = self.factory.delete("/sessions/session-1/")
        force_authenticate(request, user=self.user)

        response = views.session_delete(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)
        mock_get_session.assert_called_once_with(self.user, "session-1")
        mock_delete_session.assert_called_once_with(stub_session)

    @patch("api.views.delete_session")
    @patch("api.views.get_session_for_user")
    def test_session_delete_returns_not_found_when_missing(self, mock_get_session, mock_delete_session):
        mock_get_session.return_value = None
        request = self.factory.delete("/sessions/session-1/")
        force_authenticate(request, user=self.user)

        response = views.session_delete(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
        mock_get_session.assert_called_once_with(self.user, "session-1")
        mock_delete_session.assert_not_called()

    @patch("api.views._build_session_detail_response")
    def test_session_resource_dispatches_get_to_detail(self, mock_session_detail):
        mock_session_detail.return_value = Response(status=status.HTTP_200_OK)
        request = self.factory.get("/sessions/session-1/")
        force_authenticate(request, user=self.user)

        response = self.session_resource_view(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        mock_session_detail.assert_called_once()
        self.assertEqual(mock_session_detail.call_args.args[0].user, self.user)
        self.assertEqual(mock_session_detail.call_args.args[1], "session-1")

    @patch("api.views._build_session_update_response")
    def test_session_resource_dispatches_patch_to_update(self, mock_session_update):
        mock_session_update.return_value = Response(status=status.HTTP_200_OK)
        request = self.factory.patch("/sessions/session-1/", {"title": "Renamed"}, format="json")
        force_authenticate(request, user=self.user)

        response = self.session_resource_view(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        mock_session_update.assert_called_once()
        self.assertEqual(mock_session_update.call_args.args[0], self.user)
        self.assertEqual(mock_session_update.call_args.args[1], {"title": "Renamed"})
        self.assertEqual(mock_session_update.call_args.args[2], "session-1")

    @patch("api.views._build_session_delete_response")
    def test_session_resource_dispatches_delete_to_delete(self, mock_session_delete):
        mock_session_delete.return_value = Response(status=status.HTTP_204_NO_CONTENT)
        request = self.factory.delete("/sessions/session-1/")
        force_authenticate(request, user=self.user)

        response = self.session_resource_view(request, "session-1")

        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)
        mock_session_delete.assert_called_once()
        self.assertEqual(mock_session_delete.call_args.args[0], self.user)
        self.assertEqual(mock_session_delete.call_args.args[1], "session-1")

    @patch("api.views.get_generated_output_for_session_user")
    def test_session_output_download_csv_requires_authentication(self, mock_get_output):
        request = self.factory.get(
            f"/sessions/{self.session_id}/outputs/{self.output_id}/download/csv/"
        )

        response = views.session_output_download_csv(
            request,
            self.session_id,
            self.output_id,
        )

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)
        mock_get_output.assert_not_called()

    @patch("api.views.get_generated_output_for_session_user")
    def test_session_output_download_csv_requires_verified_user(self, mock_get_output):
        request = self.factory.get(
            f"/sessions/{self.session_id}/outputs/{self.output_id}/download/csv/"
        )
        force_authenticate(request, user=self.unverified_user)

        response = views.session_output_download_csv(
            request,
            self.session_id,
            self.output_id,
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        mock_get_output.assert_not_called()

    @patch("api.views.open")
    @patch("api.views.export_csv_to_filesystem")
    @patch("api.views.get_generated_output_for_session_user")
    def test_session_output_download_csv_returns_file_for_owned_output(
        self,
        mock_get_output,
        mock_export_csv,
        mock_open_file,
    ):
        stub_output = SimpleNamespace(
            id=self.output_id,
            output_json={
                "document_info": {
                    "filename": "invoice.pdf",
                    "source_type": "PDF",
                },
                "summary": {"table_count": 1},
                "content_data": [],
            },
        )
        mock_get_output.return_value = stub_output
        mock_export_csv.return_value = {
            "file_id": "csv_token",
            "file_name": "export_token.csv",
            "artifact_type": "csv",
            "size_bytes": 12,
            "created_at": "2026-04-26T08:00:00Z",
        }
        mock_open_file.return_value = BytesIO(b"col\n1\n")
        request = self.factory.get(
            f"/sessions/{self.session_id}/outputs/{self.output_id}/download/csv/?filename=report.csv"
        )
        force_authenticate(request, user=self.user)

        response = views.session_output_download_csv(
            request,
            self.session_id,
            self.output_id,
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertIn(
            'attachment; filename="report.csv"',
            response["Content-Disposition"],
        )
        mock_get_output.assert_called_once_with(
            self.user,
            self.session_id,
            self.output_id,
        )
        self.assertEqual(
            mock_export_csv.call_args.kwargs["output_json"],
            {
                "document_info": {
                    "filename": "invoice.pdf",
                    "source_type": "PDF",
                },
                "summary": {"table_count": 1},
                "content_data": [],
            },
        )

    @patch("api.views.get_generated_output_for_session_user")
    def test_session_output_download_csv_returns_not_found_when_output_missing(
        self,
        mock_get_output,
    ):
        mock_get_output.return_value = None
        request = self.factory.get(
            f"/sessions/{self.session_id}/outputs/{self.output_id}/download/csv/"
        )
        force_authenticate(request, user=self.user)

        response = views.session_output_download_csv(
            request,
            self.session_id,
            self.output_id,
        )

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
        self.assertEqual(response.data, {"detail": "Not found."})
        mock_get_output.assert_called_once_with(
            self.user,
            self.session_id,
            self.output_id,
        )

    @patch("api.views.export_csv_to_filesystem", side_effect=OutputCSVGenerationError("boom"))
    @patch("api.views.get_generated_output_for_session_user")
    def test_session_output_download_csv_returns_internal_error_when_generation_fails(
        self,
        mock_get_output,
        mock_export_csv,
    ):
        mock_get_output.return_value = SimpleNamespace(
            id=self.output_id,
            output_json={
                "document_info": {
                    "filename": "invoice.pdf",
                    "source_type": "PDF",
                },
                "summary": {"table_count": 1},
                "content_data": [],
            },
        )
        request = self.factory.get(
            f"/sessions/{self.session_id}/outputs/{self.output_id}/download/csv/"
        )
        force_authenticate(request, user=self.user)

        response = views.session_output_download_csv(
            request,
            self.session_id,
            self.output_id,
        )

        self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
        self.assertEqual(
            response.data,
            {
                "status": "error",
                "message": "Failed to download CSV due to internal error.",
            },
        )
        mock_get_output.assert_called_once_with(
            self.user,
            self.session_id,
            self.output_id,
        )
        mock_export_csv.assert_called_once()

    @patch("api.views.get_generated_output_for_session_user")
    def test_session_output_download_excel_requires_authentication(self, mock_get_output):
        request = self.factory.get(
            f"/sessions/{self.session_id}/outputs/{self.output_id}/download/excel/"
        )

        response = views.session_output_download_excel(
            request,
            self.session_id,
            self.output_id,
        )

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)
        mock_get_output.assert_not_called()

    @patch("api.views.get_generated_output_for_session_user")
    def test_session_output_download_excel_requires_verified_user(self, mock_get_output):
        request = self.factory.get(
            f"/sessions/{self.session_id}/outputs/{self.output_id}/download/excel/"
        )
        force_authenticate(request, user=self.unverified_user)

        response = views.session_output_download_excel(
            request,
            self.session_id,
            self.output_id,
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        mock_get_output.assert_not_called()

    @patch("api.views.open")
    @patch("api.views.export_excel_to_filesystem")
    @patch("api.views.get_generated_output_for_session_user")
    def test_session_output_download_excel_returns_file_for_owned_output(
        self,
        mock_get_output,
        mock_export_excel,
        mock_open_file,
    ):
        stub_output = SimpleNamespace(
            id=self.output_id,
            output_json={
                "document_info": {
                    "filename": "invoice.pdf",
                    "source_type": "PDF",
                },
                "summary": {"table_count": 1},
                "content_data": [],
            },
        )
        mock_get_output.return_value = stub_output
        mock_export_excel.return_value = {
            "file_id": "xlsx_token",
            "file_name": "export_token.xlsx",
            "artifact_type": "xlsx",
            "size_bytes": 24,
            "created_at": "2026-04-26T08:00:00Z",
        }
        mock_open_file.return_value = BytesIO(b"excel-bytes")
        request = self.factory.get(
            f"/sessions/{self.session_id}/outputs/{self.output_id}/download/excel/?filename=report.xlsx"
        )
        force_authenticate(request, user=self.user)

        response = views.session_output_download_excel(
            request,
            self.session_id,
            self.output_id,
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            'attachment; filename="report.xlsx"',
            response["Content-Disposition"],
        )
        mock_get_output.assert_called_once_with(
            self.user,
            self.session_id,
            self.output_id,
        )
        self.assertEqual(
            mock_export_excel.call_args.kwargs["output_json"],
            {
                "document_info": {
                    "filename": "invoice.pdf",
                    "source_type": "PDF",
                },
                "summary": {"table_count": 1},
                "content_data": [],
            },
        )

    @patch("api.views.open")
    @patch("api.views.export_excel_to_filesystem")
    @patch("api.views.get_generated_output_for_session_user")
    def test_session_output_download_excel_repairs_legacy_invalid_source_type_from_filename(
        self,
        mock_get_output,
        mock_export_excel,
        mock_open_file,
    ):
        stub_output = SimpleNamespace(
            id=self.output_id,
            output_json={
                "document_info": {
                    "source_type": "unknown",
                    "filename": "invoice.pdf",
                },
                "summary": {"table_count": 1},
                "content_data": [],
            },
        )
        mock_get_output.return_value = stub_output
        mock_export_excel.return_value = {
            "file_id": "xlsx_token",
            "file_name": "export_token.xlsx",
            "artifact_type": "xlsx",
            "size_bytes": 24,
            "created_at": "2026-04-26T08:00:00Z",
        }
        mock_open_file.return_value = BytesIO(b"excel-bytes")
        request = self.factory.get(
            f"/sessions/{self.session_id}/outputs/{self.output_id}/download/excel/"
        )
        force_authenticate(request, user=self.user)

        response = views.session_output_download_excel(
            request,
            self.session_id,
            self.output_id,
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            mock_export_excel.call_args.kwargs["output_json"],
            {
                "document_info": {
                    "source_type": "PDF",
                    "filename": "invoice.pdf",
                },
                "summary": {"table_count": 1},
                "content_data": [],
            },
        )

    @patch("api.views.get_generated_output_for_session_user")
    def test_session_output_download_excel_returns_not_found_when_output_missing(
        self,
        mock_get_output,
    ):
        mock_get_output.return_value = None
        request = self.factory.get(
            f"/sessions/{self.session_id}/outputs/{self.output_id}/download/excel/"
        )
        force_authenticate(request, user=self.user)

        response = views.session_output_download_excel(
            request,
            self.session_id,
            self.output_id,
        )

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
        self.assertEqual(response.data, {"detail": "Not found."})
        mock_get_output.assert_called_once_with(
            self.user,
            self.session_id,
            self.output_id,
        )

    @patch("api.views.export_excel_to_filesystem", side_effect=OutputExcelGenerationError("boom"))
    @patch("api.views.get_generated_output_for_session_user")
    def test_session_output_download_excel_returns_internal_error_when_generation_fails(
        self,
        mock_get_output,
        mock_export_excel,
    ):
        mock_get_output.return_value = SimpleNamespace(
            id=self.output_id,
            output_json={
                "document_info": {
                    "filename": "invoice.pdf",
                    "source_type": "PDF",
                },
                "summary": {"table_count": 1},
                "content_data": [],
            },
        )
        request = self.factory.get(
            f"/sessions/{self.session_id}/outputs/{self.output_id}/download/excel/"
        )
        force_authenticate(request, user=self.user)

        response = views.session_output_download_excel(
            request,
            self.session_id,
            self.output_id,
        )

        self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
        self.assertEqual(
            response.data,
            {
                "status": "error",
                "message": "Failed to download Excel due to internal error.",
            },
        )
        mock_get_output.assert_called_once_with(
            self.user,
            self.session_id,
            self.output_id,
        )
        mock_export_excel.assert_called_once()
