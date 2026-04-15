import os
import logging
from uuid import uuid4

from django.conf import settings
from django.utils.text import get_valid_filename
from PyPDF2 import PdfReader
from PyPDF2.errors import PdfReadError

from .excel_service import process_uploaded_excel
from .txt_service import process_uploaded_txt
from .csv_service import process_uploaded_csv
from file_processing.extractors.image_extractor import ImageExtractor
from file_processing.services.ocr_service import OCRService
from file_processing.services.non_ocr_pdf_service import NonOCRPDFService
from file_processing.services import word_validation_service
from file_processing.services.image_validation_service import validate_image
from file_processing.services.word_extraction_service import WordExtractionService
from file_processing.services.contracts import ValidationResult, ExtractionResult
from file_processing.utils.upload_constants import MAX_FILE_SIZE, FILE_TOO_LARGE_ERROR

try:
    import magic
except Exception:  # pragma: no cover - optional dependency in local envs

    class _MagicShim:
        @staticmethod
        def from_buffer(_buffer, mime=True):
            raise ImportError("python-magic unavailable")

    magic = _MagicShim()

logger = logging.getLogger(__name__)

EXT_XLSX = ".xlsx"
EXT_XLS = ".xls"
EXT_PDF = ".pdf"
EXT_DOCX = ".docx"
EXT_DOC = ".doc"
MIME_OCTET_STREAM = "application/octet-stream"
MIME_OLE_STORAGE = "application/x-ole-storage"
MIME_ZIP = "application/zip"

EXT_TXT = ".txt"
EXT_CSV = ".csv"
EXT_PNG = ".png"
EXT_JPG = ".jpg"
EXT_JPEG = ".jpeg"

MIME_OCTET_STREAM = "application/octet-stream"
MIME_OLE_STORAGE = "application/x-ole-storage"
MIME_ZIP = "application/zip"

IMAGE_EXTENSIONS = {EXT_PNG, EXT_JPG, EXT_JPEG}
ALLOWED_EXTENSIONS = [
    EXT_PDF,
    EXT_XLS,
    EXT_XLSX,
    EXT_PNG,
    EXT_JPG,
    EXT_JPEG,
    EXT_DOCX,
    EXT_DOC,
    EXT_TXT,
    EXT_CSV,
]

ALLOWED_MIME_TYPES = {
    EXT_PDF: [
        "application/pdf",
        "application/x-pdf",
        "application/vnd.pdf",
    ],
    EXT_XLS: [
        "application/vnd.ms-excel",
        "application/msexcel",
        "application/x-msexcel",
        "application/x-ms-excel",
        "application/x-excel",
        "application/xls",
        "application/x-xls",
        MIME_OCTET_STREAM,
        MIME_OLE_STORAGE,
        "application/CDFV2",
        "application/vnd.ms-office",
    ],
    EXT_XLSX: [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        MIME_ZIP,
        "application/x-zip",
        "application/x-zip-compressed",
        MIME_OCTET_STREAM,
    ],
    EXT_DOC: [
        "application/msword",
        "application/doc",
        "application/vnd.ms-word",
        "application/x-msword",
        MIME_OCTET_STREAM,
        MIME_OLE_STORAGE,
        "application/CDFV2",
    ],
    EXT_DOCX: [
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        MIME_ZIP,
        "application/x-zip",
        "application/x-zip-compressed",
        MIME_OCTET_STREAM,
        MIME_OLE_STORAGE,
    ],
    EXT_TXT: [
        "text/plain",
        "text/x-log",
    ],
    EXT_CSV: [
        "text/csv",
        "text/plain",
        "application/csv",
        "application/vnd.ms-excel",
    ],
}

MAX_PDF_PAGES = 100
MAX_EXCEL_SHEETS = 100
PDF_CORRUPT_ERROR = "PDF file is corrupt or has an invalid structure."
EXCEL_CORRUPT_ERROR = "Invalid or corrupted Excel file."
EXCEL_TOO_MANY_SHEETS_ERROR = f"Excel has too many sheets (maximum {MAX_EXCEL_SHEETS})."
EXCEL_PASSWORD_PROTECTED_ERROR = (
    "Excel file is password-protected. Please remove the password and try again."
)
MAX_WORD_PAGES = word_validation_service.MAX_WORD_PAGES
WORD_CORRUPT_ERROR = word_validation_service.WORD_CORRUPT_ERROR
TXT_CORRUPT_ERROR = "File teks tidak dapat dibaca atau rusak (corrupt)."
TXT_PROTECTED_ERROR = (
    "File terdeteksi sebagai format terproteksi atau terenkripsi. "
    "Pastikan file adalah teks biasa (.txt) yang tidak diproteksi."
)
CSV_CORRUPT_ERROR = "File CSV tidak dapat dibaca atau rusak (corrupt)."
CSV_PROTECTED_ERROR = (
    "File CSV terdeteksi sebagai format terproteksi atau terenkripsi. "
    "Pastikan file adalah CSV biasa yang tidak diproteksi."
)
MIME_TYPE_DETECTION_ERROR = "Unable to determine file type."
FILE_EXTENSION_MISMATCH_ERROR = "File content does not match its extension."
ZIP_SIGNATURE_PREFIX = b"PK"
DOES_NOT_MATCH_EXTENSION_ERROR = "File content does not match its extension."
OLE_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
ZIP_SIGNATURE_PREFIX = b"PK"

BINARY_SIGNATURES: list[tuple[bytes, str]] = [
    (b"\x50\x4b\x03\x04", FILE_EXTENSION_MISMATCH_ERROR),
    (b"\x50\x4b\x05\x06", FILE_EXTENSION_MISMATCH_ERROR),
    (OLE_SIGNATURE, TXT_PROTECTED_ERROR),
    (b"\x7fELF", FILE_EXTENSION_MISMATCH_ERROR),
    (b"MZ", FILE_EXTENSION_MISMATCH_ERROR),
    (b"%PDF", FILE_EXTENSION_MISMATCH_ERROR),
    (b"\xff\xd8\xff", FILE_EXTENSION_MISMATCH_ERROR),
    (b"\x89PNG", FILE_EXTENSION_MISMATCH_ERROR),
]


def _has_extracted_text(extracted_data):
    if not extracted_data or "content" not in extracted_data:
        return False
    for page in extracted_data["content"]:
        if page.get("text"):
            return True
    return False


def _get_empty_page_numbers(extracted_data):
    empty = []
    if not extracted_data or "content" not in extracted_data:
        return empty
    for page in extracted_data["content"]:
        if not page.get("text"):
            empty.append(page["page"])
    return empty


def _process_pdf(file_path, uploaded_file):
    is_valid, error = validate_pdf(uploaded_file)
    if not is_valid:
        return False, error, None

    try:
        extracted_data = NonOCRPDFService.extract_non_ocr_pdf_to_json(file_path)

        if not _has_extracted_text(extracted_data):
            extracted_data = OCRService.process_pdf(file_path)
        else:
            empty_pages = _get_empty_page_numbers(extracted_data)
            if empty_pages:
                ocr_data = OCRService.process_pdf_pages(file_path, empty_pages)
                # Merge OCR results into the extracted data
                ocr_by_page = {p["page"]: p for p in ocr_data.get("content", [])}
                for page in extracted_data["content"]:
                    if page["page"] in ocr_by_page:
                        page["text"] = ocr_by_page[page["page"]]["text"]
    except Exception:
        logger.exception("Non-OCR extraction failed, fallback to OCR")
        extracted_data = OCRService.process_pdf(file_path)

    return True, None, extracted_data


def _process_image_result(file_path) -> ExtractionResult:
    try:
        extractor = ImageExtractor()
        extracted_data = extractor.extract(file_path)
        return ExtractionResult.ok(extracted_data)
    except ValueError as exc:
        return ExtractionResult.fail(str(exc))
    except Exception:
        logger.exception("Image extraction failed.")
        return ExtractionResult.fail("Image OCR extraction failed.")


def _process_image(file_path):
    return _process_image_result(file_path).to_legacy_tuple()


def process_word_result(file_path, ext) -> ExtractionResult:
    try:
        extracted_data = WordExtractionService.extract_word_to_json(file_path, ext)
    except ValueError as exc:
        return ExtractionResult.fail(str(exc))
    except Exception:
        logger.exception("Word extraction failed")
        return ExtractionResult.fail(WORD_CORRUPT_ERROR)

    return ExtractionResult.ok(extracted_data)


def process_word(file_path, ext):
    return process_word_result(file_path, ext).to_legacy_tuple()



def _dispatch_upload_processing(ext, file_path, uploaded_file):
    processors = {
        EXT_PDF: lambda: _process_pdf(file_path, uploaded_file),
        EXT_XLS: lambda: process_uploaded_excel(file_path),
        EXT_XLSX: lambda: process_uploaded_excel(file_path),
        EXT_DOC: lambda: process_word(file_path, EXT_DOC),
        EXT_DOCX: lambda: process_word(file_path, EXT_DOCX),
        EXT_TXT: lambda: process_uploaded_txt(file_path),
        EXT_CSV: lambda: process_uploaded_csv(file_path),
        EXT_PNG: lambda: _process_image(file_path),
        EXT_JPG: lambda: _process_image(file_path),
        EXT_JPEG: lambda: _process_image(file_path),
    }

    processor = processors.get(ext)
    if processor is None:
        return False, "Unsupported file type", None

    return processor()


def _get_upload_extension(uploaded_file):
    return os.path.splitext(uploaded_file.name)[1].lower()


def _cleanup_temp_upload(file_path):
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception:
        logger.exception("Failed to delete temporary upload file.")


def _process_stored_upload(ext, file_path, uploaded_file):
    success, error, extracted_data = _dispatch_upload_processing(
        ext,
        file_path,
        uploaded_file,
    )
    if not success:
        return False, error, None
    return True, None, extracted_data


def process_upload(uploaded_file):
    is_valid, error = validate_file(uploaded_file)
    if not is_valid:
        return False, error, None, None

    ext = _get_upload_extension(uploaded_file)
    file_path = save_temp_file(uploaded_file)

    try:
        success, error, extracted_data = _process_stored_upload(
            ext,
            file_path,
            uploaded_file,
        )
        if not success:
            return False, error, None, None

    finally:
        _cleanup_temp_upload(file_path)

    return True, None, None, extracted_data


def validate_file(uploaded_file):
    return validate_file_result(uploaded_file).to_legacy_tuple()


def _validate_file_content(uploaded_file, ext) -> ValidationResult:
    """Validate file size, MIME type, and type-specific constraints."""
    # Image files can return early without further checks
    if ext in IMAGE_EXTENSIONS:
        is_valid, error = validate_image(uploaded_file)
        if not is_valid:
            return ValidationResult.fail(error or "Invalid image file.")
        return ValidationResult.ok()

    # All other file types: check size and MIME type
    if uploaded_file.size > MAX_FILE_SIZE:
        return ValidationResult.fail(FILE_TOO_LARGE_ERROR)

    is_valid_mime, mime_error = validate_mime_type(uploaded_file, ext)
    if not is_valid_mime:
        return ValidationResult.fail(mime_error or MIME_TYPE_DETECTION_ERROR)

    # Type-specific format validation
    if ext in {EXT_XLS, EXT_XLSX}:
        is_valid_excel, excel_error = validate_excel_sheet_count(uploaded_file, ext)
        if not is_valid_excel:
            return ValidationResult.fail(excel_error or EXCEL_CORRUPT_ERROR)

    if ext in {EXT_DOC, EXT_DOCX}:
        is_valid_word, word_error = word_validation_service.validate_word(
            uploaded_file, ext
        )
        if not is_valid_word:
            return ValidationResult.fail(word_error or WORD_CORRUPT_ERROR)

    return ValidationResult.ok()


def validate_file_result(uploaded_file) -> ValidationResult:
    ext = os.path.splitext(uploaded_file.name)[1].lower()

    if ext not in ALLOWED_EXTENSIONS:
        return ValidationResult.fail(
            "Unsupported file type. Only PDF, XLS, XLSX, TXT, CSV, PNG, JPG, JPEG, DOC, and DOCX are allowed.",
        )

    return _validate_file_content(uploaded_file, ext)


def validate_pdf(uploaded_file):
    return validate_pdf_result(uploaded_file).to_legacy_tuple()


def validate_pdf_result(uploaded_file) -> ValidationResult:
    try:
        uploaded_file.seek(0)
        reader = PdfReader(uploaded_file, strict=True)
    except Exception:
        return ValidationResult.fail(PDF_CORRUPT_ERROR)

    is_valid, error = check_pdf_encrypted(reader)
    if not is_valid:
        return ValidationResult.fail(error or PDF_CORRUPT_ERROR)

    is_valid, page_count_or_error = check_pdf_structure(reader)
    if not is_valid:
        return ValidationResult.fail(page_count_or_error or PDF_CORRUPT_ERROR)

    page_count = page_count_or_error
    is_valid, error = check_pdf_page_count(page_count)
    if not is_valid:
        return ValidationResult.fail(error or PDF_CORRUPT_ERROR)

    return ValidationResult.ok()


def check_pdf_encrypted(reader):
    if reader.is_encrypted:
        return False, "PDF file is password-protected."
    return True, None


def check_pdf_structure(reader):
    try:
        return True, len(reader.pages)
    except PdfReadError:
        return False, PDF_CORRUPT_ERROR
    except Exception:
        return False, PDF_CORRUPT_ERROR


def check_pdf_page_count(page_count):
    if page_count > MAX_PDF_PAGES:
        return False, f"PDF exceeds the maximum allowed page count of {MAX_PDF_PAGES}."
    return True, None


def validate_excel_sheet_count(uploaded_file, ext):
    try:
        if _should_parse_as_xls(uploaded_file, ext):
            sheet_count = _get_xls_sheet_count(uploaded_file)
        else:
            sheet_count = _get_xlsx_sheet_count(uploaded_file)
    except Exception:
        logger.exception("Failed to validate Excel sheet count.")
        return False, EXCEL_CORRUPT_ERROR

    return check_excel_sheet_count(sheet_count)


def check_excel_sheet_count(sheet_count):
    if sheet_count > MAX_EXCEL_SHEETS:
        return False, EXCEL_TOO_MANY_SHEETS_ERROR
    return True, None


def _should_parse_as_xls(uploaded_file, ext):
    if ext == EXT_XLS:
        return True

    if ext == EXT_XLSX and _is_ole_container(uploaded_file):
        return _is_legacy_xls_content(uploaded_file)

    return False


def _get_xlsx_sheet_count(uploaded_file):
    from openpyxl import load_workbook

    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    try:
        return len(workbook.sheetnames)
    finally:
        workbook.close()
        uploaded_file.seek(0)


def _get_xls_sheet_count(uploaded_file):
    import xlrd

    uploaded_file.seek(0)
    workbook_bytes = uploaded_file.read()
    uploaded_file.seek(0)

    workbook = xlrd.open_workbook(file_contents=workbook_bytes, on_demand=True)
    try:
        return workbook.nsheets
    finally:
        release_resources = getattr(workbook, "release_resources", None)
        if callable(release_resources):
            release_resources()
        uploaded_file.seek(0)


def _validate_xlsx_mime_structure(uploaded_file):
    if _is_ole_container(uploaded_file):
        if _is_legacy_xls_content(uploaded_file):
            return "legacy", None
        return "invalid", EXCEL_PASSWORD_PROTECTED_ERROR

    if not _has_zip_signature(uploaded_file):
        return "invalid", DOES_NOT_MATCH_EXTENSION_ERROR

    return "valid", None


def _validate_word_mime_structure(uploaded_file, ext):
    if ext == EXT_DOC and not _is_ole_container(uploaded_file):
        return DOES_NOT_MATCH_EXTENSION_ERROR

    if ext == EXT_DOCX:
        # Encrypted OOXML (.docx) is wrapped in an OLE container; allow it to
        # continue to Word validation so it can return the protected-file error.
        if _is_ole_container(uploaded_file):
            return None
        if not _has_zip_signature(uploaded_file):
            return DOES_NOT_MATCH_EXTENSION_ERROR

    return None


def _resolve_txt_detected_mime(uploaded_file, detected_mime):
    request_mime = (getattr(uploaded_file, "content_type", "") or "").lower()
    if (
        detected_mime in {MIME_OCTET_STREAM, MIME_ZIP, MIME_OLE_STORAGE}
        and request_mime
    ):
        return request_mime
    return detected_mime


def validate_mime_type(uploaded_file, ext):
    try:
        head = _read_head(uploaded_file)
        mime = _detect_mime(head, ext)

        if not mime:
            return False, MIME_TYPE_DETECTION_ERROR

        expected_mimes = ALLOWED_MIME_TYPES.get(ext, [])

        if ext == EXT_XLSX:
            xlsx_state, xlsx_error = _validate_xlsx_mime_structure(uploaded_file)
            if xlsx_state == "legacy":
                return True, None
            if xlsx_error:
                return False, xlsx_error

        if ext == EXT_TXT:
            detected_mime = _resolve_txt_detected_mime(uploaded_file, mime)
            return _validate_txt_content(uploaded_file, detected_mime)

        if ext == EXT_CSV:
            return _validate_csv_content(uploaded_file, mime)

        if ext in {EXT_DOC, EXT_DOCX}:
            word_error = _validate_word_mime_structure(uploaded_file, ext)
            if word_error:
                return False, word_error

        if mime not in expected_mimes:
            return False, DOES_NOT_MATCH_EXTENSION_ERROR

        return True, None

    except Exception:
        return False, MIME_TYPE_DETECTION_ERROR


def _read_head(uploaded_file, size=2048):
    uploaded_file.seek(0)
    head = uploaded_file.read(size)
    uploaded_file.seek(0)
    return head


def _detect_mime(head, ext):
    try:
        return magic.from_buffer(head, mime=True)
    except Exception:
        return _fallback_mime(head, ext)


def _fallback_mime(head, ext):
    if head.startswith(b"%PDF"):
        return "application/pdf"
    if head.startswith(ZIP_SIGNATURE_PREFIX):
        return MIME_ZIP
    if head.startswith(OLE_SIGNATURE):
        return MIME_OLE_STORAGE
    if ext in {EXT_XLS, EXT_DOC}:
        return MIME_OCTET_STREAM
    return None


def _is_ole_container(uploaded_file):
    try:
        uploaded_file.seek(0)
        header = uploaded_file.read(len(OLE_SIGNATURE))
        uploaded_file.seek(0)
        return header == OLE_SIGNATURE
    except Exception:
        return False


def _is_legacy_xls_content(uploaded_file):
    try:
        import xlrd
    except Exception:
        return False

    try:
        uploaded_file.seek(0)
        content = uploaded_file.read()
        uploaded_file.seek(0)
        xlrd.open_workbook(file_contents=content, on_demand=True)
        return True
    except Exception:
        return False


def _has_zip_signature(uploaded_file):
    try:
        uploaded_file.seek(0)
        header = uploaded_file.read(len(ZIP_SIGNATURE_PREFIX))
        uploaded_file.seek(0)
        return header == ZIP_SIGNATURE_PREFIX
    except Exception:
        return False


def _has_binary_signature(uploaded_file):
    try:
        max_prefix = max(len(sig) for sig, _ in BINARY_SIGNATURES)
        uploaded_file.seek(0)
        header = uploaded_file.read(max_prefix)
        uploaded_file.seek(0)

        for signature, error_msg in BINARY_SIGNATURES:
            if header.startswith(signature):
                return True, error_msg

        return False, None
    except Exception:
        return False, None


def _validate_txt_content(uploaded_file, detected_mime: str):
    is_binary, binary_error = _has_binary_signature(uploaded_file)
    if is_binary:
        return False, binary_error

    if detected_mime and detected_mime.startswith("text/"):
        return True, None

    allowed = ALLOWED_MIME_TYPES.get(EXT_TXT, [])
    if detected_mime in allowed:
        return True, None

    return False, TXT_CORRUPT_ERROR


def _validate_csv_content(uploaded_file, detected_mime: str):
    is_binary, binary_error = _has_binary_signature(uploaded_file)
    if is_binary:
        uploaded_file.seek(0)
        header = uploaded_file.read(8)
        uploaded_file.seek(0)
        if header == OLE_SIGNATURE:
            return False, CSV_PROTECTED_ERROR
        return False, binary_error

    if detected_mime and detected_mime.startswith("text/"):
        return True, None

    allowed = ALLOWED_MIME_TYPES.get(EXT_CSV, [])
    if detected_mime in allowed:
        return True, None

    return False, CSV_CORRUPT_ERROR


def save_temp_file(uploaded_file):
    os.makedirs(settings.UPLOAD_TEMP_DIR, exist_ok=True)

    original_name = os.path.basename(uploaded_file.name)
    safe_name = get_valid_filename(original_name)
    unique_name = f"{uuid4()}_{safe_name}"

    base_dir = os.path.abspath(settings.UPLOAD_TEMP_DIR)
    file_path = os.path.abspath(os.path.join(base_dir, unique_name))

    if not file_path.startswith(base_dir):
        raise ValueError("Invalid file path detected.")

    with open(file_path, "wb+") as destination:
        for chunk in uploaded_file.chunks():
            destination.write(chunk)

    return file_path
