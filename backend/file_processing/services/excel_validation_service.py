"""Excel-specific upload validation."""

import logging

from file_processing.services.mime_validation_service import (
    _is_legacy_xls_content,
    _is_ole_container,
)
from file_processing.services.upload_file_types import (
    EXT_XLS,
    EXT_XLSX,
    EXCEL_CORRUPT_ERROR,
    EXCEL_TOO_MANY_SHEETS_ERROR,
    MAX_EXCEL_SHEETS,
)

logger = logging.getLogger(__name__)


def validate_excel_sheet_count(
    uploaded_file,
    ext,
    *,
    is_ole_container_func=None,
    is_legacy_xls_content_func=None,
):
    try:
        if _should_parse_as_xls(
            uploaded_file,
            ext,
            is_ole_container_func=is_ole_container_func,
            is_legacy_xls_content_func=is_legacy_xls_content_func,
        ):
            sheet_count = _get_xls_sheet_count(uploaded_file)
        else:
            sheet_count = _get_xlsx_sheet_count(uploaded_file)
    except Exception:
        logger.exception("Failed to validate Excel sheet count.")
        return False, EXCEL_CORRUPT_ERROR

    return check_excel_sheet_count(sheet_count)


def check_excel_sheet_count(sheet_count):
    if sheet_count > MAX_EXCEL_SHEETS:
        return False, EXCEL_TOO_MANY_SHEETS_ERROR
    return True, None


def _should_parse_as_xls(
    uploaded_file,
    ext,
    *,
    is_ole_container_func=None,
    is_legacy_xls_content_func=None,
):
    is_ole_container_func = is_ole_container_func or _is_ole_container
    is_legacy_xls_content_func = (
        is_legacy_xls_content_func or _is_legacy_xls_content
    )

    if ext == EXT_XLS:
        return True

    if ext == EXT_XLSX and is_ole_container_func(uploaded_file):
        return is_legacy_xls_content_func(uploaded_file)

    return False


def _get_xlsx_sheet_count(uploaded_file):
    from openpyxl import load_workbook

    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    try:
        return len(workbook.sheetnames)
    finally:
        workbook.close()
        uploaded_file.seek(0)


def _get_xls_sheet_count(uploaded_file):
    import xlrd

    uploaded_file.seek(0)
    workbook_bytes = uploaded_file.read()
    uploaded_file.seek(0)

    workbook = xlrd.open_workbook(file_contents=workbook_bytes, on_demand=True)
    try:
        return workbook.nsheets
    finally:
        release_resources = getattr(workbook, "release_resources", None)
        if callable(release_resources):
            release_resources()
        uploaded_file.seek(0)

