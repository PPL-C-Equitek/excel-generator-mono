import csv
import io
import os
import re
import uuid
import zipfile
from datetime import datetime, timezone


class OutputLLMValidationError(Exception):
    """Raised when LLM output does not satisfy CSV validation contract."""


class OutputCSVMappingError(Exception):
    """Raised when validated LLM output cannot be mapped into CSV tabular structure."""


class OutputCSVGenerationError(Exception):
    """Raised when mapped output cannot be generated into CSV content."""


class OutputExcelGenerationError(Exception):
    """Raised when mapped output cannot be generated into Excel content."""


class OutputCSVDownloadLookupError(Exception):
    """Raised when generated CSV artifact cannot be resolved for download."""


class OutputExcelDownloadLookupError(Exception):
    """Raised when generated Excel artifact cannot be resolved for download."""


class CSVSanitizationPolicy:
    """Strategy extension point for CSV header/value sanitization."""

    def sanitize_header(self, header):
        return _sanitize_csv_value(header)

    def sanitize_value(self, value):
        return _sanitize_csv_value(value)


class CSVFileNamePolicy:
    """Strategy extension point for CSV file naming."""

    def build_filename(self, sheet_name):
        return f"{sheet_name}.csv"


_SCALAR_TYPES = (str, int, float, bool, type(None))
_ALLOWED_SOURCE_TYPES = {"Excel", "PDF"}
_REQUIRED_TOP_LEVEL_KEYS = {"document_info", "summary", "content_data"}
_CSV_FORMULA_PREFIXES = ("=", "+", "-", "@")
_EXCEL_SHEET_INVALID_CHARS = re.compile(r"[\\/*?:\[\]]")
_EXCEL_ARTIFACT_TYPE = "xlsx"
_EXCEL_FILE_ID_PREFIX = "xlsx_"
_EXCEL_FILE_NAME_PREFIX = "export_"
_EXCEL_FILE_EXTENSION = "xlsx"
_EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_EXCEL_DOWNLOAD_FILE_NAME_PATTERN = r"export_[a-z0-9]+\.xlsx"
_EXCEL_MAX_SHEET_NAME_LENGTH = 31
_EXCEL_DEFAULT_EMPTY_SHEET_NAME = "Sheet1"
_DEFAULT_CSV_SANITIZATION_POLICY = CSVSanitizationPolicy()
_DEFAULT_CSV_FILENAME_POLICY = CSVFileNamePolicy()


def validate_output_llm(output_json):
    if not isinstance(output_json, dict):
        raise OutputLLMValidationError("output_json root must be an object.")

    _validate_top_level(output_json)
    _validate_document_info(output_json["document_info"])
    _validate_summary(output_json["summary"])
    _validate_content_data(output_json["content_data"])

    return output_json


def _validate_top_level(payload):
    missing_keys = [key for key in _REQUIRED_TOP_LEVEL_KEYS if key not in payload]
    if missing_keys:
        raise OutputLLMValidationError(
            f"Missing required top-level keys: {missing_keys}."
        )

    if not isinstance(payload["document_info"], dict):
        raise OutputLLMValidationError("document_info must be an object.")
    if not isinstance(payload["summary"], dict):
        raise OutputLLMValidationError("summary must be an object.")
    if not isinstance(payload["content_data"], list):
        raise OutputLLMValidationError("content_data must be a list.")


def _validate_document_info(document_info):
    required_keys = ("source_type", "filename")
    for key in required_keys:
        if key not in document_info:
            raise OutputLLMValidationError(
                f"document_info is missing required key '{key}'."
            )

    source_type = document_info["source_type"]
    if not isinstance(source_type, str) or source_type not in _ALLOWED_SOURCE_TYPES:
        raise OutputLLMValidationError("document_info.source_type must be Excel or PDF.")

    filename = document_info["filename"]
    if not isinstance(filename, str) or not filename.strip():
        raise OutputLLMValidationError(
            "document_info.filename must be a non-empty string."
        )


def _validate_summary(summary):
    for key, value in summary.items():
        if not isinstance(key, str) or not key.strip():
            raise OutputLLMValidationError("summary keys must be non-empty strings.")
        if isinstance(value, (dict, list)):
            raise OutputLLMValidationError(
                f"summary['{key}'] must be a scalar value."
            )
        if not isinstance(value, _SCALAR_TYPES):
            raise OutputLLMValidationError(
                f"summary['{key}'] has unsupported value type."
            )

def _validate_content_data(content_data):
    table_names = set()
    for table_index, table in enumerate(content_data):
        if not isinstance(table, dict):
            raise OutputLLMValidationError(
                f"content_data[{table_index}] must be an object."
            )

        for required_key in ("table_name", "headers", "rows"):
            if required_key not in table:
                raise OutputLLMValidationError(
                    f"content_data[{table_index}] is missing required key '{required_key}'."
                )

        table_name = table["table_name"]
        if not isinstance(table_name, str) or not table_name.strip():
            raise OutputLLMValidationError(
                f"content_data[{table_index}].table_name must be a non-empty string."
            )
        if table_name in table_names:
            raise OutputLLMValidationError(
                f"Duplicate table_name found: '{table_name}'."
            )
        table_names.add(table_name)

        headers = _validate_headers(table["headers"], table_index)
        _validate_rows(table["rows"], headers, table_index)


def _validate_headers(headers, table_index):
    if not isinstance(headers, list):
        raise OutputLLMValidationError(
            f"content_data[{table_index}].headers must be a list."
        )
    if not headers:
        raise OutputLLMValidationError(
            f"content_data[{table_index}].headers must not be empty."
        )

    normalized_columns = []
    seen_columns = set()
    for header in headers:
        if not isinstance(header, str):
            raise OutputLLMValidationError(
                f"content_data[{table_index}] header names must be strings."
            )
        trimmed_column = header.strip()
        if not trimmed_column:
            raise OutputLLMValidationError(
                f"content_data[{table_index}] contains blank header name."
            )

        dedupe_key = trimmed_column.lower()
        if dedupe_key in seen_columns:
            raise OutputLLMValidationError(
                f"content_data[{table_index}] headers must be unique (case-insensitive)."
            )
        seen_columns.add(dedupe_key)
        normalized_columns.append(trimmed_column)

    return normalized_columns


def _validate_rows(rows, headers, table_index):
    if not isinstance(rows, list):
        raise OutputLLMValidationError(
            f"content_data[{table_index}].rows must be a list."
        )

    header_set = set(headers)
    for row_index, row in enumerate(rows):
        _validate_row_structure(
            row=row,
            headers=headers,
            header_set=header_set,
            table_index=table_index,
            row_index=row_index,
        )
        _validate_row_values(
            row=row,
            headers=headers,
            table_index=table_index,
            row_index=row_index,
        )


def _validate_row_structure(row, headers, header_set, table_index, row_index):
    if not isinstance(row, dict):
        raise OutputLLMValidationError(
            f"content_data[{table_index}], row {row_index} must be an object."
        )

    missing_columns = [header for header in headers if header not in row]
    if missing_columns:
        raise OutputLLMValidationError(
            f"content_data[{table_index}], row {row_index} is missing required headers: {missing_columns}."
        )

    unknown_columns = [key for key in row if key not in header_set]
    if unknown_columns:
        raise OutputLLMValidationError(
            f"content_data[{table_index}], row {row_index} has unknown headers: {unknown_columns}."
        )


def _validate_row_values(row, headers, table_index, row_index):
    for header in headers:
        value = row[header]
        if isinstance(value, (dict, list)):
            raise OutputLLMValidationError(
                f"content_data[{table_index}], row {row_index}, header '{header}' has unsupported nested value."
            )
        if not isinstance(value, _SCALAR_TYPES):
            raise OutputLLMValidationError(
                f"content_data[{table_index}], row {row_index}, header '{header}' has unsupported value type."
            )


def map_output_csv(validated_output):
    if not isinstance(validated_output, dict):
        raise OutputCSVMappingError("validated_output must be an object.")

    tables = validated_output.get("content_data")
    if not isinstance(tables, list):
        raise OutputCSVMappingError("validated_output.content_data must be a list.")

    mapped_sheets = []
    for table_index, table in enumerate(tables):
        if not isinstance(table, dict):
            raise OutputCSVMappingError(
                f"content_data[{table_index}] must be an object."
            )

        for required_key in ("table_name", "headers", "rows"):
            if required_key not in table:
                raise OutputCSVMappingError(
                    f"content_data[{table_index}] is missing required key '{required_key}'."
                )

        name = table["table_name"]
        headers = table["headers"]
        rows = table["rows"]

        if not isinstance(headers, list) or not headers:
            raise OutputCSVMappingError(
                f"content_data[{table_index}].headers must be a non-empty list."
            )
        if not isinstance(rows, list):
            raise OutputCSVMappingError(
                f"content_data[{table_index}].rows must be a list."
            )

        mapped_rows = _map_rows(headers=headers, rows=rows, table_index=table_index)
        mapped_sheets.append(
            {
                "name": name,
                "headers": headers,
                "rows": mapped_rows,
            }
        )

    return {
        "document_info": validated_output.get("document_info", {}),
        "summary": validated_output.get("summary", {}),
        "sheets": mapped_sheets,
    }


def _map_rows(headers, rows, table_index):
    mapped_rows = []
    header_set = set(headers)

    for row_index, row in enumerate(rows):
        if not isinstance(row, dict):
            raise OutputCSVMappingError(
                f"content_data[{table_index}], row {row_index} must be an object."
            )

        missing_headers = [header for header in headers if header not in row]
        if missing_headers:
            raise OutputCSVMappingError(
                f"content_data[{table_index}], row {row_index} is missing headers: {missing_headers}."
            )

        unknown_headers = [key for key in row.keys() if key not in header_set]
        if unknown_headers:
            raise OutputCSVMappingError(
                f"content_data[{table_index}], row {row_index} has unknown headers: {unknown_headers}."
            )

        mapped_rows.append([row[header] for header in headers])

    return mapped_rows


def generate_csv(mapped_output, sanitization_policy=None, filename_policy=None):
    if not isinstance(mapped_output, dict):
        raise OutputCSVGenerationError("mapped_output must be an object.")

    sheets = mapped_output.get("sheets")
    if not isinstance(sheets, list):
        raise OutputCSVGenerationError("mapped_output.sheets must be a list.")

    sanitization_policy = _resolve_sanitization_policy(
        sanitization_policy, OutputCSVGenerationError
    )
    filename_policy = _resolve_filename_policy(
        filename_policy, OutputCSVGenerationError
    )

    files = []
    for sheet_index, sheet in enumerate(sheets):
        files.append(
            _build_csv_file(
                sheet=sheet,
                sheet_index=sheet_index,
                sanitization_policy=sanitization_policy,
                filename_policy=filename_policy,
            )
        )

    return {"files": files}


def _build_csv_file(sheet, sheet_index, sanitization_policy, filename_policy):
    sheet_name, headers, rows = _validate_sheet_structure(
        sheet, sheet_index, OutputCSVGenerationError
    )

    _validate_sheet_headers(headers, sheet_index, OutputCSVGenerationError)
    normalized_headers = [sanitization_policy.sanitize_header(header) for header in headers]
    normalized_rows = _validate_sheet_rows(
        rows=rows,
        headers=headers,
        sheet_index=sheet_index,
        sanitization_policy=sanitization_policy,
        error_class=OutputCSVGenerationError,
    )
    filename = filename_policy.build_filename(sheet_name)
    if not isinstance(filename, str) or not filename.strip():
        raise OutputCSVGenerationError(
            "filename_policy.build_filename must return a non-empty string."
        )

    return {
        "name": filename,
        "content": _build_csv_content(normalized_headers, normalized_rows),
    }


def _validate_sheet_structure(sheet, sheet_index, error_class):
    if not isinstance(sheet, dict):
        raise error_class(f"Sheet {sheet_index} must be an object.")

    for required_key in ("name", "headers", "rows"):
        if required_key not in sheet:
            raise error_class(
                f"Sheet {sheet_index} is missing required key '{required_key}'."
            )

    sheet_name = sheet["name"]
    headers = sheet["headers"]
    rows = sheet["rows"]

    if not isinstance(sheet_name, str) or not sheet_name.strip():
        raise error_class(
            f"Sheet {sheet_index} name must be a non-empty string."
        )

    return sheet_name, headers, rows


def generate_csv_download_artifact(
    mapped_output,
    sanitization_policy=None,
    filename_policy=None,
):
    generated = generate_csv(
        mapped_output,
        sanitization_policy=sanitization_policy,
        filename_policy=filename_policy,
    )
    files = generated["files"]

    if len(files) == 1:
        single_file = files[0]
        return {
            "type": "csv",
            "name": single_file["name"],
            "content": single_file["content"].encode("utf-8"),
        }

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for file_item in files:
            archive.writestr(file_item["name"], file_item["content"])

    return {
        "type": "zip",
        "name": "csv_export.zip",
        "content": zip_buffer.getvalue(),
    }


def export_csv_to_filesystem(
    output_json,
    storage_dir,
    token_generator=None,
    now_provider=None,
    sanitization_policy=None,
    filename_policy=None,
):
    validated = validate_output_llm(output_json)
    mapped = map_output_csv(validated)
    artifact = generate_csv_download_artifact(
        mapped,
        sanitization_policy=sanitization_policy,
        filename_policy=filename_policy,
    )

    base_dir = _resolve_storage_dir(storage_dir, OutputCSVGenerationError)
    token = _resolve_export_token(token_generator, OutputCSVGenerationError)
    extension = "zip" if artifact["type"] == "zip" else "csv"
    file_name = f"export_{token}.{extension}"
    file_path = _build_safe_file_path(base_dir, file_name, OutputCSVGenerationError)

    try:
        with open(file_path, "wb") as destination:
            destination.write(artifact["content"])
    except OSError as exc:
        raise OutputCSVGenerationError("Failed to save generated CSV artifact.") from exc

    return {
        "file_id": f"csv_{token}",
        "file_name": file_name,
        "artifact_type": artifact["type"],
        "size_bytes": len(artifact["content"]),
        "created_at": _resolve_created_at(now_provider, OutputCSVGenerationError),
    }


def export_excel_to_filesystem(
    output_json,
    storage_dir,
    token_generator=None,
    now_provider=None,
    sanitization_policy=None,
):
    validated = validate_output_llm(output_json)
    mapped = map_output_csv(validated)
    artifact = _generate_excel_download_artifact(
        mapped,
        sanitization_policy=sanitization_policy,
    )

    token, file_name, file_path, created_at = _resolve_excel_export_context(
        storage_dir=storage_dir,
        token_generator=token_generator,
        now_provider=now_provider,
    )

    try:
        with open(file_path, "wb") as destination:
            destination.write(artifact["content"])
    except OSError as exc:
        raise OutputExcelGenerationError(
            "Failed to save generated Excel artifact."
        ) from exc

    return {
        "file_id": f"{_EXCEL_FILE_ID_PREFIX}{token}",
        "file_name": file_name,
        "artifact_type": _EXCEL_ARTIFACT_TYPE,
        "size_bytes": len(artifact["content"]),
        "created_at": created_at,
    }


def _generate_excel_download_artifact(mapped_output, sanitization_policy=None):
    sheets = _validate_excel_mapped_output(mapped_output)
    sanitization_policy = _resolve_excel_sanitization_policy(sanitization_policy)
    workbook = _build_excel_workbook(
        sheets=sheets,
        sanitization_policy=sanitization_policy,
    )

    content = _serialize_excel_workbook(workbook)

    return {
        "type": _EXCEL_ARTIFACT_TYPE,
        "name": "excel_export.xlsx",
        "content": content,
    }


def _validate_excel_mapped_output(mapped_output):
    if not isinstance(mapped_output, dict):
        raise OutputExcelGenerationError("mapped_output must be an object.")

    sheets = mapped_output.get("sheets")
    if not isinstance(sheets, list):
        raise OutputExcelGenerationError("mapped_output.sheets must be a list.")

    return sheets


def _resolve_excel_sanitization_policy(sanitization_policy):
    return _resolve_sanitization_policy(
        sanitization_policy, OutputExcelGenerationError
    )


def _resolve_excel_export_context(storage_dir, token_generator, now_provider):
    base_dir = _resolve_storage_dir(storage_dir, OutputExcelGenerationError)
    token = _resolve_export_token(token_generator, OutputExcelGenerationError)
    file_name = f"{_EXCEL_FILE_NAME_PREFIX}{token}.{_EXCEL_FILE_EXTENSION}"
    file_path = _build_safe_file_path(base_dir, file_name, OutputExcelGenerationError)
    created_at = _resolve_created_at(now_provider, OutputExcelGenerationError)

    return token, file_name, file_path, created_at


def _serialize_excel_workbook(workbook):
    try:
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
    except Exception as exc:
        raise OutputExcelGenerationError("Failed to generate Excel artifact.") from exc
    finally:
        workbook.close()


def _build_excel_workbook(sheets, sanitization_policy):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise OutputExcelGenerationError(
            "openpyxl is required to generate Excel artifacts."
        ) from exc

    workbook = Workbook()
    if workbook.worksheets:
        workbook.remove(workbook.active)

    sheet_names = set()
    for sheet_index, sheet in enumerate(sheets):
        _append_excel_worksheet(
            workbook=workbook,
            sheet=sheet,
            sheet_index=sheet_index,
            sanitization_policy=sanitization_policy,
            sheet_names=sheet_names,
        )

    if not workbook.worksheets:
        workbook.create_sheet(title=_EXCEL_DEFAULT_EMPTY_SHEET_NAME)

    return workbook


def _append_excel_worksheet(
    workbook,
    sheet,
    sheet_index,
    sanitization_policy,
    sheet_names,
):
    sheet_name, headers, rows = _validate_excel_sheet(
        sheet=sheet,
        sheet_index=sheet_index,
        sanitization_policy=sanitization_policy,
    )
    worksheet = workbook.create_sheet(
        title=_normalize_excel_sheet_name(sheet_name, sheet_names)
    )
    _write_excel_worksheet_rows(worksheet=worksheet, headers=headers, rows=rows)


def _write_excel_worksheet_rows(worksheet, headers, rows):
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)


def _validate_excel_sheet(sheet, sheet_index, sanitization_policy):
    sheet_name, headers, rows = _validate_sheet_structure(
        sheet,
        sheet_index,
        OutputExcelGenerationError,
    )
    _validate_sheet_headers(headers, sheet_index, OutputExcelGenerationError)
    normalized_rows = _validate_sheet_rows(
        rows=rows,
        headers=headers,
        sheet_index=sheet_index,
        sanitization_policy=sanitization_policy,
        error_class=OutputExcelGenerationError,
    )

    normalized_headers = [
        sanitization_policy.sanitize_header(header) for header in headers
    ]
    return sheet_name, normalized_headers, normalized_rows


def _normalize_excel_sheet_name(sheet_name, seen_names):
    normalized = _EXCEL_SHEET_INVALID_CHARS.sub("_", sheet_name).strip()
    if not normalized:
        normalized = _EXCEL_DEFAULT_EMPTY_SHEET_NAME

    normalized = normalized[:_EXCEL_MAX_SHEET_NAME_LENGTH]
    candidate = normalized
    duplicate_index = 1

    while candidate.lower() in seen_names:
        suffix = f"_{duplicate_index}"
        candidate = (
            f"{normalized[: max(0, _EXCEL_MAX_SHEET_NAME_LENGTH - len(suffix))]}{suffix}"
        )
        duplicate_index += 1

    seen_names.add(candidate.lower())
    return candidate


def resolve_csv_download_artifact(file_id, storage_dir):
    token = _resolve_download_token(file_id)
    base_dir = _resolve_download_storage_dir(storage_dir)
    discovered_artifacts = _discover_download_artifacts(base_dir)

    candidate_files = (
        ("csv", "text/csv"),
        ("zip", "application/zip"),
    )
    for extension, content_type in candidate_files:
        file_name = f"export_{token}.{extension}"
        file_path = discovered_artifacts.get(file_name)
        if file_path:
            return {
                "file_name": file_name,
                "file_path": file_path,
                "artifact_type": extension,
                "content_type": content_type,
            }

    raise OutputCSVDownloadLookupError("CSV artifact not found for given file_id.")


def resolve_excel_download_artifact(export_id, storage_dir):
    token = _resolve_excel_download_token(export_id)
    base_dir = _resolve_lookup_storage_dir(
        storage_dir, OutputExcelDownloadLookupError
    )
    discovered_artifacts = _discover_excel_download_artifacts(base_dir)

    file_name = f"{_EXCEL_FILE_NAME_PREFIX}{token}.{_EXCEL_FILE_EXTENSION}"
    file_path = discovered_artifacts.get(file_name)
    if file_path:
        return {
            "file_name": file_name,
            "file_path": file_path,
            "artifact_type": _EXCEL_ARTIFACT_TYPE,
            "content_type": _EXCEL_CONTENT_TYPE,
        }

    raise OutputExcelDownloadLookupError("Excel artifact not found for given export_id.")


def _discover_download_artifacts(base_dir):
    discovered = {}
    try:
        with os.scandir(base_dir) as entries:
            for entry in entries:
                # Avoid following symlinks; only serve regular files under base_dir.
                if not entry.is_file(follow_symlinks=False):
                    continue

                entry_name = entry.name.lower()
                if not re.fullmatch(r"export_[a-z0-9]+\.(csv|zip)", entry_name):
                    continue

                entry_path = os.path.realpath(entry.path)
                try:
                    common_path = os.path.commonpath([base_dir, entry_path])
                except ValueError:
                    continue

                if common_path != base_dir:
                    continue

                discovered[entry_name] = entry_path
    except OSError as exc:
        raise OutputCSVDownloadLookupError("CSV artifact storage is unavailable.") from exc

    return discovered


def _discover_excel_download_artifacts(base_dir):
    discovered = {}
    try:
        with os.scandir(base_dir) as entries:
            for entry in entries:
                if not entry.is_file(follow_symlinks=False):
                    continue

                entry_name = entry.name.lower()
                if not re.fullmatch(_EXCEL_DOWNLOAD_FILE_NAME_PATTERN, entry_name):
                    continue

                entry_path = os.path.realpath(entry.path)
                try:
                    common_path = os.path.commonpath([base_dir, entry_path])
                except ValueError:
                    continue

                if common_path != base_dir:
                    continue

                discovered[entry_name] = entry_path
    except OSError as exc:
        raise OutputExcelDownloadLookupError(
            "Excel artifact storage is unavailable."
        ) from exc

    return discovered


def _validate_sheet_headers(headers, sheet_index, error_class):
    if not isinstance(headers, list):
        raise error_class(f"Sheet {sheet_index} headers must be a list.")
    if not headers:
        raise error_class(
            f"Sheet {sheet_index} headers must be a non-empty list."
        )

    seen_headers = set()
    for header in headers:
        if not isinstance(header, str):
            raise error_class(
                f"Sheet {sheet_index} header names must be strings."
            )
        normalized_header = header.strip()
        if not normalized_header:
            raise error_class(
                f"Sheet {sheet_index} contains blank header name."
            )

        dedupe_key = normalized_header.lower()
        if dedupe_key in seen_headers:
            raise error_class(
                f"Sheet {sheet_index} headers must be unique (case-insensitive)."
            )
        seen_headers.add(dedupe_key)


def _validate_sheet_rows(rows, headers, sheet_index, sanitization_policy, error_class):
    if not isinstance(rows, list):
        raise error_class(f"Sheet {sheet_index} rows must be a list.")

    normalized_rows = []
    for row_index, row in enumerate(rows):
        if not isinstance(row, list):
            raise error_class(
                f"Sheet {sheet_index}, row {row_index} must be a list."
            )
        if len(row) != len(headers):
            raise error_class(
                f"Sheet {sheet_index}, row {row_index} must have {len(headers)} values."
            )

        normalized_row = []
        for value in row:
            if isinstance(value, (dict, list)):
                raise error_class(
                    f"Sheet {sheet_index}, row {row_index} contains unsupported nested value."
                )
            if not isinstance(value, _SCALAR_TYPES):
                raise error_class(
                    f"Sheet {sheet_index}, row {row_index} contains unsupported value type."
                )
            normalized_row.append(sanitization_policy.sanitize_value(value))
        normalized_rows.append(normalized_row)

    return normalized_rows


def _resolve_sanitization_policy(sanitization_policy, error_class):
    return _resolve_policy_with_methods(
        policy=sanitization_policy,
        default_policy=_DEFAULT_CSV_SANITIZATION_POLICY,
        required_methods=("sanitize_header", "sanitize_value"),
        error_message=(
            "sanitization_policy must implement callable sanitize_header and "
            "sanitize_value methods."
        ),
        error_class=error_class,
    )


def _resolve_filename_policy(filename_policy, error_class):
    return _resolve_policy_with_methods(
        policy=filename_policy,
        default_policy=_DEFAULT_CSV_FILENAME_POLICY,
        required_methods=("build_filename",),
        error_message="filename_policy must implement callable build_filename method.",
        error_class=error_class,
    )


def _resolve_policy_with_methods(
    policy,
    default_policy,
    required_methods,
    error_message,
    error_class,
):
    if policy is None:
        return default_policy

    for method_name in required_methods:
        if not callable(getattr(policy, method_name, None)):
            raise error_class(error_message)

    return policy


def _resolve_storage_dir(storage_dir, error_class):
    if not isinstance(storage_dir, str) or not storage_dir.strip():
        raise error_class("storage_dir must be a non-empty string.")

    base_dir = os.path.abspath(storage_dir)
    try:
        os.makedirs(base_dir, exist_ok=True)
    except OSError as exc:
        raise error_class("storage_dir cannot be created.") from exc

    return base_dir


def _resolve_download_storage_dir(storage_dir):
    return _resolve_lookup_storage_dir(storage_dir, OutputCSVDownloadLookupError)


def _resolve_lookup_storage_dir(storage_dir, error_class):
    if not isinstance(storage_dir, str) or not storage_dir.strip():
        raise error_class("storage_dir must be a non-empty string.")

    return os.path.realpath(os.path.abspath(storage_dir))


def _resolve_export_token(token_generator, error_class):
    if token_generator is None:
        token = uuid.uuid4().hex
    else:
        token = token_generator()

    if not isinstance(token, str) or not token.strip():
        raise error_class(
            "token_generator must return a non-empty string token."
        )

    normalized_token = token.strip().lower()
    if not re.fullmatch(r"[a-z0-9]+", normalized_token):
        raise error_class(
            "token_generator returned an unsafe token format."
        )

    return normalized_token


def _resolve_download_token(file_id):
    if not isinstance(file_id, str) or not file_id.strip():
        raise OutputCSVDownloadLookupError("file_id must be a non-empty string.")

    normalized_file_id = file_id.strip().lower()
    match = re.fullmatch(r"csv_([a-z0-9]+)", normalized_file_id)
    if not match:
        raise OutputCSVDownloadLookupError("file_id format is invalid.")

    return match.group(1)


def _resolve_excel_download_token(export_id):
    if not isinstance(export_id, str) or not export_id.strip():
        raise OutputExcelDownloadLookupError("export_id must be a non-empty string.")

    normalized_export_id = export_id.strip().lower()
    match = re.fullmatch(r"xlsx_([a-z0-9]+)", normalized_export_id)
    if not match:
        raise OutputExcelDownloadLookupError("export_id format is invalid.")

    return match.group(1)


def _build_safe_file_path(base_dir, file_name, error_class):
    base_dir_real = os.path.realpath(os.path.abspath(base_dir))
    candidate = os.path.realpath(os.path.join(base_dir_real, file_name))

    try:
        common_path = os.path.commonpath([base_dir_real, candidate])
    except ValueError as exc:
        raise error_class("Invalid storage path detected.") from exc

    if common_path != base_dir_real:
        raise error_class("Invalid storage path detected.")

    return candidate


def _resolve_created_at(now_provider, error_class):
    if now_provider is None:
        now_value = datetime.now(timezone.utc)
    else:
        now_value = now_provider()

    if isinstance(now_value, datetime):
        return now_value.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")

    if isinstance(now_value, str) and now_value.strip():
        return now_value

    raise error_class(
        "now_provider must return datetime or non-empty string."
    )


def _sanitize_csv_value(value):
    if not isinstance(value, str):
        return value

    stripped_value = value.lstrip()
    if not stripped_value:
        return value

    if stripped_value.startswith("'"):
        return value

    if stripped_value[0] in _CSV_FORMULA_PREFIXES:
        return f"'{value}"

    return value


def _build_csv_content(headers, rows):
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue()
