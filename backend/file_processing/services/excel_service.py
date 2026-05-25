from __future__ import annotations

import os
import logging
from typing import Any, IO

logger = logging.getLogger(__name__)
OLE_SIGNATURE = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"


def _load_workbook(file_or_path: str | IO[bytes] | Any):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is not installed. Run: pip install openpyxl"
        ) from exc

    if isinstance(file_or_path, str) and not os.path.exists(file_or_path):
        raise FileNotFoundError(f"File not found: {file_or_path}")

    try:
        return load_workbook(file_or_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(
            f"Excel file is corrupted or cannot be read: {exc}"
        ) from exc


def _has_ole_signature(file_or_path: str | IO[bytes] | Any) -> bool:
    if isinstance(file_or_path, str):
        try:
            with open(file_or_path, "rb") as fh:
                return fh.read(len(OLE_SIGNATURE)) == OLE_SIGNATURE
        except Exception:
            return False
    else:
        try:
            try:
                file_or_path.seek(0)
            except (AttributeError, OSError):
                pass
            sig = file_or_path.read(len(OLE_SIGNATURE))
            try:
                file_or_path.seek(0)
            except (AttributeError, OSError):
                pass
            return sig == OLE_SIGNATURE
        except Exception:
            return False


def _normalize_row(row_list: list[Any]) -> list[Any]:
    for i, val in enumerate(row_list):
        if isinstance(val, float) and val == int(val):
            row_list[i] = str(int(val))
        elif val == "":
            row_list[i] = None
    return row_list


def _trim_trailing_nulls(row_list: list[Any]) -> tuple[list[Any], int]:
    null_count = 0
    while row_list and (row_list[-1] is None or row_list[-1] == ""):
        row_list.pop()
        null_count += 1
    return row_list, null_count


def _append_null_marker(row_list: list[Any], null_count: int) -> None:
    if null_count > 1:
        row_list.append(f"nullx{null_count}")
    elif null_count == 1:
        row_list.append("null")


def _process_row(row_list: list[Any], total_cols: int) -> list[Any] | None:
    row_list, null_count = _trim_trailing_nulls(row_list)

    if not row_list and null_count == total_cols:
        return None

    _append_null_marker(row_list, null_count)
    return row_list


def _sheet_to_rows(ws) -> list[list[Any]]:
    raw_rows = list(ws.iter_rows(values_only=True))

    if not raw_rows:
        return []

    data_rows: list[list[Any]] = []
    for row in raw_rows:
        row_list = list(row)
        processed = _process_row(row_list, len(row))
        if processed is not None:
            data_rows.append(processed)

    return data_rows


def _parse_excel_calamine(file_or_path: str | IO[bytes] | Any) -> dict[str, list[list[Any]]]:
    try:
        from python_calamine import CalamineWorkbook
    except ImportError as exc:
        raise ImportError("python-calamine is not installed. Run: pip install python-calamine") from exc

    if isinstance(file_or_path, str):
        if not os.path.exists(file_or_path):
            raise FileNotFoundError(f"File not found: {file_or_path}")
        wb = CalamineWorkbook.from_path(file_or_path)
    else:
        try:
            file_or_path.seek(0)
        except (AttributeError, OSError):
            pass
        wb = CalamineWorkbook.from_filelike(file_or_path)

    result: dict[str, list[list[Any]]] = {}
    for sheet_name in wb.sheet_names:
        ws = wb.get_sheet_by_name(sheet_name)
        data_rows: list[list[Any]] = []
        for row in ws.to_python(skip_empty_area=False):
            row_list = [None if v == "" else v for v in row]
            processed = _process_row(row_list, len(row_list))
            if processed is not None:
                data_rows.append(processed)
        result[sheet_name] = data_rows
    return result


def parse_excel(file_or_path: str | IO[bytes] | Any) -> dict[str, list[list[Any]]]:
    try:
        return _parse_excel_calamine(file_or_path)
    except ImportError:
        pass

    wb = _load_workbook(file_or_path)

    result: dict[str, list[list[Any]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result[sheet_name] = _sheet_to_rows(ws)

    wb.close()

    return result


def _load_xls_workbook(file_or_path: str | IO[bytes] | Any):
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError(
            "xlrd is not installed. Run: pip install xlrd"
        ) from exc

    if isinstance(file_or_path, str):
        if not os.path.exists(file_or_path):
            raise FileNotFoundError(f"File not found: {file_or_path}")
        try:
            return xlrd.open_workbook(file_or_path)
        except Exception as exc:
            raise ValueError(
                f"Excel (.xls) file is corrupted or cannot be read: {exc}"
            ) from exc
    else:
        try:
            try:
                file_or_path.seek(0)
            except (AttributeError, OSError):
                pass
            content = file_or_path.read()
            try:
                file_or_path.seek(0)
            except (AttributeError, OSError):
                pass
            return xlrd.open_workbook(file_contents=content)
        except Exception as exc:
            raise ValueError(
                f"Excel (.xls) file is corrupted or cannot be read: {exc}"
            ) from exc


def _xls_sheet_to_rows(sheet) -> list[list[Any]]:
    data_rows: list[list[Any]] = []

    for row_idx in range(sheet.nrows):
        row_list = [
            sheet.cell_value(row_idx, col_idx)
            for col_idx in range(sheet.ncols)
        ]

        row_list = _normalize_row(row_list)
        processed = _process_row(row_list, sheet.ncols)
        if processed is not None:
            data_rows.append(processed)

    return data_rows


def parse_xls(file_or_path: str | IO[bytes] | Any) -> dict[str, list[list[Any]]]:
    wb = _load_xls_workbook(file_or_path)

    result: dict[str, list[list[Any]]] = {}

    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        result[sheet_name] = _xls_sheet_to_rows(ws)

    return result


def process_uploaded_excel(
    file_or_path: str | IO[bytes] | Any,
) -> tuple[bool, str | None, dict[str, list[list[Any]]] | None]:
    try:
        ext = ""
        if isinstance(file_or_path, str):
            ext = os.path.splitext(file_or_path)[1].lower()
        else:
            name = getattr(file_or_path, "name", None)
            if name and isinstance(name, str):
                ext = os.path.splitext(name)[1].lower()

        is_xls = ext == ".xls" or (ext == ".xlsx" and _has_ole_signature(file_or_path))

        if is_xls:
            data = parse_xls(file_or_path)
        else:
            data = parse_excel(file_or_path)

    except FileNotFoundError as exc:
        return False, str(exc), None

    except Exception:
        logger.exception("Excel parsing failed")
        return False, "Invalid or corrupted Excel file.", None

    return True, None, data
