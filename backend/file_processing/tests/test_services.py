import os
import tempfile
import builtins
import zipfile
import unittest

from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from unittest.mock import patch, MagicMock, PropertyMock
from PyPDF2.errors import PdfReadError
from openpyxl import Workbook

from file_processing.services.ocr_service import OCRService
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from file_processing.services.non_ocr_pdf_service import NonOCRPDFService
from file_processing.services import upload_service
from file_processing.services.word_extraction_service import WordExtractionService

from file_processing.services import word_validation_service


from file_processing.services.upload_service import (
    _has_extracted_text,
    _get_empty_page_numbers,
    _has_zip_signature,
    _is_legacy_xls_content,
    _is_ole_container,
    validate_file,
    validate_mime_type,
    validate_pdf,
)
from file_processing.utils.upload_constants import FILE_TOO_LARGE_ERROR, MAX_FILE_SIZE


class TestOCRService(TestCase):
    @patch("file_processing.services.ocr_service.PdfReader")
    def test_text_pdf_path(self, mock_reader):
        mock_page = MagicMock()
        mock_page.extract_text.return_value = (
            "Hello world\nSecond line\n"
            "This is a long enough text to pass the OCR threshold detection."
        )

        mock_reader.return_value.pages = [mock_page]

        result = OCRService.process_pdf("dummy.pdf")

        self.assertIn("content", result)
        self.assertEqual(result["content"][0]["page"], 1)
        self.assertEqual(len(result["content"][0]["text"]), 3)

    @patch("file_processing.services.ocr_service.OCRService._ocr_single_image")
    @patch("file_processing.services.ocr_service.PdfOcrExtractor")
    @patch("file_processing.services.ocr_service.PdfReader")
    def test_scanned_pdf_uses_ocr(
        self, mock_reader, mock_extractor_cls, mock_ocr_single
    ):
        mock_page = MagicMock()
        mock_page.extract_text.return_value = ""

        mock_reader.return_value.pages = [mock_page]

        mock_image = MagicMock()
        mock_extractor_cls.return_value.convert_pages.return_value = [(1, mock_image)]

        mock_ocr_single.return_value = "OCR line one\nOCR line two"

        result = OCRService.process_pdf("dummy.pdf")

        self.assertEqual(result["content"][0]["page"], 1)
        self.assertGreater(len(result["content"][0]["text"]), 0)
        mock_ocr_single.assert_called_once()

    @patch("file_processing.services.ocr_service.PdfReader")
    def test_exception_wrapping(self, mock_reader):

        mock_reader.side_effect = Exception("boom")

        with self.assertRaises(ValueError) as context:
            OCRService.process_pdf("bad.pdf")

        self.assertIn("OCRService failed", str(context.exception))

    def test_split_lines_basic(self):
        text = "Hello world. How are you?\nFine.\nRp 1.000.000"

        result = OCRService.split_lines(text)

        self.assertEqual(result, ["Hello world. How are you?", "Fine.", "Rp 1.000.000"])

    @patch("file_processing.services.ocr_service.OCRService._ocr_single_image")
    @patch("file_processing.services.ocr_service.PdfOcrExtractor")
    def test_process_pdf_pages_success(self, mock_extractor_cls, mock_ocr_single):

        mock_image = MagicMock()
        mock_extractor_cls.return_value.convert_pages.return_value = [(1, mock_image)]

        mock_ocr_single.return_value = "Hello OCR from single image"

        result = OCRService.process_pdf_pages("/tmp/file.pdf", [1])

        self.assertEqual(result["content"][0]["page"], 1)
        self.assertEqual(result["content"][0]["text"], ["Hello OCR from single image"])

    @patch("file_processing.services.ocr_service.PdfOcrExtractor")
    def test_process_pdf_pages_no_images(self, mock_extractor_cls):

        mock_extractor_cls.return_value.convert_pages.return_value = []

        result = OCRService.process_pdf_pages("/tmp/file.pdf", [1])

        self.assertEqual(result["content"], [])

    @patch("file_processing.services.ocr_service.TesseractEngine")
    def test_ocr_single_image_success(self, mock_tesseract):
        mock_engine = MagicMock()
        mock_engine.extract_text_with_confidence.return_value = ("OCR text", 60.0)
        mock_tesseract.return_value = mock_engine

        text = OCRService._ocr_single_image("mock_image")

        self.assertEqual(text, "OCR text")


    @patch("file_processing.services.ocr_service.OCRService._ocr_single_image")
    def test_process_image_success(self, mock_ocr_single):
        mock_ocr_single.return_value = "Standalone image OCR"

        result = OCRService.process_image("mock_image")

        self.assertEqual(result["content"][0]["page"], 1)
        self.assertEqual(result["content"][0]["text"], ["Standalone image OCR"])

    @patch("file_processing.services.ocr_service.PdfOcrExtractor")
    def test_process_pdf_pages_exception(self, mock_extractor_cls):

        mock_extractor_cls.return_value.convert_pages.side_effect = Exception("fail")

        with self.assertRaises(ValueError):
            OCRService.process_pdf_pages("/tmp/file.pdf", [1])

    @patch("file_processing.services.ocr_service.PdfReader")
    def test_process_pdf_text_based(self, mock_reader):

        mock_page = MagicMock()
        mock_page.extract_text.return_value = "Hello world " * 20

        mock_reader.return_value.pages = [mock_page]

        result = OCRService.process_pdf("/tmp/file.pdf")

        self.assertEqual(result["content"][0]["page"], 1)

    @patch("file_processing.services.ocr_service.OCRService._ocr_single_image")
    @patch("file_processing.services.ocr_service.PdfOcrExtractor")
    @patch("file_processing.services.ocr_service.PdfReader")
    def test_process_pdf_ocr_branch(
        self, mock_reader, mock_extractor_cls, mock_ocr_single
    ):

        mock_page = MagicMock()
        mock_page.extract_text.return_value = ""

        mock_reader.return_value.pages = [mock_page]

        mock_image = MagicMock()
        mock_extractor_cls.return_value.convert_pages.return_value = [(1, mock_image)]

        mock_ocr_single.return_value = "OCR TEXT"

        result = OCRService.process_pdf("/tmp/file.pdf")

        self.assertEqual(result["content"][0]["text"], ["OCR TEXT"])

    @patch("file_processing.services.ocr_service.TesseractEngine")
    def test_ocr_single_image_import_error(self, mock_tesseract):
        mock_tesseract.side_effect = ImportError()

        with self.assertRaises(ValueError) as context:
            OCRService._ocr_single_image("img")

        self.assertIn("OCR dependency error", str(context.exception))

    @patch("file_processing.services.ocr_service.TesseractEngine")
    def test_ocr_single_image_generic_exception(self, mock_tesseract):
        mock_engine = MagicMock()
        mock_engine.extract_text_with_confidence.side_effect = Exception("boom")

        mock_tesseract.return_value = mock_engine

        with self.assertRaises(ValueError) as context:
            OCRService._ocr_single_image("img")

        self.assertIn("OCR runtime error", str(context.exception))


class TestNonOCRPDFService(TestCase):
    """Tests covering extract_pdf_to_json."""

    def _create_pdf(self, texts: list[str]) -> str:
        """Create a real PDF with one page per text entry and return its path."""
        buf = BytesIO()
        p = canvas.Canvas(buf)
        for text in texts:
            p.drawString(72, 720, text)
            p.showPage()
        p.save()
        buf.seek(0)

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        tmp.write(buf.read())
        tmp.close()
        return tmp.name

    def _create_blank_pdf(self) -> str:
        """Create a PDF whose page returns empty text."""
        buf = BytesIO()
        p = canvas.Canvas(buf)
        p.showPage()
        p.save()
        buf.seek(0)

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        tmp.write(buf.read())
        tmp.close()
        return tmp.name

    def _create_table_pdf(self) -> str:
        """Create a PDF containing a table (triggers table-extraction path)."""
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        tmp.close()

        doc = SimpleDocTemplate(tmp.name, pagesize=A4)
        data = [
            ["NAMA", "NPM", "SEMESTER"],
            ["Alice", "123", "6"],
            ["Bob", "456", "4"],
        ]
        table = Table(data)
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        doc.build([table])
        return tmp.name

    def _create_table_with_none_cell_pdf(self) -> str:
        """Create a table PDF where a cell is empty (None → '')."""
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        tmp.close()

        doc = SimpleDocTemplate(tmp.name, pagesize=A4)
        data = [
            ["COL1", "COL2"],
            ["val", ""],
        ]
        table = Table(data)
        table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 1, colors.black)]))
        doc.build([table])
        return tmp.name

    def _create_table_plus_text_pdf(self) -> str:
        """Create a PDF with a table AND plain text below it."""
        from reportlab.platypus import Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        tmp.close()

        doc = SimpleDocTemplate(tmp.name, pagesize=A4)
        styles = getSampleStyleSheet()

        data = [
            ["NAMA", "NPM"],
            ["Alice", "123"],
        ]
        table = Table(data)
        table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 1, colors.black)]))

        text_paragraph = Paragraph(
            "Please ignore this. THIS IS ONLY FOR TESTING.", styles["Normal"]
        )

        doc.build([table, Spacer(1, 20), text_paragraph])
        return tmp.name

    def test_single_page_text_returns_lines(self):
        """Plain-text page → text is a list of line strings."""
        path = self._create_pdf(["Hello World"])
        try:
            result = NonOCRPDFService.extract_non_ocr_pdf_to_json(path)
            self.assertEqual(len(result["content"]), 1)
            self.assertEqual(result["content"][0]["page"], 1)

            text = result["content"][0]["text"]
            self.assertIsInstance(text, list)
            joined = " ".join(text) if isinstance(text[0], str) else str(text)
            self.assertIn("Hello World", joined)
        finally:
            os.unlink(path)

    def test_multi_page_pdf(self):
        """Covers loop iteration for multiple pages."""
        path = self._create_pdf(["Page one", "Page two", "Page three"])
        try:
            result = NonOCRPDFService.extract_non_ocr_pdf_to_json(path)
            self.assertEqual(len(result["content"]), 3)
            for i, entry in enumerate(result["content"], start=1):
                self.assertEqual(entry["page"], i)
                self.assertIsInstance(entry["text"], list)
        finally:
            os.unlink(path)

    def test_blank_page_returns_empty_list(self):
        """Covers the branch when page has no text and no tables."""
        path = self._create_blank_pdf()
        try:
            result = NonOCRPDFService.extract_non_ocr_pdf_to_json(path)
            self.assertEqual(result["content"][0]["text"], [])
        finally:
            os.unlink(path)

    def test_table_pdf_returns_rows(self):
        """Pages with tables → text is list of row-arrays."""
        path = self._create_table_pdf()
        try:
            result = NonOCRPDFService.extract_non_ocr_pdf_to_json(path)
            text = result["content"][0]["text"]
            self.assertIsInstance(text, list)
            self.assertGreaterEqual(len(text), 3)
            for row in text:
                self.assertIsInstance(row, list)
            self.assertIn("NAMA", text[0])
            self.assertIn("NPM", text[0])
        finally:
            os.unlink(path)

    def test_table_none_cell_replaced_with_empty_string(self):
        """None cells in tables are replaced with empty strings."""
        path = self._create_table_with_none_cell_pdf()
        try:
            result = NonOCRPDFService.extract_non_ocr_pdf_to_json(path)
            text = result["content"][0]["text"]
            for row in text:
                if isinstance(row, list):
                    for cell in row:
                        self.assertIsNotNone(cell)
                        self.assertIsInstance(cell, str)
        finally:
            os.unlink(path)

    def test_table_plus_outside_text(self):
        """Text outside table area is also captured as plain strings."""
        path = self._create_table_plus_text_pdf()
        try:
            result = NonOCRPDFService.extract_non_ocr_pdf_to_json(path)
            text = result["content"][0]["text"]

            table_rows = [e for e in text if isinstance(e, list)]
            plain_lines = [e for e in text if isinstance(e, str)]

            self.assertGreaterEqual(len(table_rows), 2)
            self.assertGreaterEqual(len(plain_lines), 1)
            combined = " ".join(plain_lines)
            self.assertIn("TESTING", combined)
        finally:
            os.unlink(path)

    def test_corrupt_file_raises_exception(self):
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        tmp.write(b"not a pdf at all")
        tmp.close()
        try:
            with self.assertRaises(Exception):
                NonOCRPDFService.extract_non_ocr_pdf_to_json(tmp.name)
        finally:
            os.unlink(tmp.name)

    def test_nonexistent_file_raises_exception(self):
        with self.assertRaises(Exception):
            NonOCRPDFService.extract_non_ocr_pdf_to_json(
                "/tmp/nonexistent_file_12345.pdf"
            )

    def test_return_structure_keys(self):
        path = self._create_pdf(["structure test"])
        try:
            result = NonOCRPDFService.extract_non_ocr_pdf_to_json(path)
            self.assertIn("content", result)
            self.assertIsInstance(result["content"], list)
        finally:
            os.unlink(path)

    def test_content_entry_keys(self):
        path = self._create_pdf(["key test"])
        try:
            result = NonOCRPDFService.extract_non_ocr_pdf_to_json(path)
            for entry in result["content"]:
                self.assertIn("page", entry)
                self.assertIn("text", entry)
        finally:
            os.unlink(path)


class TestUploadService(TestCase):
    def _post_file(self, name, content, content_type):
        f = SimpleUploadedFile(name, content, content_type=content_type)
        return self.client.post("/upload/", {"file": f}, format="multipart")

    def generate_valid_pdf_bytes(self):
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        p.drawString(100, 750, "Hello PDF")
        p.save()
        buffer.seek(0)
        return buffer.read()

    def generate_valid_xlsx_bytes(self):
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Hello"
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def generate_valid_xls_bytes(self):
        import xlwt

        buffer = BytesIO()
        wb = xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        ws.write(0, 0, "Hello XLS")
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def test_is_ole_container_returns_true_for_ole_signature(self):
        xls_content = self.generate_valid_xls_bytes()
        file_obj = SimpleUploadedFile(
            "legacy.xls",
            xls_content,
            content_type="application/vnd.ms-excel",
        )

        self.assertTrue(_is_ole_container(file_obj))

    def test_is_ole_container_returns_false_on_seek_error(self):
        class BrokenFile:
            def seek(self, *_args, **_kwargs):
                raise OSError("seek failed")

        self.assertFalse(_is_ole_container(BrokenFile()))

    def test_is_legacy_xls_content_returns_true_for_valid_xls(self):
        xls_content = self.generate_valid_xls_bytes()
        file_obj = SimpleUploadedFile(
            "legacy.xls",
            xls_content,
            content_type="application/vnd.ms-excel",
        )

        self.assertTrue(_is_legacy_xls_content(file_obj))

    def test_is_legacy_xls_content_returns_false_for_non_xls_payload(self):
        file_obj = SimpleUploadedFile(
            "fake.xlsx",
            b"not an xls payload",
            content_type="application/octet-stream",
        )

        self.assertFalse(_is_legacy_xls_content(file_obj))

    def test_is_legacy_xls_content_returns_false_when_xlrd_unavailable(self):
        file_obj = SimpleUploadedFile(
            "legacy.xls",
            self.generate_valid_xls_bytes(),
            content_type="application/vnd.ms-excel",
        )
        real_import = builtins.__import__

        def fake_import(name, *args, **kwargs):
            if name == "xlrd":
                raise ImportError("xlrd unavailable")
            return real_import(name, *args, **kwargs)

        with patch("builtins.__import__", side_effect=fake_import):
            self.assertFalse(_is_legacy_xls_content(file_obj))

    def test_has_zip_signature_returns_false_on_seek_error(self):
        class BrokenFile:
            def seek(self, *_args, **_kwargs):
                raise OSError("seek failed")

        self.assertFalse(_has_zip_signature(BrokenFile()))

    @patch(
        "file_processing.services.upload_service.word_validation_service.validate_word"
    )
    @patch("file_processing.services.upload_service.validate_mime_type")
    def test_validate_file_accepts_doc_with_valid_stub(
        self, mock_validate_mime, mock_validate_word
    ):
        mock_validate_mime.return_value = (True, None)
        mock_validate_word.return_value = (True, None)

        f = SimpleUploadedFile(
            "contract.doc",
            b"dummy-doc-content",
            content_type="application/msword",
        )

        is_valid, error = validate_file(f)

        self.assertTrue(is_valid)
        self.assertIsNone(error)
        mock_validate_mime.assert_called_once_with(f, ".doc")
        mock_validate_word.assert_called_once_with(f, ".doc")

    @patch(
        "file_processing.services.upload_service.word_validation_service.validate_word"
    )
    @patch("file_processing.services.upload_service.validate_mime_type")
    def test_validate_file_accepts_docx_with_valid_stub(
        self, mock_validate_mime, mock_validate_word
    ):
        mock_validate_mime.return_value = (True, None)
        mock_validate_word.return_value = (True, None)

        f = SimpleUploadedFile(
            "contract.docx",
            b"dummy-docx-content",
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

        is_valid, error = validate_file(f)

        self.assertTrue(is_valid)
        self.assertIsNone(error)
        mock_validate_mime.assert_called_once_with(f, ".docx")
        mock_validate_word.assert_called_once_with(f, ".docx")

    @patch(
        "file_processing.services.upload_service.word_validation_service.validate_word"
    )
    @patch("file_processing.services.upload_service.validate_mime_type")
    def test_validate_file_returns_word_error(
        self, mock_validate_mime, mock_validate_word
    ):
        mock_validate_mime.return_value = (True, None)
        mock_validate_word.return_value = (False, "Word file is password-protected.")

        f = SimpleUploadedFile(
            "contract.docx",
            b"dummy-docx-content",
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

        is_valid, error = validate_file(f)

        self.assertFalse(is_valid)
        self.assertEqual(error, "Word file is password-protected.")

    def test_validate_word_docx_happy_path(self):
        content = BytesIO()
        with zipfile.ZipFile(content, "w") as archive:
            archive.writestr("[Content_Types].xml", "<Types></Types>")
            archive.writestr("word/document.xml", "<w:document></w:document>")
            archive.writestr(
                "docProps/app.xml", "<Properties><Pages>8</Pages></Properties>"
            )

        f = SimpleUploadedFile(
            "contract.docx",
            content.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

        is_valid, error = upload_service.word_validation_service.validate_word(
            f, ".docx"
        )

        self.assertTrue(is_valid)
        self.assertIsNone(error)

    @patch("file_processing.services.word_validation_service.zipfile.ZipFile")
    @patch(
        "file_processing.services.word_validation_service.is_ole_container",
        return_value=True,
    )
    def test_validate_word_docx_stops_on_encrypted(
        self,
        _mock_is_ole,
        mock_zip,
    ):
        f = SimpleUploadedFile(
            "contract.docx",
            b"dummy-docx-content",
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

        is_valid, error = upload_service.word_validation_service.validate_word(
            f, ".docx"
        )

        self.assertFalse(is_valid)
        self.assertEqual(error, "Word file is password-protected.")
        mock_zip.assert_not_called()

    @patch(
        "file_processing.services.word_validation_service.WordPageCountValidationHandler.handle",
        return_value=(False, "Word exceeds the maximum allowed page count of 100."),
    )
    def test_validate_word_doc_stops_on_page_count_limit(
        self,
        mock_page_count,
    ):
        f = SimpleUploadedFile(
            "contract.doc",
            upload_service.OLE_SIGNATURE + b"WordDocument",
            content_type="application/msword",
        )

        is_valid, error = upload_service.word_validation_service.validate_word(
            f, ".doc"
        )

        self.assertFalse(is_valid)
        self.assertEqual(error, "Word exceeds the maximum allowed page count of 100.")
        mock_page_count.assert_called_once()

    def test_validate_word_doc_happy_path(self):
        f = SimpleUploadedFile(
            "contract.doc",
            upload_service.OLE_SIGNATURE + b"WordDocument",
            content_type="application/msword",
        )

        is_valid, error = upload_service.word_validation_service.validate_word(
            f, ".doc"
        )

        self.assertTrue(is_valid)
        self.assertIsNone(error)

    @patch(
        "file_processing.services.word_validation_service.DocStructureValidationHandler.handle"
    )
    def test_validate_word_doc_stops_on_encrypted(
        self,
        mock_doc_structure,
    ):
        f = SimpleUploadedFile(
            "contract.doc",
            upload_service.OLE_SIGNATURE + b"EncryptedPackage",
            content_type="application/msword",
        )

        is_valid, error = upload_service.word_validation_service.validate_word(
            f, ".doc"
        )

        self.assertFalse(is_valid)
        self.assertEqual(error, "Word file is password-protected.")
        mock_doc_structure.assert_not_called()

    @patch(
        "file_processing.services.word_validation_service.WordPageCountValidationHandler.handle"
    )
    def test_validate_word_docx_stops_on_structure_error(
        self,
        mock_page_count_handler,
    ):
        content = BytesIO()
        with zipfile.ZipFile(content, "w") as archive:
            archive.writestr(
                "docProps/app.xml", "<Properties><Pages>8</Pages></Properties>"
            )

        f = SimpleUploadedFile(
            "contract.docx",
            content.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

        is_valid, error = upload_service.word_validation_service.validate_word(
            f, ".docx"
        )

        self.assertFalse(is_valid)
        self.assertEqual(error, "Word file is corrupt or has an invalid structure.")
        mock_page_count_handler.assert_not_called()

    def test_validate_word_rejects_unsupported_extension(self):
        f = SimpleUploadedFile(
            "contract.txt",
            b"dummy-text-content",
            content_type="text/plain",
        )

        is_valid, error = upload_service.word_validation_service.validate_word(
            f, ".txt"
        )

        self.assertFalse(is_valid)
        self.assertEqual(error, "Unsupported file type.")

    def test_estimate_doc_page_count_returns_zero_for_empty_content(self):
        page_count = upload_service.word_validation_service.estimate_doc_page_count(b"")
        self.assertEqual(page_count, 0)

    def test_estimate_doc_page_count_handles_decode_exception(self):
        class _DecodeErrorContent:
            def __bool__(self):
                return True

            def count(self, _marker):
                return 0

            def decode(self, *_args, **_kwargs):
                raise ValueError("decode failed")

        page_count = upload_service.word_validation_service.estimate_doc_page_count(
            _DecodeErrorContent()
        )
        self.assertEqual(page_count, 0)

    @patch("file_processing.services.word_validation_service.is_ole_container")
    def test_check_docx_encrypted_rejects_ole_container(self, mock_is_ole):
        mock_is_ole.return_value = True

        f = SimpleUploadedFile(
            "contract.docx",
            b"dummy-docx-content",
            content_type="application/octet-stream",
        )

        is_valid, error = upload_service.word_validation_service.check_docx_encrypted(f)

        self.assertFalse(is_valid)
        self.assertEqual(error, "Word file is password-protected.")

    def test_check_docx_structure_rejects_invalid_zip(self):
        f = SimpleUploadedFile(
            "broken.docx",
            b"not-a-zip-file",
            content_type="application/octet-stream",
        )

        is_valid, error = upload_service.word_validation_service.check_docx_structure(f)

        self.assertFalse(is_valid)
        self.assertEqual(error, "Word file is corrupt or has an invalid structure.")

    def test_check_docx_structure_accepts_minimal_valid_docx(self):
        content = BytesIO()
        with zipfile.ZipFile(content, "w") as archive:
            archive.writestr("[Content_Types].xml", "<Types></Types>")
            archive.writestr("word/document.xml", "<w:document></w:document>")
        content.seek(0)

        f = SimpleUploadedFile(
            "valid.docx",
            content.read(),
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

        is_valid, page_count = (
            upload_service.word_validation_service.check_docx_structure(f)
        )

        self.assertTrue(is_valid)
        self.assertEqual(page_count, 0)

    def test_check_doc_encrypted_detects_encrypted_markers(self):
        encrypted_doc = upload_service.OLE_SIGNATURE + b"EncryptedPackage" + b"padding"
        f = SimpleUploadedFile(
            "encrypted.doc",
            encrypted_doc,
            content_type="application/msword",
        )

        is_valid, error = upload_service.word_validation_service.check_doc_encrypted(f)

        self.assertFalse(is_valid)
        self.assertEqual(error, "Word file is password-protected.")

    def test_check_doc_encrypted_returns_corrupt_when_stream_fails(self):
        class BrokenFile:
            def seek(self, *_args, **_kwargs):
                raise OSError("seek failed")

            def read(self, *_args, **_kwargs):
                return b""

        with patch(
            "file_processing.services.word_validation_service.is_ole_container",
            return_value=True,
        ):
            is_valid, error = (
                upload_service.word_validation_service.check_doc_encrypted(BrokenFile())
            )

        self.assertFalse(is_valid)
        self.assertEqual(error, "Word file is corrupt or has an invalid structure.")

    def test_check_doc_structure_rejects_missing_worddocument_stream(self):
        content_without_word_stream = upload_service.OLE_SIGNATURE + b"not-a-word-doc"
        f = SimpleUploadedFile(
            "broken.doc",
            content_without_word_stream,
            content_type="application/msword",
        )

        is_valid, error = upload_service.word_validation_service.check_doc_structure(f)

        self.assertFalse(is_valid)
        self.assertEqual(error, "Word file is corrupt or has an invalid structure.")

    def test_check_doc_structure_accepts_worddocument_stream(self):
        content_with_word_stream = upload_service.OLE_SIGNATURE + b"WordDocument"
        f = SimpleUploadedFile(
            "valid.doc",
            content_with_word_stream,
            content_type="application/msword",
        )

        is_valid, page_count = (
            upload_service.word_validation_service.check_doc_structure(f)
        )

        self.assertTrue(is_valid)
        self.assertEqual(page_count, 0)

    @patch("file_processing.services.upload_service.magic.from_buffer")
    @patch("file_processing.services.upload_service._is_ole_container")
    def test_validate_mime_type_doc_rejects_non_ole_signature(
        self, mock_is_ole, mock_magic
    ):
        mock_magic.return_value = "application/msword"
        mock_is_ole.return_value = False

        f = SimpleUploadedFile(
            "contract.doc",
            b"not-ole-content",
            content_type="application/msword",
        )

        is_valid, error = validate_mime_type(f, ".doc")

        self.assertFalse(is_valid)
        self.assertEqual(error, "File content does not match its extension.")

    @patch("file_processing.services.upload_service.magic.from_buffer")
    @patch("file_processing.services.upload_service._is_ole_container")
    @patch("file_processing.services.upload_service._has_zip_signature")
    def test_validate_mime_type_docx_rejects_non_zip_signature(
        self, mock_zip_signature, mock_is_ole, mock_magic
    ):
        mock_magic.return_value = (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )
        mock_zip_signature.return_value = False
        mock_is_ole.return_value = False

        f = SimpleUploadedFile(
            "contract.docx",
            b"not-zip-content",
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

        is_valid, error = validate_mime_type(f, ".docx")

        self.assertFalse(is_valid)
        self.assertEqual(error, "File content does not match its extension.")

    @patch("file_processing.services.upload_service.magic.from_buffer")
    @patch("file_processing.services.upload_service._is_ole_container")
    @patch("file_processing.services.upload_service._has_zip_signature")
    def test_validate_mime_type_docx_accepts_zip_signature(
        self, mock_zip_signature, mock_is_ole, mock_magic
    ):
        mock_magic.return_value = (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )
        mock_zip_signature.return_value = True
        mock_is_ole.return_value = False

        f = SimpleUploadedFile(
            "contract.docx",
            b"zip-like-content",
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

        is_valid, error = validate_mime_type(f, ".docx")

        self.assertTrue(is_valid)
        self.assertIsNone(error)

    @patch("file_processing.services.upload_service.magic.from_buffer")
    @patch("file_processing.services.upload_service._is_ole_container")
    @patch("file_processing.services.upload_service._has_zip_signature")
    def test_validate_mime_type_docx_accepts_ole_for_encrypted_ooxml(
        self, mock_zip_signature, mock_is_ole, mock_magic
    ):
        mock_magic.return_value = "application/x-ole-storage"
        mock_is_ole.return_value = True
        mock_zip_signature.return_value = False

        f = SimpleUploadedFile(
            "encrypted.docx",
            upload_service.OLE_SIGNATURE + b"EncryptedPackage",
            content_type="application/octet-stream",
        )

        is_valid, error = validate_mime_type(f, ".docx")

        self.assertTrue(is_valid)
        self.assertIsNone(error)

    def test_validate_pdf_reader_creation_fails(self):
        with patch("file_processing.services.upload_service.PdfReader") as mock_reader:
            mock_reader.side_effect = Exception("parse error")

            f = SimpleUploadedFile(
                "doc.pdf",
                b"%PDF-test",
                content_type="application/pdf",
            )
            is_valid, error = validate_pdf(f)
            self.assertFalse(is_valid)
            self.assertIn("corrupt", error.lower())

    def test_validate_pdf_structure_pdf_read_error(self):
        with patch("file_processing.services.upload_service.PdfReader") as mock_cls:
            mock_instance = MagicMock()
            mock_instance.is_encrypted = False
            type(mock_instance).pages = PropertyMock(
                side_effect=PdfReadError("bad xref")
            )
            mock_cls.return_value = mock_instance

            f = SimpleUploadedFile(
                "doc.pdf",
                b"%PDF-test",
                content_type="application/pdf",
            )
            is_valid, error = validate_pdf(f)
            self.assertFalse(is_valid)
            self.assertIn("corrupt", error.lower())

    def test_validate_pdf_structure_generic_exception(self):
        with patch("file_processing.services.upload_service.PdfReader") as mock_cls:
            mock_instance = MagicMock()
            mock_instance.is_encrypted = False
            type(mock_instance).pages = PropertyMock(
                side_effect=Exception("unexpected")
            )
            mock_cls.return_value = mock_instance

            f = SimpleUploadedFile(
                "doc.pdf",
                b"%PDF-test",
                content_type="application/pdf",
            )
            is_valid, error = validate_pdf(f)
            self.assertFalse(is_valid)
            self.assertIn("corrupt", error.lower())

    @patch("file_processing.services.upload_service.validate_mime_type")
    def test_process_upload_mime_validation_failure(self, mock_validate):
        from file_processing.services.upload_service import process_upload

        mock_validate.return_value = (False, "Invalid MIME")

        f = SimpleUploadedFile(
            "file.pdf",
            b"%PDF-1.4",
            content_type="application/pdf",
        )

        success, error, _, _ = process_upload(f)

        self.assertFalse(success)
        self.assertEqual(error, "Invalid MIME")

    @patch("file_processing.services.upload_service.process_uploaded_excel")
    def test_process_upload_excel_failure(self, mock_excel):
        mock_excel.return_value = (False, "Excel error", None)

        xlsx_content = self.generate_valid_xlsx_bytes()

        resp = self._post_file(
            "sheet.xlsx",
            xlsx_content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.data["status"], "error")

    @patch("file_processing.services.upload_service.validate_file")
    @patch("file_processing.services.upload_service.process_uploaded_excel")
    @patch("file_processing.services.upload_service.save_temp_file")
    def test_process_upload_processing_exception(
        self, mock_save, mock_excel, mock_validate
    ):
        from file_processing.services.upload_service import process_upload

        mock_validate.return_value = (True, None)
        mock_save.return_value = "/tmp/test.xlsx"
        mock_excel.side_effect = Exception("disk failure")

        f = SimpleUploadedFile(
            "file.xlsx",
            b"dummy",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with self.assertRaises(Exception):
            process_upload(f)

    @patch("file_processing.services.upload_service._process_image")
    @patch("file_processing.services.upload_service.save_temp_file")
    @patch("file_processing.services.upload_service.validate_file")
    def test_process_upload_image_success_with_extraction(
        self, mock_validate, mock_save, mock_process_image
    ):
        from file_processing.services.upload_service import process_upload

        mock_validate.return_value = (True, None)
        mock_save.return_value = "/tmp/photo.png"
        mock_process_image.return_value = (
            True,
            None,
            {"content": [{"page": 1, "text": ["hello from image"]}]},
        )

        f = SimpleUploadedFile(
            "photo.png",
            b"dummy-image-content",
            content_type="image/png",
        )

        success, error, file_path, extracted = process_upload(f)

        self.assertTrue(success)
        self.assertIsNone(error)
        self.assertIsNone(file_path)
        self.assertEqual(extracted["content"][0]["text"], ["hello from image"])
        mock_save.assert_called_once()
        mock_process_image.assert_called_once_with("/tmp/photo.png")

    @patch("file_processing.services.upload_service.validate_file")
    def test_process_upload_image_validation_failure(self, mock_validate):
        from file_processing.services.upload_service import process_upload

        mock_validate.return_value = (
            False,
            "File content does not match its extension.",
        )

        f = SimpleUploadedFile(
            "photo.png",
            b"not-an-image",
            content_type="image/png",
        )

        success, error, file_path, extracted = process_upload(f)

        self.assertFalse(success)
        self.assertEqual(error, "File content does not match its extension.")
        self.assertIsNone(file_path)
        self.assertIsNone(extracted)

    @patch("file_processing.services.upload_service._dispatch_upload_processing")
    def test_process_stored_upload_returns_extraction_result(self, mock_dispatch):
        payload = {"content": [{"page": 1, "text": ["ok"]}]}
        mock_dispatch.return_value = upload_service.ExtractionResult.ok(payload)

        result = upload_service._process_stored_upload(
            ".pdf",
            "/tmp/file.pdf",
            SimpleUploadedFile(
                "doc.pdf",
                self.generate_valid_pdf_bytes(),
                content_type="application/pdf",
            ),
        )

        self.assertTrue(result.success)
        self.assertEqual(result.extracted_data, payload)
        mock_dispatch.assert_called_once()

    @patch("file_processing.services.upload_service._dispatch_upload_processing")
    def test_process_stored_upload_raises_when_extraction_fails(self, mock_dispatch):
        mock_dispatch.return_value = upload_service.ExtractionResult.fail(
            "bad extraction"
        )

        with self.assertRaises(upload_service.UploadExtractionError) as context:
            upload_service._process_stored_upload(
                ".pdf",
                "/tmp/file.pdf",
                SimpleUploadedFile(
                    "doc.pdf",
                    self.generate_valid_pdf_bytes(),
                    content_type="application/pdf",
                ),
            )

        self.assertEqual(str(context.exception), "bad extraction")

    @patch("file_processing.services.upload_service._dispatch_upload_processing")
    def test_process_stored_upload_re_raises_validation_error(self, mock_dispatch):
        mock_dispatch.side_effect = upload_service.UploadValidationError(
            "bad validation"
        )

        with self.assertRaises(upload_service.UploadValidationError) as context:
            upload_service._process_stored_upload(
                ".pdf",
                "/tmp/file.pdf",
                SimpleUploadedFile(
                    "doc.pdf",
                    self.generate_valid_pdf_bytes(),
                    content_type="application/pdf",
                ),
            )

        self.assertEqual(str(context.exception), "bad validation")

    @patch("file_processing.services.upload_service._cleanup_temp_upload")
    @patch(
        "file_processing.services.upload_service._process_stored_upload",
        side_effect=upload_service.UploadExtractionError("boom"),
    )
    @patch("file_processing.services.upload_service.save_temp_file")
    @patch("file_processing.services.upload_service.validate_file")
    def test_process_upload_handles_upload_extraction_error_and_cleans_up(
        self, mock_validate, mock_save, _mock_process, mock_cleanup
    ):
        from file_processing.services.upload_service import process_upload

        mock_validate.return_value = (True, None)
        mock_save.return_value = "/tmp/file.pdf"

        success, error, _, extracted = process_upload(
            SimpleUploadedFile(
                "doc.pdf",
                self.generate_valid_pdf_bytes(),
                content_type="application/pdf",
            )
        )

        self.assertFalse(success)
        self.assertEqual(error, "boom")
        self.assertIsNone(extracted)
        mock_cleanup.assert_called_once_with("/tmp/file.pdf")

    @patch(
        "file_processing.services.upload_service.save_temp_file",
        side_effect=ValueError("Invalid file path detected."),
    )
    @patch("file_processing.services.upload_service.validate_file")
    def test_process_upload_wraps_save_temp_value_error(
        self, mock_validate, mock_save
    ):
        from file_processing.services.upload_service import process_upload

        mock_validate.return_value = (True, None)

        success, error, _, extracted = process_upload(
            SimpleUploadedFile(
                "doc.pdf",
                self.generate_valid_pdf_bytes(),
                content_type="application/pdf",
            )
        )

        self.assertFalse(success)
        self.assertEqual(error, "Invalid file path detected.")
        self.assertIsNone(extracted)
        mock_save.assert_called_once()

    @patch(
        "file_processing.services.upload_service.save_temp_file",
        side_effect=upload_service.UploadStorageError("storage boom"),
    )
    @patch("file_processing.services.upload_service.validate_file")
    def test_process_upload_handles_upload_storage_error(
        self, mock_validate, mock_save
    ):
        from file_processing.services.upload_service import process_upload

        mock_validate.return_value = (True, None)

        success, error, _, extracted = process_upload(
            SimpleUploadedFile(
                "doc.pdf",
                self.generate_valid_pdf_bytes(),
                content_type="application/pdf",
            )
        )

        self.assertFalse(success)
        self.assertEqual(error, "storage boom")
        self.assertIsNone(extracted)
        mock_save.assert_called_once()

    @patch("file_processing.services.upload_service._process_image")
    @patch("file_processing.services.upload_service.save_temp_file")
    @patch("file_processing.services.upload_service.validate_file")
    def test_process_upload_image_extraction_failure(
        self, mock_validate, mock_save, mock_process_image
    ):
        from file_processing.services.upload_service import process_upload

        mock_validate.return_value = (True, None)
        mock_save.return_value = "/tmp/bad.png"
        mock_process_image.return_value = (
            False,
            "Image file is corrupted or unreadable.",
            None,
        )

        f = SimpleUploadedFile(
            "bad.png",
            b"dummy-image-content",
            content_type="image/png",
        )

        success, error, file_path, extracted = process_upload(f)

        self.assertFalse(success)
        self.assertEqual(error, "Image file is corrupted or unreadable.")
        self.assertIsNone(file_path)
        self.assertIsNone(extracted)

    def test_save_temp_file_wraps_unexpected_storage_error(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            original_dir = upload_service.settings.UPLOAD_TEMP_DIR
            upload_service.settings.UPLOAD_TEMP_DIR = tmpdir
            try:
                f = SimpleUploadedFile("my file.xlsx", b"abc")

                with patch.object(f, "chunks", side_effect=OSError("boom")):
                    with self.assertRaises(upload_service.UploadStorageError) as context:
                        upload_service.save_temp_file(f)

                self.assertEqual(str(context.exception), "Failed to save temporary file")
            finally:
                upload_service.settings.UPLOAD_TEMP_DIR = original_dir

    def test_save_temp_file_re_raises_explicit_upload_storage_error(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            original_dir = upload_service.settings.UPLOAD_TEMP_DIR
            upload_service.settings.UPLOAD_TEMP_DIR = tmpdir
            try:
                f = SimpleUploadedFile("my file.xlsx", b"abc")

                with patch(
                    "file_processing.services.upload_service.get_valid_filename",
                    side_effect=upload_service.UploadStorageError("explicit storage error"),
                ):
                    with self.assertRaises(upload_service.UploadStorageError) as context:
                        upload_service.save_temp_file(f)

                self.assertEqual(str(context.exception), "explicit storage error")
            finally:
                upload_service.settings.UPLOAD_TEMP_DIR = original_dir

    @patch("file_processing.services.upload_service.OCRService.process_pdf")
    @patch(
        "file_processing.services.upload_service.NonOCRPDFService.extract_non_ocr_pdf_to_json"
    )
    def test_process_pdf_ocr_fallback_called(self, mock_non_ocr, mock_ocr):
        from file_processing.services.upload_service import _process_pdf

        mock_non_ocr.return_value = None
        mock_ocr.return_value = {"text": "ocr"}

        pdf_doc = self.generate_valid_pdf_bytes()
        f = SimpleUploadedFile(
            "doc.pdf",
            pdf_doc,
            content_type="application/pdf",
        )

        success, error, data = _process_pdf("/tmp/file.pdf", f)

        self.assertTrue(success)
        mock_ocr.assert_called_once()

    @patch("file_processing.services.upload_service.OCRService.process_pdf")
    @patch(
        "file_processing.services.upload_service.NonOCRPDFService.extract_non_ocr_pdf_to_json"
    )
    def test_process_pdf_non_ocr_exception_triggers_ocr(self, mock_non_ocr, mock_ocr):
        from file_processing.services.upload_service import _process_pdf

        mock_non_ocr.side_effect = Exception("Non-OCR crash")
        mock_ocr.return_value = {"content": "ocr"}

        pdf_doc = self.generate_valid_pdf_bytes()

        f = SimpleUploadedFile(
            "doc.pdf",
            pdf_doc,
            content_type="application/pdf",
        )

        success, error, data = _process_pdf("/tmp/file.pdf", f)

        self.assertTrue(success)
        mock_ocr.assert_called_once()

    def test_process_upload_service_unsupported_extension(self):
        from file_processing.services.upload_service import process_upload

        f = SimpleUploadedFile(
            "file.xyz",
            b"data",
            content_type="application/octet-stream",
        )

        success, error, _, _ = process_upload(f)

        self.assertFalse(success)
        self.assertIn("Unsupported file type", error)

    @patch("file_processing.services.upload_service.validate_file")
    def test_process_upload_validate_file_failure(self, mock_validate):
        from file_processing.services.upload_service import process_upload

        mock_validate.return_value = (False, "Invalid file")

        f = SimpleUploadedFile(
            "doc.pdf",
            b"%PDF-1.4",
            content_type="application/pdf",
        )

        success, error, _, _ = process_upload(f)

        self.assertFalse(success)
        self.assertEqual(error, "Invalid file")

    @patch("file_processing.services.upload_service._process_pdf")
    @patch("file_processing.services.upload_service.save_temp_file")
    @patch("file_processing.services.upload_service.validate_mime_type")
    @patch("file_processing.services.upload_service.validate_file")
    def test_process_upload_pdf_processing_failure(
        self, mock_validate_file, mock_mime, mock_save, mock_pdf
    ):
        from file_processing.services.upload_service import process_upload

        mock_validate_file.return_value = (True, None)
        mock_mime.return_value = (True, None)
        mock_save.return_value = "/tmp/test.pdf"
        mock_pdf.return_value = (False, "PDF error", None)

        f = SimpleUploadedFile(
            "doc.pdf",
            b"%PDF-1.4",
            content_type="application/pdf",
        )

        success, error, _, _ = process_upload(f)

        self.assertFalse(success)
        self.assertEqual(error, "PDF error")

    @patch("file_processing.services.upload_service.validate_file")
    @patch("file_processing.services.upload_service.validate_mime_type")
    @patch("file_processing.services.upload_service.save_temp_file")
    def test_process_upload_else_branch_unsupported_extension(
        self, mock_save, mock_mime, mock_validate
    ):
        from file_processing.services.upload_service import process_upload

        mock_validate.return_value = (True, None)
        mock_mime.return_value = (True, None)
        mock_save.return_value = "/tmp/test.bin"

        f = SimpleUploadedFile(
            "file.bin",
            b"dummy",
            content_type="application/octet-stream",
        )

        success, error, _, _ = process_upload(f)

        self.assertFalse(success)
        self.assertEqual(error, "Unsupported file type")

    def test_get_empty_page_numbers_invalid_data(self):
        self.assertEqual(_get_empty_page_numbers(None), [])
        self.assertEqual(_get_empty_page_numbers({}), [])
        self.assertEqual(_get_empty_page_numbers({"other_key": "val"}), [])


class TestWordExtractionService(TestCase):
    def _create_docx_file(self, xml_content: str) -> str:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".docx")
        tmp.close()
        with zipfile.ZipFile(tmp.name, "w") as archive:
            archive.writestr("word/document.xml", xml_content)
        return tmp.name

    def test_extract_word_to_json_docx_success(self):
        file_path = self._create_docx_file(
            """
            <w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
              <w:body>
                <w:p><w:r><w:t>Hello</w:t></w:r></w:p>
                <w:p><w:r><w:t>World</w:t></w:r></w:p>
              </w:body>
            </w:document>
            """
        )
        try:
            result = WordExtractionService.extract_word_to_json(file_path, ".docx")
            self.assertEqual(result["content"][0]["page"], 1)
            self.assertEqual(result["content"][0]["text"], ["Hello", "World"])
        finally:
            os.unlink(file_path)

    @patch("file_processing.services.word_extraction_service.olefile")
    def test_extract_word_to_json_doc_success(self, mock_olefile):
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".doc")
        try:
            tmp.write(b"placeholder")
            tmp.close()

            mock_ole = MagicMock()
            mock_ole.exists.side_effect = lambda name: name == "WordDocument"
            mock_ole.openstream.return_value.read.return_value = (
                "Hello legacy doc\nSecond line".encode("utf-16-le")
            )
            mock_olefile.OleFileIO.return_value.__enter__.return_value = mock_ole

            result = WordExtractionService.extract_word_to_json(tmp.name, ".doc")
            extracted = " ".join(result["content"][0]["text"])
            self.assertIn("Hello legacy doc", extracted)
            self.assertIn("Second line", extracted)
        finally:
            os.unlink(tmp.name)

    def test_extract_word_to_json_unsupported_extension_raises(self):
        with self.assertRaises(ValueError):
            WordExtractionService.extract_word_to_json("/tmp/file.txt", ".txt")

    def test_extract_docx_to_json_invalid_structure_raises(self):
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".docx")
        tmp.close()
        with zipfile.ZipFile(tmp.name, "w") as archive:
            archive.writestr("[Content_Types].xml", "<Types></Types>")

        try:
            with self.assertRaises(ValueError):
                WordExtractionService._extract_docx_to_json(tmp.name)
        finally:
            os.unlink(tmp.name)

    def test_extract_docx_to_json_handles_tab_br_cr_nodes(self):
        file_path = self._create_docx_file(
            """
            <w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
              <w:body>
                <w:p>
                  <w:r><w:t>Hello</w:t></w:r>
                  <w:r><w:tab/></w:r>
                  <w:r><w:t>Word</w:t></w:r>
                  <w:r><w:br/></w:r>
                  <w:r><w:t>Line</w:t></w:r>
                  <w:r><w:cr/></w:r>
                </w:p>
              </w:body>
            </w:document>
            """
        )
        try:
            result = WordExtractionService._extract_docx_to_json(file_path)
            self.assertEqual(result["content"][0]["text"], ["Hello Word Line"])
        finally:
            os.unlink(file_path)

    def test_extract_docx_to_json_appends_only_non_empty_paragraph_text(self):
        file_path = self._create_docx_file(
            """
            <w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
              <w:body>
                <w:p><w:r><w:t>   </w:t></w:r></w:p>
                <w:p><w:r><w:t>Alpha</w:t></w:r></w:p>
              </w:body>
            </w:document>
            """
        )
        try:
            result = WordExtractionService._extract_docx_to_json(file_path)
            self.assertEqual(result["content"][0]["text"], ["Alpha"])
        finally:
            os.unlink(file_path)

    @patch("file_processing.services.word_extraction_service.olefile")
    def test_extract_doc_to_json_skips_non_alpha_and_duplicates(self, mock_olefile):
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".doc")
        try:
            tmp.write(b"placeholder")
            tmp.close()

            mock_ole = MagicMock()
            mock_ole.exists.side_effect = lambda name: name == "WordDocument"
            mock_ole.openstream.return_value.read.return_value = (
                b"Hello\nHello\n12345\nWorld\n"
            )
            mock_olefile.OleFileIO.return_value.__enter__.return_value = mock_ole

            result = WordExtractionService._extract_doc_to_json(tmp.name)
            lines = result["content"][0]["text"]
            self.assertEqual(lines.count("Hello"), 1)
            self.assertIn("World", lines)
            self.assertNotIn("12345", lines)
        finally:
            os.unlink(tmp.name)

    @patch("file_processing.services.word_extraction_service.olefile")
    def test_extract_doc_to_json_raises_on_unreadable_text(self, mock_olefile):
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".doc")
        try:
            tmp.write(b"placeholder")
            tmp.close()

            mock_ole = MagicMock()
            mock_ole.exists.side_effect = lambda name: name == "WordDocument"
            mock_ole.openstream.return_value.read.return_value = b"\x00\x01\x02\x03"
            mock_olefile.OleFileIO.return_value.__enter__.return_value = mock_ole

            with self.assertRaises(ValueError) as exc:
                WordExtractionService._extract_doc_to_json(tmp.name)

            self.assertEqual(
                str(exc.exception),
                "Unable to extract readable text from legacy .doc file.",
            )
        finally:
            os.unlink(tmp.name)

    @patch("file_processing.services.word_extraction_service.olefile", None)
    def test_extract_doc_to_json_raises_when_olefile_unavailable(self):
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".doc")
        try:
            tmp.write(b"placeholder")
            tmp.close()

            with self.assertRaises(ValueError) as exc:
                WordExtractionService._extract_doc_to_json(tmp.name)

            self.assertEqual(
                str(exc.exception),
                "Word file is corrupt or has an invalid structure.",
            )
        finally:
            os.unlink(tmp.name)

    @patch("file_processing.services.word_extraction_service.olefile")
    def test_extract_doc_to_json_raises_when_worddocument_stream_missing(
        self, mock_olefile
    ):
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".doc")
        try:
            tmp.write(b"placeholder")
            tmp.close()

            mock_ole = MagicMock()
            mock_ole.exists.return_value = False
            mock_olefile.OleFileIO.return_value.__enter__.return_value = mock_ole

            with self.assertRaises(ValueError) as exc:
                WordExtractionService._extract_doc_to_json(tmp.name)

            self.assertEqual(
                str(exc.exception),
                "Word file is corrupt or has an invalid structure.",
            )
        finally:
            os.unlink(tmp.name)

    def test_extract_printable_lines_skips_low_density_rows(self):
        payload = b"abc1234567890\nHello World\n"

        lines = WordExtractionService._extract_printable_lines(payload)

        self.assertIn("Hello World", lines)
        self.assertNotIn("abc1234567890", lines)

    @patch(
        "file_processing.services.word_extraction_service.open",
        side_effect=OSError("boom"),
    )
    def test_extract_doc_to_json_open_error_raises_value_error(self, _mock_open):
        with self.assertRaises(ValueError):
            WordExtractionService._extract_doc_to_json("/tmp/missing.doc")

    def test_local_name_handles_namespaced_and_plain_tags(self):
        self.assertEqual(WordExtractionService._local_name("{ns}p"), "p")
        self.assertEqual(WordExtractionService._local_name("p"), "p")

    @patch(
        "file_processing.services.upload_service.WordExtractionService.extract_word_to_json"
    )
    def test_process_word_success(self, mock_extract):
        mock_extract.return_value = {"content": [{"page": 1, "text": ["ok"]}]}

        success, error, data = upload_service.process_word("/tmp/f.docx", ".docx")

        self.assertTrue(success)
        self.assertIsNone(error)
        self.assertEqual(data, {"content": [{"page": 1, "text": ["ok"]}]})

    @patch(
        "file_processing.services.upload_service.WordExtractionService.extract_word_to_json"
    )
    def test_process_word_value_error_propagates_message(self, mock_extract):
        mock_extract.side_effect = ValueError("bad word")

        success, error, data = upload_service.process_word("/tmp/f.docx", ".docx")

        self.assertFalse(success)
        self.assertEqual(error, "bad word")
        self.assertIsNone(data)

    @patch("file_processing.services.upload_service.logger.exception")
    @patch(
        "file_processing.services.upload_service.WordExtractionService.extract_word_to_json"
    )
    def test_process_word_generic_exception_returns_word_corrupt_error(
        self, mock_extract, mock_logger
    ):
        mock_extract.side_effect = RuntimeError("boom")

        success, error, data = upload_service.process_word("/tmp/f.doc", ".doc")

        self.assertFalse(success)
        self.assertEqual(error, upload_service.WORD_CORRUPT_ERROR)
        self.assertIsNone(data)
        mock_logger.assert_called_once()


class TestUploadServiceCoverageGaps(TestCase):
    def _pdf_file(self):
        return SimpleUploadedFile(
            "doc.pdf", b"%PDF-1.4\n%%EOF", content_type="application/pdf"
        )

    def _xlsx_file(self):
        return SimpleUploadedFile(
            "sheet.xlsx",
            b"PK\x03\x04dummy",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_has_extracted_text_true_and_false(self):
        self.assertFalse(_has_extracted_text({"content": [{"page": 1, "text": []}]}))
        self.assertTrue(_has_extracted_text({"content": [{"page": 1, "text": ["ok"]}]}))

    def test_fallback_mime_returns_zip_for_zip_signature(self):
        mime = upload_service._fallback_mime(b"PK\x03\x04dummy", ".xlsx")
        self.assertEqual(mime, upload_service.MIME_ZIP)

    @patch(
        "file_processing.services.upload_service.validate_pdf",
        return_value=(False, "bad"),
    )
    def test_process_pdf_short_circuit_on_invalid_pdf(self, _mock_validate_pdf):
        success, error, data = upload_service._process_pdf(
            "/tmp/f.pdf", self._pdf_file()
        )
        self.assertFalse(success)
        self.assertEqual(error, "bad")
        self.assertIsNone(data)

    @patch(
        "file_processing.services.upload_service.validate_pdf",
        return_value=(True, None),
    )
    @patch(
        "file_processing.services.upload_service.NonOCRPDFService.extract_non_ocr_pdf_to_json"
    )
    @patch("file_processing.services.upload_service.OCRService.process_pdf_pages")
    def test_process_pdf_merges_ocr_for_empty_pages(
        self, mock_pdf_pages, mock_non_ocr, _mock_validate
    ):
        mock_non_ocr.return_value = {
            "content": [
                {"page": 1, "text": ["native"]},
                {"page": 2, "text": []},
            ]
        }
        mock_pdf_pages.return_value = {"content": [{"page": 2, "text": ["ocr"]}]}

        success, error, data = upload_service._process_pdf(
            "/tmp/f.pdf", self._pdf_file()
        )
        self.assertTrue(success)
        self.assertIsNone(error)
        self.assertEqual(data["content"][1]["text"], ["ocr"])

    @patch(
        "file_processing.services.upload_service.validate_file",
        return_value=(True, None),
    )
    @patch(
        "file_processing.services.upload_service.save_temp_file",
        return_value="/tmp/f.pdf",
    )
    @patch(
        "file_processing.services.upload_service._process_pdf",
        return_value=(True, None, {"content": []}),
    )
    @patch("file_processing.services.upload_service.os.path.exists", return_value=False)
    def test_process_upload_pdf_success_when_temp_already_missing(
        self, _exists, _proc_pdf, _save, _validate
    ):
        success, error, _, extracted = upload_service.process_upload(self._pdf_file())
        self.assertTrue(success)
        self.assertIsNone(error)
        self.assertEqual(extracted, {"content": []})

    @patch(
        "file_processing.services.upload_service.validate_file",
        return_value=(True, None),
    )
    @patch(
        "file_processing.services.upload_service.save_temp_file",
        return_value="/tmp/f.xlsx",
    )
    @patch(
        "file_processing.services.upload_service.process_uploaded_excel",
        return_value=(False, "excel bad", None),
    )
    @patch("file_processing.services.upload_service.os.path.exists", return_value=False)
    def test_process_upload_excel_failure_direct(
        self, _exists, _excel, _save, _validate
    ):
        success, error, _, extracted = upload_service.process_upload(self._xlsx_file())
        self.assertFalse(success)
        self.assertEqual(error, "excel bad")
        self.assertIsNone(extracted)

    @patch(
        "file_processing.services.upload_service.validate_file",
        return_value=(True, None),
    )
    @patch(
        "file_processing.services.upload_service.save_temp_file",
        return_value="/tmp/f.pdf",
    )
    @patch(
        "file_processing.services.upload_service._process_pdf",
        return_value=(True, None, {"content": []}),
    )
    @patch("file_processing.services.upload_service.os.path.exists", return_value=True)
    @patch(
        "file_processing.services.upload_service.os.remove",
        side_effect=OSError("cannot remove"),
    )
    @patch("file_processing.services.upload_service.logger.exception")
    def test_process_upload_logs_remove_error(
        self, mock_log, _remove, _exists, _proc_pdf, _save, _validate
    ):
        success, _error, _, _extracted = upload_service.process_upload(self._pdf_file())
        self.assertTrue(success)
        mock_log.assert_called_once()

    @patch(
        "file_processing.services.upload_service.validate_mime_type",
        return_value=(True, None),
    )
    def test_validate_file_size_limit(self, _mock_mime):
        f = SimpleUploadedFile("large.pdf", b"x")
        f.size = MAX_FILE_SIZE + 1
        is_valid, error = upload_service.validate_file(f)
        self.assertFalse(is_valid)
        self.assertEqual(error, FILE_TOO_LARGE_ERROR)

    @patch(
        "file_processing.services.upload_service.validate_mime_type",
        return_value=(True, None),
    )
    @patch(
        "file_processing.services.upload_service.validate_excel_sheet_count",
        return_value=(False, "too many"),
    )
    def test_validate_file_excel_sheet_error(self, _excel_count, _mime):
        f = SimpleUploadedFile("sheet.xlsx", b"PK", content_type="application/zip")
        is_valid, error = upload_service.validate_file(f)
        self.assertFalse(is_valid)
        self.assertEqual(error, "too many")

    @patch(
        "file_processing.services.word_validation_service.DocStructureValidationHandler.handle",
        return_value=(False, "bad doc"),
    )
    def test_validate_word_doc_structure_error(self, _structure):
        payload = upload_service.OLE_SIGNATURE + b"WordDocument"
        is_valid, error = upload_service.word_validation_service.validate_word(
            SimpleUploadedFile("x.doc", payload), ".doc"
        )
        self.assertFalse(is_valid)
        self.assertEqual(error, "bad doc")

    def test_check_docx_encrypted_non_ole(self):
        is_valid, error = upload_service.word_validation_service.check_docx_encrypted(
            SimpleUploadedFile("x.docx", b"PK")
        )
        self.assertTrue(is_valid)
        self.assertIsNone(error)

    def test_check_docx_structure_missing_required_entries(self):
        content = BytesIO()
        with zipfile.ZipFile(content, "w") as archive:
            archive.writestr("docProps/app.xml", "<Properties></Properties>")
        f = SimpleUploadedFile(
            "broken.docx", content.getvalue(), content_type="application/zip"
        )
        is_valid, error = upload_service.word_validation_service.check_docx_structure(f)
        self.assertFalse(is_valid)
        self.assertEqual(error, upload_service.WORD_CORRUPT_ERROR)

    def test_extract_docx_page_count_with_pages_and_without_pages(self):
        class ArchiveWithPages:
            def read(self, _name):
                return b"<Properties><Pages>7</Pages></Properties>"

        class ArchiveNoPages:
            def read(self, _name):
                return b"<Properties><Template>Normal</Template></Properties>"

        self.assertEqual(
            upload_service.word_validation_service.extract_docx_page_count(
                ArchiveWithPages()
            ),
            7,
        )
        self.assertEqual(
            upload_service.word_validation_service.extract_docx_page_count(
                ArchiveNoPages()
            ),
            0,
        )

    def test_check_doc_encrypted_rejects_non_ole(self):
        is_valid, error = upload_service.word_validation_service.check_doc_encrypted(
            SimpleUploadedFile("x.doc", b"not-ole")
        )
        self.assertFalse(is_valid)
        self.assertEqual(error, upload_service.WORD_CORRUPT_ERROR)

    def test_check_doc_encrypted_valid_doc(self):
        payload = upload_service.OLE_SIGNATURE + b"WordDocument"
        is_valid, error = upload_service.word_validation_service.check_doc_encrypted(
            SimpleUploadedFile("x.doc", payload)
        )
        self.assertTrue(is_valid)
        self.assertIsNone(error)

    def test_check_doc_structure_rejects_non_ole_and_handles_exception(self):
        is_valid, error = upload_service.word_validation_service.check_doc_structure(
            SimpleUploadedFile("x.doc", b"not-ole")
        )
        self.assertFalse(is_valid)
        self.assertEqual(error, upload_service.WORD_CORRUPT_ERROR)

        class BrokenFile:
            def seek(self, *_args, **_kwargs):
                raise OSError("boom")

            def read(self, *_args, **_kwargs):
                return b""

        with patch(
            "file_processing.services.word_validation_service.is_ole_container",
            return_value=True,
        ):
            is_valid2, error2 = (
                upload_service.word_validation_service.check_doc_structure(BrokenFile())
            )

        self.assertFalse(is_valid2)
        self.assertEqual(error2, upload_service.WORD_CORRUPT_ERROR)

    def test_check_doc_structure_page_count_detected_from_form_feed(self):
        payload = upload_service.OLE_SIGNATURE + b"WordDocument" + (b"Body\x0c" * 120)
        is_valid, page_count = (
            upload_service.word_validation_service.check_doc_structure(
                SimpleUploadedFile("x.doc", payload)
            )
        )
        self.assertTrue(is_valid)
        self.assertGreater(page_count, upload_service.MAX_WORD_PAGES)

    def test_validate_word_doc_page_limit_triggers_without_mocking(self):
        payload = upload_service.OLE_SIGNATURE + b"WordDocument" + (b"P\x0c" * 120)
        is_valid, error = upload_service.word_validation_service.validate_word(
            SimpleUploadedFile("x.doc", payload),
            ".doc",
        )
        self.assertFalse(is_valid)
        self.assertIn("maximum allowed page count", error)

    def test_validate_pdf_encrypted_and_too_many_pages(self):
        encrypted_reader = MagicMock()
        encrypted_reader.is_encrypted = True
        with patch(
            "file_processing.services.upload_service.PdfReader",
            return_value=encrypted_reader,
        ):
            is_valid, error = upload_service.validate_pdf(self._pdf_file())
            self.assertFalse(is_valid)
            self.assertIn("password-protected", error)

        many_pages_reader = MagicMock()
        many_pages_reader.is_encrypted = False
        type(many_pages_reader).pages = PropertyMock(
            return_value=[None] * (upload_service.MAX_PDF_PAGES + 1)
        )
        with patch(
            "file_processing.services.upload_service.PdfReader",
            return_value=many_pages_reader,
        ):
            is_valid, error = upload_service.validate_pdf(self._pdf_file())
            self.assertFalse(is_valid)
            self.assertIn("maximum allowed page count", error)

    def test_check_pdf_helpers(self):
        encrypted_reader = MagicMock(is_encrypted=True)
        self.assertEqual(
            upload_service.check_pdf_encrypted(encrypted_reader),
            (False, "PDF file is password-protected."),
        )
        self.assertEqual(
            upload_service.check_pdf_page_count(upload_service.MAX_PDF_PAGES + 1)[0],
            False,
        )

    @patch(
        "file_processing.services.upload_service._should_parse_as_xls",
        return_value=True,
    )
    @patch(
        "file_processing.services.upload_service._get_xls_sheet_count", return_value=3
    )
    def test_validate_excel_sheet_count_xls_path(self, _xls_count, _should_xls):
        is_valid, error = upload_service.validate_excel_sheet_count(
            self._xlsx_file(), ".xlsx"
        )
        self.assertTrue(is_valid)
        self.assertIsNone(error)

    @patch(
        "file_processing.services.upload_service._should_parse_as_xls",
        return_value=False,
    )
    @patch(
        "file_processing.services.upload_service._get_xlsx_sheet_count",
        return_value=upload_service.MAX_EXCEL_SHEETS + 1,
    )
    def test_validate_excel_sheet_count_too_many(self, _xlsx_count, _should_xls):
        is_valid, error = upload_service.validate_excel_sheet_count(
            self._xlsx_file(), ".xlsx"
        )
        self.assertFalse(is_valid)
        self.assertEqual(error, upload_service.EXCEL_TOO_MANY_SHEETS_ERROR)

    @patch(
        "file_processing.services.upload_service._should_parse_as_xls",
        side_effect=Exception("bad workbook"),
    )
    def test_validate_excel_sheet_count_exception(self, _should_xls):
        is_valid, error = upload_service.validate_excel_sheet_count(
            self._xlsx_file(), ".xlsx"
        )
        self.assertFalse(is_valid)
        self.assertEqual(error, upload_service.EXCEL_CORRUPT_ERROR)

    def test_should_parse_as_xls_branches(self):
        f = self._xlsx_file()
        self.assertTrue(upload_service._should_parse_as_xls(f, ".xls"))
        with (
            patch(
                "file_processing.services.upload_service._is_ole_container",
                return_value=True,
            ),
            patch(
                "file_processing.services.upload_service._is_legacy_xls_content",
                return_value=True,
            ),
        ):
            self.assertTrue(upload_service._should_parse_as_xls(f, ".xlsx"))
        with patch(
            "file_processing.services.upload_service._is_ole_container",
            return_value=False,
        ):
            self.assertFalse(upload_service._should_parse_as_xls(f, ".xlsx"))

    def test_get_xlsx_sheet_count(self):
        buffer = BytesIO()
        wb = Workbook()
        wb.create_sheet("S2")
        wb.save(buffer)
        f = SimpleUploadedFile(
            "book.xlsx", buffer.getvalue(), content_type="application/zip"
        )
        self.assertEqual(upload_service._get_xlsx_sheet_count(f), 2)

    @patch("xlrd.open_workbook")
    def test_get_xls_sheet_count_release_resources_callable_and_not_callable(
        self, mock_open_workbook
    ):
        workbook = MagicMock()
        workbook.nsheets = 4
        workbook.release_resources = MagicMock()
        mock_open_workbook.return_value = workbook
        self.assertEqual(
            upload_service._get_xls_sheet_count(SimpleUploadedFile("f.xls", b"data")), 4
        )
        workbook.release_resources.assert_called_once()

        workbook2 = MagicMock()
        workbook2.nsheets = 2
        workbook2.release_resources = None
        mock_open_workbook.return_value = workbook2
        self.assertEqual(
            upload_service._get_xls_sheet_count(SimpleUploadedFile("f.xls", b"data2")),
            2,
        )

    def test_validate_mime_type_xlsx_ole_legacy_and_password_protected(self):
        with (
            patch(
                "file_processing.services.upload_service.magic.from_buffer",
                return_value="application/zip",
            ),
            patch(
                "file_processing.services.upload_service._is_ole_container",
                return_value=True,
            ),
            patch(
                "file_processing.services.upload_service._is_legacy_xls_content",
                return_value=True,
            ),
        ):
            ok, err = upload_service.validate_mime_type(self._xlsx_file(), ".xlsx")
            self.assertTrue(ok)
            self.assertIsNone(err)

        with (
            patch(
                "file_processing.services.upload_service.magic.from_buffer",
                return_value="application/zip",
            ),
            patch(
                "file_processing.services.upload_service._is_ole_container",
                return_value=True,
            ),
            patch(
                "file_processing.services.upload_service._is_legacy_xls_content",
                return_value=False,
            ),
        ):
            ok, err = upload_service.validate_mime_type(self._xlsx_file(), ".xlsx")
            self.assertFalse(ok)
            self.assertEqual(err, upload_service.EXCEL_PASSWORD_PROTECTED_ERROR)

    def test_validate_mime_type_xlsx_non_zip_and_unexpected_mime_and_exception(self):
        with (
            patch(
                "file_processing.services.upload_service.magic.from_buffer",
                return_value="application/zip",
            ),
            patch(
                "file_processing.services.upload_service._is_ole_container",
                return_value=False,
            ),
            patch(
                "file_processing.services.upload_service._has_zip_signature",
                return_value=False,
            ),
        ):
            ok, err = upload_service.validate_mime_type(self._xlsx_file(), ".xlsx")
            self.assertFalse(ok)
            self.assertIn("does not match", err)

        with (
            patch(
                "file_processing.services.upload_service.magic.from_buffer",
                return_value="text/plain",
            ),
            patch(
                "file_processing.services.upload_service._is_ole_container",
                return_value=False,
            ),
            patch(
                "file_processing.services.upload_service._has_zip_signature",
                return_value=True,
            ),
        ):
            ok, err = upload_service.validate_mime_type(self._xlsx_file(), ".xlsx")
            self.assertFalse(ok)
            self.assertIn("does not match", err)

        with patch(
            "file_processing.services.upload_service.magic.from_buffer",
            side_effect=Exception("boom"),
        ):
            ok, err = upload_service.validate_mime_type(
                SimpleUploadedFile(
                    "sheet.xlsx",
                    b"dummy content",
                    content_type="application/octet-stream",
                ),
                ".xlsx",
            )
            self.assertFalse(ok)
            self.assertIn("Unable to determine", err)

    def test_validate_mime_type_fallback_pdf_ole_octet_none_and_outer_exception(self):
        with patch(
            "file_processing.services.upload_service.magic.from_buffer",
            side_effect=Exception("boom"),
        ):
            ok_pdf, err_pdf = upload_service.validate_mime_type(
                SimpleUploadedFile(
                    "doc.pdf", b"%PDF-1.4\n", content_type="application/pdf"
                ),
                ".pdf",
            )
            self.assertTrue(ok_pdf)
            self.assertIsNone(err_pdf)

            ole_payload = upload_service.OLE_SIGNATURE + b"WordDocument"
            ok_doc, err_doc = upload_service.validate_mime_type(
                SimpleUploadedFile(
                    "doc.doc", ole_payload, content_type="application/msword"
                ),
                ".doc",
            )
            self.assertTrue(ok_doc)
            self.assertIsNone(err_doc)

            ok_xls, err_xls = upload_service.validate_mime_type(
                SimpleUploadedFile(
                    "sheet.xls", b"RANDOM", content_type="application/octet-stream"
                ),
                ".xls",
            )
            self.assertTrue(ok_xls)
            self.assertIsNone(err_xls)

            ok_txt, err_txt = upload_service.validate_mime_type(
                SimpleUploadedFile("note.txt", b"plain", content_type="text/plain"),
                ".txt",
            )
            self.assertFalse(ok_txt)
            self.assertIn("Unable to determine", err_txt)

        class BrokenFile:
            def seek(self, *_args, **_kwargs):
                raise OSError("seek failed")

            def read(self, *_args, **_kwargs):
                return b""

        ok_broken, err_broken = upload_service.validate_mime_type(BrokenFile(), ".pdf")
        self.assertFalse(ok_broken)
        self.assertIn("Unable to determine", err_broken)

    def test_has_zip_signature_true(self):
        self.assertTrue(_has_zip_signature(SimpleUploadedFile("x.zip", b"PKpayload")))

    def test_save_temp_file_success_and_guard(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            original_dir = upload_service.settings.UPLOAD_TEMP_DIR
            upload_service.settings.UPLOAD_TEMP_DIR = tmpdir
            try:
                f = SimpleUploadedFile("my file.xlsx", b"abc")
                saved_path = upload_service.save_temp_file(f)
                self.assertTrue(os.path.exists(saved_path))
                with open(saved_path, "rb") as saved:
                    self.assertEqual(saved.read(), b"abc")
            finally:
                upload_service.settings.UPLOAD_TEMP_DIR = original_dir

        f2 = SimpleUploadedFile("name.xlsx", b"abc")
        with patch(
            "file_processing.services.upload_service.os.path.abspath",
            side_effect=["/safe", "/unsafe/file"],
        ):
            with self.assertRaises(ValueError):
                upload_service.save_temp_file(f2)

    @patch(
        "file_processing.services.upload_service.validate_pdf",
        return_value=(True, None),
    )
    @patch(
        "file_processing.services.upload_service.NonOCRPDFService.extract_non_ocr_pdf_to_json"
    )
    @patch("file_processing.services.upload_service.OCRService.process_pdf_pages")
    def test_process_pdf_no_empty_pages_skips_partial_ocr(
        self, mock_pdf_pages, mock_non_ocr, _mock_validate
    ):
        mock_non_ocr.return_value = {
            "content": [
                {"page": 1, "text": ["native-a"]},
                {"page": 2, "text": ["native-b"]},
            ]
        }

        success, error, data = upload_service._process_pdf(
            "/tmp/f.pdf", self._pdf_file()
        )

        self.assertTrue(success)
        self.assertIsNone(error)
        self.assertEqual(data["content"][1]["text"], ["native-b"])
        mock_pdf_pages.assert_not_called()

    @patch(
        "file_processing.services.upload_service.validate_file",
        return_value=(True, None),
    )
    @patch(
        "file_processing.services.upload_service.save_temp_file",
        return_value="/tmp/f.xlsx",
    )
    @patch(
        "file_processing.services.upload_service.process_uploaded_excel",
        return_value=(True, None, {"rows": 2}),
    )
    @patch("file_processing.services.upload_service.os.path.exists", return_value=False)
    def test_process_upload_excel_success_sets_extracted_data(
        self, _exists, _excel, _save, _validate
    ):
        success, error, _, extracted = upload_service.process_upload(self._xlsx_file())
        self.assertTrue(success)
        self.assertIsNone(error)
        self.assertEqual(extracted, {"rows": 2})

    @patch(
        "file_processing.services.upload_service.validate_mime_type",
        return_value=(True, None),
    )
    @patch(
        "file_processing.services.upload_service.validate_excel_sheet_count",
        return_value=(True, None),
    )
    def test_validate_file_valid_xlsx_hits_non_word_path(self, _excel_count, _mime):
        is_valid, error = upload_service.validate_file(self._xlsx_file())
        self.assertTrue(is_valid)
        self.assertIsNone(error)

    def test_check_word_page_count_over_limit(self):
        is_valid, error = upload_service.word_validation_service.check_word_page_count(
            upload_service.MAX_WORD_PAGES + 1
        )
        self.assertFalse(is_valid)
        self.assertIn("maximum allowed page count", error)

    def test_check_word_page_count_within_limit(self):
        is_valid, error = upload_service.word_validation_service.check_word_page_count(
            upload_service.MAX_WORD_PAGES
        )
        self.assertTrue(is_valid)
        self.assertIsNone(error)


class TestWordValidationServiceCoverage(TestCase):
    class _StubFile:
        def __init__(self, payload=b""):
            self._payload = payload

        def seek(self, *_args, **_kwargs):
            return None

        def read(self, *_args, **_kwargs):
            return self._payload

    def test_validate_word_returns_unsupported_for_unknown_extension(self):
        is_valid, error = word_validation_service.validate_word(
            self._StubFile(), ".txt"
        )
        self.assertFalse(is_valid)
        self.assertEqual(error, "Unsupported file type.")

    @patch(
        "file_processing.services.word_validation_service.is_ole_container",
        return_value=True,
    )
    def test_docx_encrypted_handler_blocks_ole_using_stub(self, _mock_is_ole):
        context = word_validation_service.WordValidationContext(
            uploaded_file=self._StubFile(),
            ext=".docx",
        )
        handler = word_validation_service.DocxEncryptedValidationHandler()

        is_valid, error = handler.handle(context)

        self.assertFalse(is_valid)
        self.assertEqual(error, word_validation_service.WORD_PROTECTED_ERROR)

    @patch(
        "file_processing.services.word_validation_service.is_ole_container",
        return_value=True,
    )
    def test_doc_structure_handler_stream_failure_returns_corrupt(self, _mock_is_ole):
        class _BrokenStubFile:
            def seek(self, *_args, **_kwargs):
                raise OSError("seek failed")

            def read(self, *_args, **_kwargs):
                return b""

        context = word_validation_service.WordValidationContext(
            uploaded_file=_BrokenStubFile(),
            ext=".doc",
        )

        handler = word_validation_service.DocStructureValidationHandler()
        is_valid, error = handler.handle(context)

        self.assertFalse(is_valid)
        self.assertEqual(error, word_validation_service.WORD_CORRUPT_ERROR)

    def test_build_word_validation_chain_returns_none_for_unknown_extension(self):
        chain = word_validation_service._build_word_validation_chain(".bin")
        self.assertIsNone(chain)


class TestWordValidationService(unittest.TestCase):
    def _build_valid_docx(self, pages=1):
        content = BytesIO()
        with zipfile.ZipFile(content, "w") as archive:
            archive.writestr("[Content_Types].xml", "<Types></Types>")
            archive.writestr("word/document.xml", "<w:document></w:document>")
            archive.writestr(
                "docProps/app.xml",
                f"<Properties><Pages>{pages}</Pages></Properties>",
            )
        return content.getvalue()

    def test_validate_word_docx_happy_path(self):
        f = SimpleUploadedFile(
            "ok.docx",
            self._build_valid_docx(pages=8),
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

        is_valid, error = word_validation_service.validate_word(f, ".docx")

        self.assertTrue(is_valid)
        self.assertIsNone(error)

    @patch("file_processing.services.word_validation_service.zipfile.ZipFile")
    @patch(
        "file_processing.services.word_validation_service.is_ole_container",
        return_value=True,
    )
    def test_validate_word_docx_stops_on_encrypted(self, _mock_ole, mock_zip):
        f = SimpleUploadedFile(
            "enc.docx", b"dummy", content_type="application/octet-stream"
        )

        is_valid, error = word_validation_service.validate_word(f, ".docx")

        self.assertFalse(is_valid)
        self.assertEqual(error, "Word file is password-protected.")
        mock_zip.assert_not_called()

    def test_validate_word_doc_too_many_pages_from_chain(self):
        payload = word_validation_service.OLE_SIGNATURE + b"WordDocument"
        f = SimpleUploadedFile("x.doc", payload, content_type="application/msword")

        with patch(
            "file_processing.services.word_validation_service.DocStructureValidationHandler.handle",
            return_value=(
                False,
                f"Word exceeds the maximum allowed page count of {word_validation_service.MAX_WORD_PAGES}.",
            ),
        ):
            is_valid, error = word_validation_service.validate_word(f, ".doc")

        self.assertFalse(is_valid)
        self.assertIn("maximum allowed page count", error)

    def test_validate_word_rejects_unsupported_extension(self):
        f = SimpleUploadedFile("x.txt", b"abc", content_type="text/plain")

        is_valid, error = word_validation_service.validate_word(f, ".txt")

        self.assertFalse(is_valid)
        self.assertEqual(error, "Unsupported file type.")

    def test_check_docx_structure_invalid_zip(self):
        f = SimpleUploadedFile(
            "broken.docx", b"not-zip", content_type="application/octet-stream"
        )
        is_valid, error = word_validation_service.check_docx_structure(f)

        self.assertFalse(is_valid)
        self.assertEqual(error, word_validation_service.WORD_CORRUPT_ERROR)

    def test_extract_docx_page_count_without_pages_returns_zero(self):
        class ArchiveNoPages:
            def read(self, _name):
                return b"<Properties><Template>Normal</Template></Properties>"

        self.assertEqual(
            word_validation_service.extract_docx_page_count(ArchiveNoPages()), 0
        )
