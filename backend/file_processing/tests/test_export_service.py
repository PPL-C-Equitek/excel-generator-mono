import csv
import io
import os
import re
import tempfile
import types
import unittest
import zipfile
from copy import deepcopy
from unittest.mock import patch
from openpyxl import load_workbook

import file_processing.services.export_service as export_service
from file_processing.services.export_service import (
    OutputCSVMappingError,
    OutputLLMValidationError,
    map_output_csv,
    validate_output_llm,
)


class ValidateOutputLLMTest(unittest.TestCase):
    def _build_valid_payload(self):
        return {
            "document_info": {
                "source_type": "Excel",
                "filename": "laporan_tahunan.xlsx",
            },
            "summary": {
                "grand_total": 1500000,
                "period": "2026",
            },
            "content_data": [
                {
                    "table_name": "Sheet1_Januari",
                    "headers": ["item_name", "quantity", "price"],
                    "rows": [
                        {"item_name": "Kertas", "quantity": 10, "price": 50000},
                        {"item_name": "Pena", "quantity": 5, "price": 10000},
                    ],
                }
            ],
        }

    def test_validate_output_llm_accepts_valid_ok_payload_single_sheet(self):
        output_json = self._build_valid_payload()

        result = validate_output_llm(output_json)

        self.assertEqual(result, output_json)

    def test_validate_output_llm_accepts_valid_multiple_sheets_payload(self):
        output_json = self._build_valid_payload()
        output_json["content_data"] = [
            {
                "table_name": "Sheet1_Januari",
                "headers": ["item_name", "quantity", "price"],
                "rows": [{"item_name": "Kertas", "quantity": 10, "price": 50000}],
            },
            {
                "table_name": "Sheet2_Februari",
                "headers": ["item_name", "quantity", "price"],
                "rows": [{"item_name": "Tinta", "quantity": 1, "price": 400000}],
            },
        ]

        result = validate_output_llm(output_json)

        self.assertEqual(result, output_json)

    def test_validate_output_llm_rejects_non_object_or_array_root(self):
        with self.assertRaises(OutputLLMValidationError):
            validate_output_llm("not-valid")

    def test_validate_output_llm_rejects_missing_required_top_level_keys(self):
        required_keys = ("document_info", "summary", "content_data")
        for key in required_keys:
            with self.subTest(missing_key=key):
                payload = self._build_valid_payload()
                payload.pop(key)
                with self.assertRaises(OutputLLMValidationError):
                    validate_output_llm(payload)

    def test_validate_output_llm_rejects_invalid_top_level_values(self):
        cases = [
            ("document_info_non_object", {"document_info": 1}),
            ("summary_non_object", {"summary": 123}),
            ("content_data_non_list", {"content_data": {}}),
        ]
        for case_name, mutation in cases:
            with self.subTest(case=case_name):
                payload = self._build_valid_payload()
                payload.update(mutation)
                with self.assertRaises(OutputLLMValidationError):
                    validate_output_llm(payload)

    def test_validate_output_llm_rejects_invalid_document_info_fields(self):
        cases = [
            ("missing_source_type", {"filename": "laporan.xlsx"}),
            ("missing_filename", {"source_type": "Excel"}),
            ("source_type_non_string", {"source_type": 1, "filename": "laporan.xlsx"}),
            (
                "source_type_not_allowed",
                {"source_type": "Word", "filename": "laporan.xlsx"},
            ),
            ("filename_non_string", {"source_type": "Excel", "filename": 123}),
            ("filename_blank", {"source_type": "Excel", "filename": "   "}),
        ]
        for case_name, document_info in cases:
            with self.subTest(case=case_name):
                payload = self._build_valid_payload()
                payload["document_info"] = document_info
                with self.assertRaises(OutputLLMValidationError):
                    validate_output_llm(payload)

    def test_validate_output_llm_rejects_invalid_summary_values(self):
        payload = self._build_valid_payload()
        payload["summary"] = {
            "grand_total": 1500000,
            "period": {"year": "2026"},
        }
        with self.assertRaises(OutputLLMValidationError):
            validate_output_llm(payload)

        payload = self._build_valid_payload()
        payload["summary"] = {"": "blank-key"}
        with self.assertRaises(OutputLLMValidationError):
            validate_output_llm(payload)

        payload = self._build_valid_payload()
        payload["summary"] = {"grand_total": {1}}
        with self.assertRaises(OutputLLMValidationError):
            validate_output_llm(payload)

        payload = self._build_valid_payload()
        payload["summary"] = {"total_items": -1}
        with self.assertRaises(OutputLLMValidationError):
            validate_output_llm(payload)

    def test_validate_output_llm_rejects_summary_total_items_boolean(self):
        payload = self._build_valid_payload()
        payload["summary"] = {"total_items": True}

        with self.assertRaises(OutputLLMValidationError):
            validate_output_llm(payload)

    def test_validate_output_llm_accepts_summary_total_items_non_negative_integer(self):
        payload = self._build_valid_payload()
        payload["summary"] = {"total_items": 0}

        result = validate_output_llm(payload)

        self.assertEqual(result["summary"]["total_items"], 0)

    def test_validate_output_llm_rejects_empty_content_data_list(self):
        payload = self._build_valid_payload()
        payload["content_data"] = []
        with self.assertRaises(OutputLLMValidationError):
            validate_output_llm(payload)

    def test_validate_output_llm_rejects_invalid_content_data_structure(self):
        cases = [
            ("table_non_object", ["x"]),
            (
                "missing_table_name",
                [{"headers": ["item_name"], "rows": [{"item_name": "Kertas"}]}],
            ),
            (
                "missing_headers",
                [{"table_name": "Sheet1", "rows": [{"item_name": "Kertas"}]}],
            ),
            ("missing_rows", [{"table_name": "Sheet1", "headers": ["item_name"]}]),
            (
                "table_name_non_string",
                [{"table_name": 123, "headers": ["item_name"], "rows": [{"item_name": "Kertas"}]}],
            ),
            (
                "table_name_blank",
                [{"table_name": "   ", "headers": ["item_name"], "rows": [{"item_name": "Kertas"}]}],
            ),
            (
                "duplicate_table_name",
                [
                    {"table_name": "Sheet1", "headers": ["item_name"], "rows": [{"item_name": "Kertas"}]},
                    {"table_name": "Sheet1", "headers": ["item_name"], "rows": [{"item_name": "Pena"}]},
                ],
            ),
        ]
        for case_name, content_data in cases:
            with self.subTest(case=case_name):
                payload = self._build_valid_payload()
                payload["content_data"] = content_data
                with self.assertRaises(OutputLLMValidationError):
                    validate_output_llm(payload)

    def test_validate_output_llm_rejects_invalid_headers(self):
        cases = [
            ("headers_non_list", {"headers": "item_name"}),
            ("headers_empty", {"headers": []}),
            ("header_non_string", {"headers": ["item_name", 1]}),
            ("header_blank", {"headers": ["item_name", "   "]}),
            ("header_duplicate_case_insensitive", {"headers": ["Item", "item"]}),
        ]
        for case_name, mutation in cases:
            with self.subTest(case=case_name):
                payload = self._build_valid_payload()
                payload["content_data"][0].update(mutation)
                if case_name == "header_duplicate_case_insensitive":
                    payload["content_data"][0]["rows"] = [{"Item": "Kertas", "item": "Pena"}]
                with self.assertRaises(OutputLLMValidationError):
                    validate_output_llm(payload)

    def test_validate_output_llm_rejects_invalid_rows_container_and_row_type(self):
        cases = [
            ("rows_non_list", "not-list"),
            ("row_non_object", ["not-dict"]),
        ]
        for case_name, rows_value in cases:
            with self.subTest(case=case_name):
                payload = self._build_valid_payload()
                payload["content_data"][0]["rows"] = rows_value
                with self.assertRaises(OutputLLMValidationError):
                    validate_output_llm(payload)

    def test_validate_output_llm_rejects_row_header_mismatch_and_invalid_values(self):
        cases = [
            ("missing_required_header", [{"item_name": "Kertas", "quantity": 10}]),
            (
                "unknown_extra_header",
                [{"item_name": "Kertas", "quantity": 10, "price": 50000, "note": "x"}],
            ),
            (
                "nested_dict_value",
                [{"item_name": "Kertas", "quantity": 10, "price": {"raw": 50000}}],
            ),
            (
                "nested_list_value",
                [{"item_name": "Kertas", "quantity": 10, "price": [50000]}],
            ),
            (
                "non_scalar_value_type",
                [{"item_name": "Kertas", "quantity": 10, "price": {50000}}],
            ),
        ]
        for case_name, rows in cases:
            with self.subTest(case=case_name):
                payload = self._build_valid_payload()
                payload["content_data"][0]["rows"] = rows
                with self.assertRaises(OutputLLMValidationError):
                    validate_output_llm(payload)

    def test_validate_output_llm_allows_empty_rows_per_sheet(self):
        output_json = self._build_valid_payload()
        output_json["content_data"][0]["rows"] = []

        result = validate_output_llm(output_json)

        self.assertEqual(result, output_json)

    def test_validate_output_llm_accepts_unicode_and_formula_like_strings(self):
        output_json = self._build_valid_payload()
        output_json["content_data"][0] = {
            "table_name": "Sheet1_Januari",
            "headers": ["item_name", "note"],
            "rows": [{"item_name": "शोफ़ी", "note": "=SUM(A1:A2)"}],
        }

        result = validate_output_llm(output_json)

        self.assertEqual(result, output_json)

    def test_validate_output_llm_handles_large_payload_smoke(self):
        rows = []
        for index in range(1000):
            rows.append({"item_name": f"user-{index}", "quantity": index, "price": index})

        output_json = self._build_valid_payload()
        output_json["content_data"][0]["rows"] = rows

        result = validate_output_llm(output_json)

        self.assertEqual(len(result["content_data"][0]["rows"]), 1000)

    def test_validate_output_llm_rejects_empty_output_object(self):
        with self.assertRaises(OutputLLMValidationError):
            validate_output_llm({})


class MapOutputCSVTest(unittest.TestCase):
    def _build_validated_output(self):
        return {
            "document_info": {
                "source_type": "Excel",
                "filename": "laporan_tahunan.xlsx",
            },
            "summary": {
                "grand_total": 1500000,
                "period": "2026",
            },
            "content_data": [
                {
                    "table_name": "Sheet1_Januari",
                    "headers": ["item_name", "quantity", "price"],
                    "rows": [
                        {"item_name": "Kertas", "quantity": 10, "price": 50000},
                        {"item_name": "Pena", "quantity": 5, "price": 10000},
                    ],
                }
            ],
        }

    def test_mapping_output_csv_maps_single_sheet_successfully(self):
        validated_output = self._build_validated_output()

        result = map_output_csv(validated_output)

        self.assertEqual(result["document_info"], validated_output["document_info"])
        self.assertEqual(result["summary"], validated_output["summary"])
        self.assertEqual(result["sheets"][0]["name"], "Sheet1_Januari")
        self.assertEqual(
            result["sheets"][0]["headers"],
            ["item_name", "quantity", "price"],
        )
        self.assertEqual(
            result["sheets"][0]["rows"],
            [["Kertas", 10, 50000], ["Pena", 5, 10000]],
        )

    def test_mapping_output_csv_maps_multiple_sheets_successfully(self):
        validated_output = self._build_validated_output()
        validated_output["content_data"].append(
            {
                "table_name": "Sheet2_Februari",
                "headers": ["item_name", "quantity", "price"],
                "rows": [{"item_name": "Tinta", "quantity": 1, "price": 400000}],
            }
        )

        result = map_output_csv(validated_output)

        self.assertEqual(len(result["sheets"]), 2)
        self.assertEqual(
            result["sheets"][1]["headers"],
            ["item_name", "quantity", "price"],
        )
        self.assertEqual(result["sheets"][1]["rows"], [["Tinta", 1, 400000]])

    def test_mapping_output_csv_allows_empty_rows(self):
        validated_output = self._build_validated_output()
        validated_output["content_data"][0]["rows"] = []

        result = map_output_csv(validated_output)

        self.assertEqual(
            result["sheets"][0]["headers"],
            ["item_name", "quantity", "price"],
        )
        self.assertEqual(result["sheets"][0]["rows"], [])

    def test_mapping_output_csv_keeps_unicode_and_formula_like_values(self):
        validated_output = self._build_validated_output()
        validated_output["content_data"][0]["headers"] = ["item_name", "note"]
        validated_output["content_data"][0]["rows"] = [
            {"item_name": "शोफ़ी", "note": "=SUM(A1:A2)"},
        ]

        result = map_output_csv(validated_output)

        self.assertEqual(result["sheets"][0]["rows"], [["शोफ़ी", "=SUM(A1:A2)"]])

    def test_mapping_output_csv_rejects_invalid_root_or_sheets(self):
        cases = [
            ("root_non_object", "invalid"),
            ("sheets_missing", {"status": "ok"}),
            ("content_data_non_list", {"content_data": {}}),
            ("table_item_non_object", {"content_data": ["invalid"]}),
        ]
        for case_name, payload in cases:
            with self.subTest(case=case_name):
                with self.assertRaises(OutputCSVMappingError):
                    map_output_csv(payload)

    def test_mapping_output_csv_rejects_sheet_missing_required_fields(self):
        cases = [
            (
                "missing_table_name",
                {"headers": ["item_name"], "rows": [{"item_name": "Kertas"}]},
            ),
            (
                "missing_headers",
                {"table_name": "Sheet1_Januari", "rows": [{"item_name": "Kertas"}]},
            ),
            ("missing_rows", {"table_name": "Sheet1_Januari", "headers": ["item_name"]}),
        ]
        for case_name, table_payload in cases:
            with self.subTest(case=case_name):
                validated_output = self._build_validated_output()
                validated_output["content_data"] = [table_payload]
                with self.assertRaises(OutputCSVMappingError):
                    map_output_csv(validated_output)

    def test_mapping_output_csv_rejects_invalid_columns_and_rows_container(self):
        cases = [
            ("headers_non_list", {"headers": "item_name"}),
            ("headers_empty", {"headers": []}),
            ("rows_non_list", {"rows": "not-list"}),
        ]
        for case_name, mutation in cases:
            with self.subTest(case=case_name):
                validated_output = self._build_validated_output()
                validated_output["content_data"][0].update(mutation)
                with self.assertRaises(OutputCSVMappingError):
                    map_output_csv(validated_output)

    def test_mapping_output_csv_rejects_invalid_row_items_and_key_mismatch(self):
        cases = [
            ("row_non_object", ["not-dict"]),
            ("missing_column", [{"item_name": "Kertas", "quantity": 10}]),
            (
                "unknown_column",
                [{"item_name": "Kertas", "quantity": 10, "price": 50000, "zip": 12345}],
            ),
        ]
        for case_name, rows in cases:
            with self.subTest(case=case_name):
                validated_output = self._build_validated_output()
                validated_output["content_data"][0]["rows"] = rows
                with self.assertRaises(OutputCSVMappingError):
                    map_output_csv(validated_output)

    def test_mapping_output_csv_does_not_mutate_input_payload(self):
        validated_output = self._build_validated_output()
        original = deepcopy(validated_output)

        _ = map_output_csv(validated_output)

        self.assertEqual(validated_output, original)

    def test_mapping_output_csv_handles_large_payload_smoke(self):
        rows = []
        for index in range(1000):
            rows.append(
                {"item_name": f"user-{index}", "quantity": index, "price": index}
            )

        validated_output = self._build_validated_output()
        validated_output["content_data"][0]["rows"] = rows

        result = map_output_csv(validated_output)

        self.assertEqual(len(result["sheets"][0]["rows"]), 1000)

    def test_mapping_output_csv_uses_stubbed_row_mapper(self):
        validated_output = self._build_validated_output()
        stubbed_rows = [["stubbed-row"]]

        with patch(
            "file_processing.services.export_service._map_rows",
            return_value=stubbed_rows,
        ) as mocked_map_rows:
            result = map_output_csv(validated_output)

        mocked_map_rows.assert_called_once_with(
            headers=["item_name", "quantity", "price"],
            rows=validated_output["content_data"][0]["rows"],
            table_index=0,
        )
        self.assertEqual(result["sheets"][0]["rows"], stubbed_rows)


class GenerateCSVTest(unittest.TestCase):
    class PrefixSanitizationPolicy:
        def sanitize_header(self, header):
            return f"HDR::{header}"

        def sanitize_value(self, value):
            if isinstance(value, str):
                return f"VAL::{value}"
            return value

    class PrefixFileNamePolicy:
        def build_filename(self, sheet_name):
            return f"csv_{sheet_name.lower()}.export.csv"

    class BlankFileNamePolicy:
        def build_filename(self, sheet_name):
            return "   "

    def _build_mapped_output(self):
        return {
            "sheets": [
                {
                    "name": "Sheet1",
                    "headers": ["name", "age", "city"],
                    "rows": [
                        ["Zufar", 21, "Depok"],
                        ["Siti", 22, "Jakarta"],
                    ],
                }
            ]
        }

    def _read_csv_rows(self, csv_content):
        return list(csv.reader(io.StringIO(csv_content)))

    def test_generate_csv_exposes_expected_api(self):
        self.assertTrue(
            hasattr(export_service, "generate_csv"),
            "generate_csv must be implemented in export_service.",
        )
        self.assertTrue(
            hasattr(export_service, "generate_csv_download_artifact"),
            "generate_csv_download_artifact must be implemented in export_service.",
        )
        self.assertTrue(
            hasattr(export_service, "OutputCSVGenerationError"),
            "OutputCSVGenerationError must be implemented in export_service.",
        )

    def test_generate_csv_single_sheet_successfully(self):
        mapped_output = self._build_mapped_output()

        result = export_service.generate_csv(mapped_output)

        self.assertEqual(len(result["files"]), 1)
        self.assertEqual(result["files"][0]["name"], "Sheet1.csv")
        self.assertEqual(
            self._read_csv_rows(result["files"][0]["content"]),
            [
                ["name", "age", "city"],
                ["Zufar", "21", "Depok"],
                ["Siti", "22", "Jakarta"],
            ],
        )

    def test_generate_csv_multiple_sheets_successfully(self):
        mapped_output = self._build_mapped_output()
        mapped_output["sheets"].append(
            {
                "name": "Sheet2",
                "headers": ["sku", "price"],
                "rows": [["A-1", 15000]],
            }
        )

        result = export_service.generate_csv(mapped_output)

        self.assertEqual(len(result["files"]), 2)
        self.assertEqual(result["files"][0]["name"], "Sheet1.csv")
        self.assertEqual(result["files"][1]["name"], "Sheet2.csv")
        self.assertEqual(
            self._read_csv_rows(result["files"][1]["content"]),
            [
                ["sku", "price"],
                ["A-1", "15000"],
            ],
        )

    def test_generate_csv_allows_empty_rows(self):
        mapped_output = self._build_mapped_output()
        mapped_output["sheets"][0]["rows"] = []

        result = export_service.generate_csv(mapped_output)

        self.assertEqual(
            self._read_csv_rows(result["files"][0]["content"]),
            [["name", "age", "city"]],
        )

    def test_generate_csv_rejects_invalid_root_or_sheets(self):
        cases = [
            ("root_non_object", "invalid"),
            ("sheets_missing", {}),
            ("sheets_non_list", {"sheets": {}}),
            ("sheet_non_object", {"sheets": ["invalid"]}),
        ]
        for case_name, payload in cases:
            with self.subTest(case=case_name):
                with self.assertRaises(export_service.OutputCSVGenerationError):
                    export_service.generate_csv(payload)

    def test_generate_csv_rejects_sheet_missing_required_fields(self):
        cases = [
            ("missing_name", {"headers": ["name"], "rows": [["Zufar"]]}),
            ("missing_headers", {"name": "Sheet1", "rows": [["Zufar"]]}),
            ("missing_rows", {"name": "Sheet1", "headers": ["name"]}),
        ]
        for case_name, sheet_payload in cases:
            with self.subTest(case=case_name):
                mapped_output = {"sheets": [sheet_payload]}
                with self.assertRaises(export_service.OutputCSVGenerationError):
                    export_service.generate_csv(mapped_output)

    def test_generate_csv_rejects_invalid_sheet_name(self):
        cases = [
            ("name_non_string", 123),
            ("name_blank", "   "),
        ]
        for case_name, name_value in cases:
            with self.subTest(case=case_name):
                mapped_output = self._build_mapped_output()
                mapped_output["sheets"][0]["name"] = name_value
                with self.assertRaises(export_service.OutputCSVGenerationError):
                    export_service.generate_csv(mapped_output)

    def test_generate_csv_rejects_invalid_headers_and_rows(self):
        cases = [
            ("headers_non_list", {"headers": "name"}),
            ("headers_empty", {"headers": []}),
            ("header_non_string", {"headers": ["name", 1]}),
            ("rows_non_list", {"rows": "not-list"}),
            ("row_non_list", {"rows": ["not-list-row"]}),
        ]
        for case_name, mutation in cases:
            with self.subTest(case=case_name):
                mapped_output = self._build_mapped_output()
                mapped_output["sheets"][0].update(mutation)
                with self.assertRaises(export_service.OutputCSVGenerationError):
                    export_service.generate_csv(mapped_output)

    def test_generate_csv_rejects_blank_or_duplicate_headers(self):
        cases = [
            ("blank_header", ["name", "   "], [["Zufar", "x"]]),
            ("duplicate_header_case_insensitive", ["Name", "name"], [["Zufar", "x"]]),
        ]
        for case_name, headers, rows in cases:
            with self.subTest(case=case_name):
                mapped_output = self._build_mapped_output()
                mapped_output["sheets"][0]["headers"] = headers
                mapped_output["sheets"][0]["rows"] = rows
                with self.assertRaises(export_service.OutputCSVGenerationError):
                    export_service.generate_csv(mapped_output)

    def test_generate_csv_rejects_row_length_mismatch(self):
        mapped_output = self._build_mapped_output()
        mapped_output["sheets"][0]["rows"] = [["Zufar", 21]]

        with self.assertRaises(export_service.OutputCSVGenerationError):
            export_service.generate_csv(mapped_output)

    def test_generate_csv_rejects_nested_or_unsupported_row_values(self):
        cases = [
            ("nested_dict_value", [["Zufar", {"age": 21}, "Depok"]]),
            ("nested_list_value", [["Zufar", [21], "Depok"]]),
            ("unsupported_set_value", [["Zufar", {21}, "Depok"]]),
        ]
        for case_name, rows in cases:
            with self.subTest(case=case_name):
                mapped_output = self._build_mapped_output()
                mapped_output["sheets"][0]["rows"] = rows
                with self.assertRaises(export_service.OutputCSVGenerationError):
                    export_service.generate_csv(mapped_output)

    def test_generate_csv_handles_csv_special_characters(self):
        mapped_output = {
            "sheets": [
                {
                    "name": "Sheet1",
                    "headers": ["text"],
                    "rows": [['Hello, "CSV"\nWorld']],
                }
            ]
        }

        result = export_service.generate_csv(mapped_output)

        self.assertEqual(
            self._read_csv_rows(result["files"][0]["content"]),
            [["text"], ['Hello, "CSV"\nWorld']],
        )

    def test_generate_csv_keeps_unicode_and_sanitizes_formula_like_values(self):
        mapped_output = {
            "sheets": [
                {
                    "name": "Sheet1",
                    "headers": ["name", "note"],
                    "rows": [["शोफ़ी", "=SUM(A1:A2)"]],
                }
            ]
        }

        result = export_service.generate_csv(mapped_output)

        self.assertEqual(
            self._read_csv_rows(result["files"][0]["content"]),
            [["name", "note"], ["शोफ़ी", "'=SUM(A1:A2)"]],
        )

    def test_generate_csv_sanitizes_formula_like_values_to_prevent_csv_injection(self):
        dangerous_values = [
            "=SUM(A1:A2)",
            "+CMD|'/C calc'!A0",
            "-10+20",
            "@HYPERLINK(\"http://evil.com\")",
        ]

        for value in dangerous_values:
            with self.subTest(value=value):
                mapped_output = {
                    "sheets": [
                        {
                            "name": "Sheet1",
                            "headers": ["name", "note"],
                            "rows": [["Zufar", value]],
                        }
                    ]
                }

                result = export_service.generate_csv(mapped_output)
                rows = self._read_csv_rows(result["files"][0]["content"])

                self.assertEqual(rows[1][1], f"'{value}")

    def test_generate_csv_keeps_whitespace_only_and_prequoted_values_unchanged(self):
        mapped_output = {
            "sheets": [
                {
                    "name": "Sheet1",
                    "headers": ["note", "formula"],
                    "rows": [["   ", "'=SUM(A1:A2)"]],
                }
            ]
        }

        result = export_service.generate_csv(mapped_output)

        self.assertEqual(
            self._read_csv_rows(result["files"][0]["content"]),
            [["note", "formula"], ["   ", "'=SUM(A1:A2)"]],
        )

    def test_generate_csv_sanitizes_formula_like_headers_to_prevent_csv_injection(self):
        mapped_output = {
            "sheets": [
                {
                    "name": "Sheet1",
                    "headers": ["=InjectedHeader", "name"],
                    "rows": [["x", "Zufar"]],
                }
            ]
        }

        result = export_service.generate_csv(mapped_output)
        rows = self._read_csv_rows(result["files"][0]["content"])

        self.assertEqual(rows[0][0], "'=InjectedHeader")

    def test_generate_csv_uses_mocked_csv_builder(self):
        mapped_output = self._build_mapped_output()
        mocked_csv_content = "name,age,city\r\nstub,0,stub-city\r\n"

        with patch(
            "file_processing.services.export_service._build_csv_content",
            return_value=mocked_csv_content,
        ) as mocked_builder:
            result = export_service.generate_csv(mapped_output)

        mocked_builder.assert_called_once_with(
            ["name", "age", "city"],
            [["Zufar", 21, "Depok"], ["Siti", 22, "Jakarta"]],
        )
        self.assertEqual(result["files"][0]["content"], mocked_csv_content)

    def test_generate_csv_supports_custom_sanitization_policy(self):
        mapped_output = self._build_mapped_output()
        policy = self.PrefixSanitizationPolicy()

        result = export_service.generate_csv(
            mapped_output,
            sanitization_policy=policy,
        )

        self.assertEqual(
            self._read_csv_rows(result["files"][0]["content"]),
            [
                ["HDR::name", "HDR::age", "HDR::city"],
                ["VAL::Zufar", "21", "VAL::Depok"],
                ["VAL::Siti", "22", "VAL::Jakarta"],
            ],
        )

    def test_generate_csv_rejects_invalid_sanitization_policy_contract(self):
        mapped_output = self._build_mapped_output()

        with self.assertRaises(export_service.OutputCSVGenerationError):
            export_service.generate_csv(
                mapped_output,
                sanitization_policy=object(),
            )

    def test_generate_csv_supports_custom_filename_policy(self):
        mapped_output = self._build_mapped_output()
        filename_policy = self.PrefixFileNamePolicy()

        result = export_service.generate_csv(
            mapped_output,
            filename_policy=filename_policy,
        )

        self.assertEqual(result["files"][0]["name"], "csv_sheet1.export.csv")

    def test_generate_csv_rejects_invalid_filename_policy_contract(self):
        mapped_output = self._build_mapped_output()

        with self.assertRaises(export_service.OutputCSVGenerationError):
            export_service.generate_csv(
                mapped_output,
                filename_policy=object(),
            )

    def test_generate_csv_rejects_blank_filename_from_policy(self):
        mapped_output = self._build_mapped_output()
        filename_policy = self.BlankFileNamePolicy()

        with self.assertRaises(export_service.OutputCSVGenerationError):
            export_service.generate_csv(
                mapped_output,
                filename_policy=filename_policy,
            )

    def test_generate_csv_normalizes_path_traversal_member_filenames(self):
        cases = [
            ("../evil.csv", "evil.csv"),
            ("folder/evil.csv", "evil.csv"),
            ("folder\\evil.csv", "evil.csv"),
        ]
        
        for unsafe_name, expected_name in cases:
            with self.subTest(unsafe_name=unsafe_name):
                class UnsafeFilenamePolicy:
                    def build_filename(self, sheet_name):
                        return unsafe_name
                        
                mapped_output = self._build_mapped_output()
                result = export_service.generate_csv(
                    mapped_output,
                    filename_policy=UnsafeFilenamePolicy(),
                )
                
                self.assertEqual(result["files"][0]["name"], expected_name)
                self.assertIn("name", result["files"][0]["content"])

    def test_generate_csv_replaces_invalid_filesystem_characters(self):
        cases = [
            ("file*.csv", "file_.csv"),
            ("file?.csv", "file_.csv"),
            ('file".csv', "file_.csv"),
            ("file:.csv", "file_.csv"),
            ("file<.csv", "file_.csv"),
            ("file>.csv", "file_.csv"),
            ("file|.csv", "file_.csv"),
            ("my*awesome?file.csv", "my_awesome_file.csv"),
        ]
        
        for unsafe_name, expected_name in cases:
            with self.subTest(unsafe_name=unsafe_name):
                class UnsafeFilenamePolicy:
                    def build_filename(self, sheet_name):
                        return unsafe_name
                        
                mapped_output = self._build_mapped_output()
                result = export_service.generate_csv(
                    mapped_output,
                    filename_policy=UnsafeFilenamePolicy(),
                )
                
                self.assertEqual(result["files"][0]["name"], expected_name)

    def test_generate_csv_deduplicates_member_filenames_deterministically(self):
        mapped_output = {
            "sheets": [
                {"name": "Sheet", "headers": ["h"], "rows": []},
                {"name": "Sheet", "headers": ["h"], "rows": []},
                {"name": "folder/Sheet", "headers": ["h"], "rows": []},
                {"name": 'S*h?e:e"t', "headers": ["h"], "rows": []},
                {"name": "S_h_e_e_t", "headers": ["h"], "rows": []},
            ]
        }
        
        result = export_service.generate_csv(mapped_output)
        filenames = [file["name"] for file in result["files"]]
        
        expected = [
            "Sheet.csv",
            "Sheet_1.csv",
            "Sheet_2.csv",
            "S_h_e_e_t.csv",
            "S_h_e_e_t_1.csv",
        ]
        
        self.assertEqual(filenames, expected)

    def test_generate_csv_download_artifact_single_sheet_returns_plain_csv(self):
        mapped_output = self._build_mapped_output()

        result = export_service.generate_csv_download_artifact(mapped_output)

        self.assertEqual(result["type"], "csv")
        self.assertEqual(result["name"], "Sheet1.csv")
        rows = self._read_csv_rows(result["content"].decode("utf-8"))
        self.assertEqual(
            rows,
            [
                ["name", "age", "city"],
                ["Zufar", "21", "Depok"],
                ["Siti", "22", "Jakarta"],
            ],
        )

    def test_generate_csv_download_artifact_multiple_sheets_returns_zip(self):
        mapped_output = self._build_mapped_output()
        mapped_output["sheets"].append(
            {
                "name": "Sheet2",
                "headers": ["sku", "price"],
                "rows": [["A-1", 15000]],
            }
        )

        result = export_service.generate_csv_download_artifact(mapped_output)

        self.assertEqual(result["type"], "zip")
        self.assertEqual(result["name"], "csv_export.zip")
        with zipfile.ZipFile(io.BytesIO(result["content"]), "r") as archive:
            names = sorted(archive.namelist())
            self.assertEqual(names, ["Sheet1.csv", "Sheet2.csv"])

            sheet2_rows = list(
                csv.reader(io.StringIO(archive.read("Sheet2.csv").decode("utf-8")))
            )
            self.assertEqual(
                sheet2_rows,
                [
                    ["sku", "price"],
                    ["A-1", "15000"],
                ],
            )


class ExportCSVToFilesystemTest(unittest.TestCase):
    def _build_valid_output_json(self):
        return {
            "document_info": {
                "source_type": "Excel",
                "filename": "laporan_tahunan.xlsx",
            },
            "summary": {
                "grand_total": 1500000,
                "period": "2026",
            },
            "content_data": [
                {
                    "table_name": "Sheet1_Januari",
                    "headers": ["item_name", "quantity", "price"],
                    "rows": [
                        {"item_name": "Kertas", "quantity": 10, "price": 50000},
                        {"item_name": "Pena", "quantity": 5, "price": 10000},
                    ],
                }
            ],
        }

    def test_export_csv_to_filesystem_exposes_expected_api(self):
        self.assertTrue(
            hasattr(export_service, "export_csv_to_filesystem"),
            "export_csv_to_filesystem must be implemented in export_service.",
        )

    def test_export_csv_to_filesystem_saves_single_sheet_as_csv(self):
        output_json = self._build_valid_output_json()
        with tempfile.TemporaryDirectory() as temp_dir:
            result = export_service.export_csv_to_filesystem(
                output_json=output_json,
                storage_dir=temp_dir,
            )

            self.assertTrue(re.match(r"^csv_[a-z0-9]{32}$", result["file_id"]))
            self.assertEqual(result["artifact_type"], "csv")
            self.assertTrue(result["file_name"].startswith("export_"))
            self.assertTrue(result["file_name"].endswith(".csv"))
            self.assertGreater(result["size_bytes"], 0)
            self.assertIn("created_at", result)

            file_path = os.path.join(temp_dir, result["file_name"])
            self.assertTrue(os.path.exists(file_path))
            with open(file_path, encoding="utf-8") as generated:
                rows = list(csv.reader(generated))
            self.assertEqual(rows[0], ["item_name", "quantity", "price"])

    def test_export_csv_to_filesystem_saves_multi_sheet_as_zip(self):
        output_json = self._build_valid_output_json()
        output_json["content_data"].append(
            {
                "table_name": "Sheet2_Februari",
                "headers": ["item_name", "quantity", "price"],
                "rows": [{"item_name": "Tinta", "quantity": 1, "price": 400000}],
            }
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            result = export_service.export_csv_to_filesystem(
                output_json=output_json,
                storage_dir=temp_dir,
            )

            self.assertEqual(result["artifact_type"], "zip")
            self.assertTrue(result["file_name"].startswith("export_"))
            self.assertTrue(result["file_name"].endswith(".zip"))

            file_path = os.path.join(temp_dir, result["file_name"])
            self.assertTrue(os.path.exists(file_path))
            with zipfile.ZipFile(file_path, "r") as archive:
                self.assertEqual(sorted(archive.namelist()), ["Sheet1_Januari.csv", "Sheet2_Februari.csv"])

    def test_export_csv_to_filesystem_rejects_invalid_output_json(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(OutputLLMValidationError):
                export_service.export_csv_to_filesystem(
                    output_json={},
                    storage_dir=temp_dir,
                )

    def test_export_csv_to_filesystem_rejects_invalid_storage_dir(self):
        output_json = self._build_valid_output_json()

        with self.assertRaises(export_service.OutputCSVGenerationError):
            export_service.export_csv_to_filesystem(
                output_json=output_json,
                storage_dir="",
            )

    def test_export_csv_to_filesystem_supports_test_doubles_for_id_and_timestamp(self):
        output_json = self._build_valid_output_json()
        with tempfile.TemporaryDirectory() as temp_dir:
            result = export_service.export_csv_to_filesystem(
                output_json=output_json,
                storage_dir=temp_dir,
                token_generator=lambda: "abc123",
                now_provider=lambda: "2026-03-07T10:00:00Z",
            )

            self.assertEqual(result["file_id"], "csv_abc123")
            self.assertEqual(result["created_at"], "2026-03-07T10:00:00Z")
            self.assertEqual(result["file_name"], "export_abc123.csv")

    def test_export_csv_to_filesystem_generates_unique_file_id(self):
        output_json = self._build_valid_output_json()
        with tempfile.TemporaryDirectory() as temp_dir:
            first = export_service.export_csv_to_filesystem(
                output_json=output_json,
                storage_dir=temp_dir,
            )
            second = export_service.export_csv_to_filesystem(
                output_json=output_json,
                storage_dir=temp_dir,
            )

            self.assertNotEqual(first["file_id"], second["file_id"])
            self.assertNotEqual(first["file_name"], second["file_name"])

    @patch("file_processing.services.export_service.open")
    def test_export_csv_to_filesystem_raises_when_writing_file_fails(self, mocked_open):
        output_json = self._build_valid_output_json()
        mocked_open.side_effect = OSError("disk write failed")

        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(export_service.OutputCSVGenerationError):
                export_service.export_csv_to_filesystem(
                    output_json=output_json,
                    storage_dir=temp_dir,
                )

    @patch("file_processing.services.export_service.os.makedirs")
    def test_export_csv_to_filesystem_raises_when_storage_dir_creation_fails(
        self,
        mocked_makedirs,
    ):
        output_json = self._build_valid_output_json()
        mocked_makedirs.side_effect = OSError("cannot create dir")

        with self.assertRaises(export_service.OutputCSVGenerationError):
            export_service.export_csv_to_filesystem(
                output_json=output_json,
                storage_dir="C:/invalid/dir",
            )

    def test_export_csv_to_filesystem_rejects_empty_token_from_generator(self):
        output_json = self._build_valid_output_json()
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(export_service.OutputCSVGenerationError):
                export_service.export_csv_to_filesystem(
                    output_json=output_json,
                    storage_dir=temp_dir,
                    token_generator=lambda: "   ",
                )

    def test_export_csv_to_filesystem_rejects_unsafe_token_from_generator(self):
        output_json = self._build_valid_output_json()
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(export_service.OutputCSVGenerationError):
                export_service.export_csv_to_filesystem(
                    output_json=output_json,
                    storage_dir=temp_dir,
                    token_generator=lambda: "abc-123",
                )

    def test_export_csv_to_filesystem_rejects_invalid_now_provider_value(self):
        output_json = self._build_valid_output_json()
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(export_service.OutputCSVGenerationError):
                export_service.export_csv_to_filesystem(
                    output_json=output_json,
                    storage_dir=temp_dir,
                    now_provider=lambda: "",
                )

    def test_build_safe_file_path_rejects_path_traversal(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(export_service.OutputCSVGenerationError):
                export_service._build_safe_file_path(temp_dir, "../evil.csv", export_service.OutputCSVGenerationError)

    def test_build_safe_file_path_raises_when_commonpath_fails(self):
        with patch(
            "file_processing.services.export_service.os.path.commonpath",
            side_effect=ValueError("invalid path roots"),
        ):
            with self.assertRaises(export_service.OutputCSVGenerationError):
                export_service._build_safe_file_path(r"C:\safe\storage", "export_abc123.csv", export_service.OutputCSVGenerationError)


class ExportExcelToFilesystemTest(unittest.TestCase):
    def _build_valid_output_json(self):
        return {
            "document_info": {
                "source_type": "Excel",
                "filename": "laporan_tahunan.xlsx",
            },
            "summary": {
                "grand_total": 1500000,
                "period": "2026",
            },
            "content_data": [
                {
                    "table_name": "Sheet1_Januari",
                    "headers": ["item_name", "quantity", "price"],
                    "rows": [
                        {"item_name": "Kertas", "quantity": 10, "price": 50000},
                    ],
                },
                {
                    "table_name": "Sheet2_Februari",
                    "headers": ["item_name", "quantity", "price"],
                    "rows": [
                        {"item_name": "Tinta", "quantity": 1, "price": 400000},
                    ],
                },
            ],
        }

    def _read_xlsx_sheet_names(self, file_path):
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            return workbook.sheetnames
        finally:
            workbook.close()

    def _count_xlsx_worksheet_files(self, file_path):
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            return len(workbook.worksheets)
        finally:
            workbook.close()

    def _read_xlsx_sheet_rows(self, file_path, target_sheet_name):
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook[target_sheet_name]
            return [list(row) for row in worksheet.iter_rows(values_only=True)]
        finally:
            workbook.close()

    def test_export_excel_to_filesystem_exposes_expected_api(self):
        self.assertTrue(
            hasattr(export_service, "export_excel_to_filesystem"),
            "export_excel_to_filesystem must be implemented in export_service.",
        )
        self.assertTrue(
            hasattr(export_service, "OutputExcelGenerationError"),
            "OutputExcelGenerationError must be implemented in export_service.",
        )

    def test_export_excel_to_filesystem_saves_multi_sheet_as_single_xlsx(self):
        output_json = self._build_valid_output_json()
        with tempfile.TemporaryDirectory() as temp_dir:
            result = export_service.export_excel_to_filesystem(
                output_json=output_json,
                storage_dir=temp_dir,
                token_generator=lambda: "abc123",
                now_provider=lambda: "2026-03-08T10:00:00Z",
            )

            self.assertEqual(result["file_id"], "xlsx_abc123")
            self.assertEqual(result["file_name"], "export_abc123.xlsx")
            self.assertEqual(result["artifact_type"], "xlsx")
            self.assertEqual(result["created_at"], "2026-03-08T10:00:00Z")
            self.assertGreater(result["size_bytes"], 0)

            file_path = os.path.join(temp_dir, result["file_name"])
            self.assertTrue(os.path.exists(file_path))
            self.assertEqual(self._count_xlsx_worksheet_files(file_path), 2)
            self.assertEqual(
                self._read_xlsx_sheet_names(file_path),
                ["Sheet1_Januari", "Sheet2_Februari"],
            )

    def test_export_excel_to_filesystem_sanitizes_formula_like_headers_and_values(self):
        output_json = self._build_valid_output_json()
        output_json["content_data"] = [
            {
                "table_name": "FormulaSheet",
                "headers": ["=header", "note"],
                "rows": [
                    {"=header": "=SUM(A1:A2)", "note": "@danger"},
                ],
            }
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            result = export_service.export_excel_to_filesystem(
                output_json=output_json,
                storage_dir=temp_dir,
                token_generator=lambda: "safe01",
            )

            file_path = os.path.join(temp_dir, result["file_name"])
            rows = self._read_xlsx_sheet_rows(file_path, "FormulaSheet")

            self.assertEqual(rows[0], ["'=header", "note"])
            self.assertEqual(rows[1], ["'=SUM(A1:A2)", "'@danger"])

    def test_export_excel_to_filesystem_rejects_invalid_output_json(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(OutputLLMValidationError):
                export_service.export_excel_to_filesystem(
                    output_json={},
                    storage_dir=temp_dir,
                )

    def test_export_excel_to_filesystem_rejects_invalid_storage_dir(self):
        output_json = self._build_valid_output_json()

        with self.assertRaises(export_service.OutputExcelGenerationError):
            export_service.export_excel_to_filesystem(
                output_json=output_json,
                storage_dir="",
            )

    @patch("file_processing.services.export_service.open")
    def test_export_excel_to_filesystem_raises_when_writing_file_fails(self, mocked_open):
        output_json = self._build_valid_output_json()
        mocked_open.side_effect = OSError("disk write failed")

        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(export_service.OutputExcelGenerationError):
                export_service.export_excel_to_filesystem(
                    output_json=output_json,
                    storage_dir=temp_dir,
                )

    def test_export_excel_to_filesystem_rejects_empty_token_from_generator(self):
        output_json = self._build_valid_output_json()
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(export_service.OutputExcelGenerationError):
                export_service.export_excel_to_filesystem(
                    output_json=output_json,
                    storage_dir=temp_dir,
                    token_generator=lambda: "   ",
                )

    def test_export_excel_to_filesystem_rejects_unsafe_token_from_generator(self):
        output_json = self._build_valid_output_json()
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(export_service.OutputExcelGenerationError):
                export_service.export_excel_to_filesystem(
                    output_json=output_json,
                    storage_dir=temp_dir,
                    token_generator=lambda: "bad-token",
                )

    def test_export_excel_to_filesystem_rejects_invalid_now_provider_value(self):
        output_json = self._build_valid_output_json()
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(export_service.OutputExcelGenerationError):
                export_service.export_excel_to_filesystem(
                    output_json=output_json,
                    storage_dir=temp_dir,
                    now_provider=lambda: "",
                )


class ExcelExportInternalHelperCoverageTest(unittest.TestCase):
    class NoOpSanitizationPolicy:
        def sanitize_header(self, header):
            return header

        def sanitize_value(self, value):
            return value

    class BrokenWorkbook:
        def __init__(self):
            self.closed = False

        def save(self, _buffer):
            raise RuntimeError("save failure")

        def close(self):
            self.closed = True

    class _FakeWorksheet:
        def __init__(self, title):
            self.title = title
            self.rows = []

        def append(self, row):
            self.rows.append(row)

    class _FakeWorkbookWithoutDefaultSheet:
        def __init__(self):
            self.worksheets = []
            self.active = None
            self.removed_called = False

        def remove(self, _worksheet):
            self.removed_called = True

        def create_sheet(self, title):
            worksheet = ExcelExportInternalHelperCoverageTest._FakeWorksheet(title)
            self.worksheets.append(worksheet)
            return worksheet

        def close(self):
            return None

    def test_validate_excel_mapped_output_rejects_non_object_root(self):
        with self.assertRaises(export_service.OutputExcelGenerationError):
            export_service._validate_excel_mapped_output("invalid")

    def test_validate_excel_mapped_output_rejects_non_list_sheets(self):
        with self.assertRaises(export_service.OutputExcelGenerationError):
            export_service._validate_excel_mapped_output({"sheets": "invalid"})

    def test_serialize_excel_workbook_wraps_unexpected_error_and_closes_workbook(self):
        workbook = self.BrokenWorkbook()

        with self.assertRaises(export_service.OutputExcelGenerationError):
            export_service._serialize_excel_workbook(workbook)

        self.assertTrue(workbook.closed)

    def test_build_excel_workbook_raises_clear_error_when_openpyxl_missing(self):
        original_import = __import__

        def _import_stub(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("openpyxl not installed")
            return original_import(name, *args, **kwargs)

        with patch("builtins.__import__", side_effect=_import_stub):
            with self.assertRaises(export_service.OutputExcelGenerationError):
                export_service._build_excel_workbook(
                    sheets=[],
                    sanitization_policy=self.NoOpSanitizationPolicy(),
                )

    def test_build_excel_workbook_creates_default_sheet_when_sheet_input_is_empty(self):
        workbook = export_service._build_excel_workbook(
            sheets=[],
            sanitization_policy=self.NoOpSanitizationPolicy(),
        )
        try:
            self.assertEqual(workbook.sheetnames, ["Sheet1"])
        finally:
            workbook.close()

    def test_build_excel_workbook_skips_default_sheet_removal_when_workbook_starts_empty(self):
        original_import = __import__

        def _import_stub(name, *args, **kwargs):
            if name == "openpyxl":
                return types.SimpleNamespace(
                    Workbook=self._FakeWorkbookWithoutDefaultSheet
                )
            return original_import(name, *args, **kwargs)

        with patch("builtins.__import__", side_effect=_import_stub):
            workbook = export_service._build_excel_workbook(
                sheets=[
                    {
                        "name": "SheetA",
                        "headers": ["col1"],
                        "rows": [["value1"]],
                    }
                ],
                sanitization_policy=self.NoOpSanitizationPolicy(),
            )

        self.assertEqual(len(workbook.worksheets), 1)
        self.assertEqual(workbook.worksheets[0].title, "SheetA")
        self.assertFalse(workbook.removed_called)

    def test_normalize_excel_sheet_name_uses_default_when_normalized_is_blank(self):
        seen_names = set()

        result = export_service._normalize_excel_sheet_name("   ", seen_names)

        self.assertEqual(result, "Sheet1")
        self.assertIn("sheet1", seen_names)

    def test_normalize_excel_sheet_name_increments_suffix_until_unique(self):
        seen_names = {"sheet", "sheet_1"}

        result = export_service._normalize_excel_sheet_name("Sheet", seen_names)

        self.assertEqual(result, "Sheet_2")
        self.assertIn("sheet_2", seen_names)


class ResolveCSVDownloadArtifactTest(unittest.TestCase):
    def test_resolve_csv_download_artifact_exposes_expected_api(self):
        self.assertTrue(
            hasattr(export_service, "resolve_csv_download_artifact"),
            "resolve_csv_download_artifact must be implemented in export_service.",
        )
        self.assertTrue(
            hasattr(export_service, "OutputCSVDownloadLookupError"),
            "OutputCSVDownloadLookupError must be implemented in export_service.",
        )

    def test_resolve_csv_download_artifact_returns_csv_metadata_for_existing_file(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            token = "abc123"
            file_name = f"export_{token}.csv"
            file_path = os.path.join(temp_dir, file_name)
            with open(file_path, "w", encoding="utf-8") as generated:
                generated.write("name,age\r\nZufar,21\r\n")

            result = export_service.resolve_csv_download_artifact(
                file_id=f"csv_{token}",
                storage_dir=temp_dir,
            )

            self.assertEqual(result["artifact_type"], "csv")
            self.assertEqual(result["file_name"], file_name)
            self.assertEqual(os.path.realpath(result["file_path"]), os.path.realpath(file_path))
            self.assertEqual(result["content_type"], "text/csv")

    def test_resolve_csv_download_artifact_returns_zip_metadata_for_existing_file(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            token = "abc123"
            file_name = f"export_{token}.zip"
            file_path = os.path.join(temp_dir, file_name)
            with zipfile.ZipFile(file_path, "w", zipfile.ZIP_DEFLATED) as archive:
                archive.writestr("Sheet1.csv", "name,age\r\nZufar,21\r\n")

            result = export_service.resolve_csv_download_artifact(
                file_id=f"csv_{token}",
                storage_dir=temp_dir,
            )

            self.assertEqual(result["artifact_type"], "zip")
            self.assertEqual(result["file_name"], file_name)
            self.assertEqual(os.path.realpath(result["file_path"]), os.path.realpath(file_path))
            self.assertEqual(result["content_type"], "application/zip")

    def test_resolve_csv_download_artifact_rejects_invalid_file_id(self):
        invalid_file_ids = [
            "",
            "csv_",
            "csv_abc-123",
            "../evil",
            "C:\\evil",
            "/tmp/evil",
            "abc123",
        ]

        for file_id in invalid_file_ids:
            with self.subTest(file_id=file_id):
                with self.assertRaises(export_service.OutputCSVDownloadLookupError):
                    export_service.resolve_csv_download_artifact(
                        file_id=file_id,
                        storage_dir="C:/safe/storage",
                    )

    def test_resolve_csv_download_artifact_raises_not_found_when_file_missing(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(export_service.OutputCSVDownloadLookupError):
                export_service.resolve_csv_download_artifact(
                    file_id="csv_deadbeef",
                    storage_dir=temp_dir,
                )

    def test_resolve_csv_download_artifact_prefers_csv_before_zip(self):
        token = "abc123"
        with tempfile.TemporaryDirectory() as temp_dir:
            csv_file_name = f"export_{token}.csv"
            zip_file_name = f"export_{token}.zip"
            csv_file_path = os.path.join(temp_dir, csv_file_name)
            zip_file_path = os.path.join(temp_dir, zip_file_name)

            with open(csv_file_path, "w", encoding="utf-8") as generated:
                generated.write("name,age\r\nZufar,21\r\n")
            with zipfile.ZipFile(zip_file_path, "w", zipfile.ZIP_DEFLATED) as archive:
                archive.writestr("Sheet1.csv", "name,age\r\nZufar,21\r\n")

            result = export_service.resolve_csv_download_artifact(
                file_id=f"csv_{token}",
                storage_dir=temp_dir,
            )

            self.assertEqual(result["artifact_type"], "csv")
            self.assertEqual(result["file_name"], csv_file_name)
            self.assertEqual(
                os.path.realpath(result["file_path"]),
                os.path.realpath(csv_file_path),
            )
            self.assertEqual(result["content_type"], "text/csv")

    def test_resolve_csv_download_artifact_rejects_invalid_storage_dir(self):
        with self.assertRaises(export_service.OutputCSVDownloadLookupError):
            export_service.resolve_csv_download_artifact(
                file_id="csv_abc123",
                storage_dir="",
            )


class _StaticScandir:
    def __init__(self, entries):
        self._entries = entries

    def __enter__(self):
        return self._entries

    def __exit__(self, exc_type, exc_val, exc_tb):
        return False


class _FakeDirEntry:
    def __init__(self, name, path, is_file_result):
        self.name = name
        self.path = path
        self._is_file_result = is_file_result

    def is_file(self, follow_symlinks=False):
        return self._is_file_result


class DiscoverDownloadArtifactsTest(unittest.TestCase):
    def test_discover_download_artifacts_skips_non_file_entries(self):
        base_dir = r"C:\safe\storage"
        entries = [
            _FakeDirEntry("export_abc123.csv", r"C:\safe\storage\export_abc123.csv", False),
            _FakeDirEntry("export_def456.csv", r"C:\safe\storage\export_def456.csv", True),
        ]

        with patch(
            "file_processing.services.export_service.os.scandir",
            return_value=_StaticScandir(entries),
        ):
            with patch(
                "file_processing.services.export_service.os.path.realpath",
                side_effect=lambda value: value,
            ):
                with patch(
                    "file_processing.services.export_service.os.path.commonpath",
                    side_effect=lambda paths: paths[0],
                ):
                    discovered = export_service._discover_download_artifacts(base_dir)

        self.assertNotIn("export_abc123.csv", discovered)
        self.assertIn("export_def456.csv", discovered)

    def test_discover_download_artifacts_skips_non_matching_file_names(self):
        base_dir = r"C:\safe\storage"
        entries = [
            _FakeDirEntry("notes.txt", r"C:\safe\storage\notes.txt", True),
            _FakeDirEntry("export_def456.csv", r"C:\safe\storage\export_def456.csv", True),
        ]

        with patch(
            "file_processing.services.export_service.os.scandir",
            return_value=_StaticScandir(entries),
        ):
            with patch(
                "file_processing.services.export_service.os.path.realpath",
                side_effect=lambda value: value,
            ):
                with patch(
                    "file_processing.services.export_service.os.path.commonpath",
                    side_effect=lambda paths: paths[0],
                ):
                    discovered = export_service._discover_download_artifacts(base_dir)

        self.assertNotIn("notes.txt", discovered)
        self.assertIn("export_def456.csv", discovered)

    def test_discover_download_artifacts_skips_entries_when_commonpath_raises(self):
        base_dir = r"C:\safe\storage"
        entries = [
            _FakeDirEntry("export_abc123.csv", r"C:\safe\storage\export_abc123.csv", True),
            _FakeDirEntry("export_def456.csv", r"C:\safe\storage\export_def456.csv", True),
        ]

        with patch(
            "file_processing.services.export_service.os.scandir",
            return_value=_StaticScandir(entries),
        ):
            with patch(
                "file_processing.services.export_service.os.path.realpath",
                side_effect=lambda value: value,
            ):
                with patch(
                    "file_processing.services.export_service.os.path.commonpath",
                    side_effect=[ValueError("invalid path"), base_dir],
                ):
                    discovered = export_service._discover_download_artifacts(base_dir)

        self.assertNotIn("export_abc123.csv", discovered)
        self.assertIn("export_def456.csv", discovered)

    def test_discover_download_artifacts_skips_entries_outside_base_dir(self):
        base_dir = r"C:\safe\storage"
        entries = [
            _FakeDirEntry("export_abc123.csv", r"C:\other\export_abc123.csv", True),
        ]

        with patch(
            "file_processing.services.export_service.os.scandir",
            return_value=_StaticScandir(entries),
        ):
            with patch(
                "file_processing.services.export_service.os.path.realpath",
                side_effect=lambda value: value,
            ):
                with patch(
                    "file_processing.services.export_service.os.path.commonpath",
                    return_value=r"C:\other",
                ):
                    discovered = export_service._discover_download_artifacts(base_dir)

        self.assertEqual(discovered, {})

    def test_discover_download_artifacts_raises_lookup_error_when_scandir_fails(self):
        with patch(
            "file_processing.services.export_service.os.scandir",
            side_effect=OSError("storage not readable"),
        ):
            with self.assertRaises(export_service.OutputCSVDownloadLookupError):
                export_service._discover_download_artifacts(r"C:\safe\storage")


class ResolveExcelDownloadArtifactTest(unittest.TestCase):
    def test_resolve_excel_download_artifact_exposes_expected_api(self):
        self.assertTrue(
            hasattr(export_service, "resolve_excel_download_artifact"),
            "resolve_excel_download_artifact must be implemented in export_service.",
        )
        self.assertTrue(
            hasattr(export_service, "OutputExcelDownloadLookupError"),
            "OutputExcelDownloadLookupError must be implemented in export_service.",
        )

    def test_resolve_excel_download_artifact_returns_xlsx_metadata_for_existing_file(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            token = "abc123"
            file_name = f"export_{token}.xlsx"
            file_path = os.path.join(temp_dir, file_name)
            with open(file_path, "wb") as generated:
                generated.write(b"fake xlsx bytes")

            result = export_service.resolve_excel_download_artifact(
                export_id=f"xlsx_{token}",
                storage_dir=temp_dir,
            )

            self.assertEqual(result["artifact_type"], "xlsx")
            self.assertEqual(result["file_name"], file_name)
            self.assertEqual(os.path.realpath(result["file_path"]), os.path.realpath(file_path))
            self.assertEqual(
                result["content_type"],
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    def test_resolve_excel_download_artifact_rejects_invalid_export_id(self):
        invalid_export_ids = [
            "",
            "xlsx_",
            "xlsx_abc-123",
            "../evil",
            "C:\\evil",
            "/tmp/evil",
            "abc123",
        ]

        for export_id in invalid_export_ids:
            with self.subTest(export_id=export_id):
                with self.assertRaises(export_service.OutputExcelDownloadLookupError):
                    export_service.resolve_excel_download_artifact(
                        export_id=export_id,
                        storage_dir="C:/safe/storage",
                    )

    def test_resolve_excel_download_artifact_raises_not_found_when_file_missing(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(export_service.OutputExcelDownloadLookupError):
                export_service.resolve_excel_download_artifact(
                    export_id="xlsx_deadbeef",
                    storage_dir=temp_dir,
                )

    def test_resolve_excel_download_artifact_rejects_invalid_storage_dir(self):
        with self.assertRaises(export_service.OutputExcelDownloadLookupError):
            export_service.resolve_excel_download_artifact(
                export_id="xlsx_abc123",
                storage_dir="",
            )


class DiscoverExcelDownloadArtifactsTest(unittest.TestCase):
    def test_discover_excel_download_artifacts_skips_non_file_entries(self):
        base_dir = r"C:\safe\storage"
        entries = [
            _FakeDirEntry("export_abc123.xlsx", r"C:\safe\storage\export_abc123.xlsx", False),
            _FakeDirEntry("export_def456.xlsx", r"C:\safe\storage\export_def456.xlsx", True),
        ]

        with patch(
            "file_processing.services.export_service.os.scandir",
            return_value=_StaticScandir(entries),
        ):
            with patch(
                "file_processing.services.export_service.os.path.realpath",
                side_effect=lambda value: value,
            ):
                with patch(
                    "file_processing.services.export_service.os.path.commonpath",
                    side_effect=lambda paths: paths[0],
                ):
                    discovered = export_service._discover_excel_download_artifacts(
                        base_dir
                    )

        self.assertNotIn("export_abc123.xlsx", discovered)
        self.assertIn("export_def456.xlsx", discovered)

    def test_discover_excel_download_artifacts_skips_non_matching_file_names(self):
        base_dir = r"C:\safe\storage"
        entries = [
            _FakeDirEntry("notes.txt", r"C:\safe\storage\notes.txt", True),
            _FakeDirEntry("export_bad.csv", r"C:\safe\storage\export_bad.csv", True),
            _FakeDirEntry("export_def456.xlsx", r"C:\safe\storage\export_def456.xlsx", True),
        ]

        with patch(
            "file_processing.services.export_service.os.scandir",
            return_value=_StaticScandir(entries),
        ):
            with patch(
                "file_processing.services.export_service.os.path.realpath",
                side_effect=lambda value: value,
            ):
                with patch(
                    "file_processing.services.export_service.os.path.commonpath",
                    side_effect=lambda paths: paths[0],
                ):
                    discovered = export_service._discover_excel_download_artifacts(
                        base_dir
                    )

        self.assertNotIn("notes.txt", discovered)
        self.assertNotIn("export_bad.csv", discovered)
        self.assertIn("export_def456.xlsx", discovered)

    def test_discover_excel_download_artifacts_skips_entries_when_commonpath_raises(self):
        base_dir = r"C:\safe\storage"
        entries = [
            _FakeDirEntry("export_abc123.xlsx", r"C:\safe\storage\export_abc123.xlsx", True),
            _FakeDirEntry("export_def456.xlsx", r"C:\safe\storage\export_def456.xlsx", True),
        ]

        with patch(
            "file_processing.services.export_service.os.scandir",
            return_value=_StaticScandir(entries),
        ):
            with patch(
                "file_processing.services.export_service.os.path.realpath",
                side_effect=lambda value: value,
            ):
                with patch(
                    "file_processing.services.export_service.os.path.commonpath",
                    side_effect=[ValueError("invalid path"), base_dir],
                ):
                    discovered = export_service._discover_excel_download_artifacts(
                        base_dir
                    )

        self.assertNotIn("export_abc123.xlsx", discovered)
        self.assertIn("export_def456.xlsx", discovered)

    def test_discover_excel_download_artifacts_skips_entries_outside_base_dir(self):
        base_dir = r"C:\safe\storage"
        entries = [
            _FakeDirEntry("export_abc123.xlsx", r"C:\other\export_abc123.xlsx", True),
        ]

        with patch(
            "file_processing.services.export_service.os.scandir",
            return_value=_StaticScandir(entries),
        ):
            with patch(
                "file_processing.services.export_service.os.path.realpath",
                side_effect=lambda value: value,
            ):
                with patch(
                    "file_processing.services.export_service.os.path.commonpath",
                    return_value=r"C:\other",
                ):
                    discovered = export_service._discover_excel_download_artifacts(
                        base_dir
                    )

        self.assertEqual(discovered, {})

    def test_discover_excel_download_artifacts_raises_storage_error_when_scandir_fails(self):
        with patch(
            "file_processing.services.export_service.os.scandir",
            side_effect=OSError("storage not readable"),
        ):
            with self.assertRaises(export_service.OutputExcelDownloadStorageError):
                export_service._discover_excel_download_artifacts(r"C:\safe\storage")
