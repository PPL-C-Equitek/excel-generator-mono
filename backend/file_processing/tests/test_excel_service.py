import io
import builtins
from unittest.mock import patch
from django.test import TestCase, SimpleTestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from file_processing.services.excel_service import process_uploaded_excel, _load_workbook, _trim_trailing_nulls

def _build_excel(sheets: dict) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError("openpyxl is required for these tests.")

    wb = Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_xls(sheets: dict) -> bytes:
    try:
        import xlwt
    except ImportError:
        raise RuntimeError("xlwt is required for these tests.")

    wb = xlwt.Workbook()
    for sheet_name, rows in sheets.items():
        ws = wb.add_sheet(sheet_name)
        for row_idx, row in enumerate(rows):
            for col_idx, cell_value in enumerate(row):
                if cell_value is not None:
                    ws.write(row_idx, col_idx, cell_value)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _uploaded(name: str, data: bytes) -> SimpleUploadedFile:
    return SimpleUploadedFile(
        name, data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

class ExcelDataExtractionTests(TestCase):
    EXTRACT_URL = '/upload/'

    def test_single_sheet_first_row_is_header(self):
        data = _build_excel({
            "Mahasiswa": [
                ["NIM", "Nama", "Jurusan"],
                ["12345", "Alice", "Teknik Informatika"],
                ["67890", "Bob", "Sistem Informasi"],
            ]
        })
        response = self.client.post(self.EXTRACT_URL, {'file': _uploaded("data.xlsx", data)})

        self.assertEqual(response.status_code, 200)
        body = response.json()

        rows = self._get_sheet_rows(body, "Mahasiswa")
        self.assertEqual(len(rows), 3)

        self.assertEqual(rows[0][0], "NIM")
        self.assertEqual(rows[0][1], "Nama")
        self.assertEqual(rows[0][2], "Jurusan")

        self.assertEqual(rows[1][0], "12345")
        self.assertEqual(rows[1][1], "Alice")
        self.assertEqual(rows[2][0], "67890")
        self.assertEqual(rows[2][1], "Bob")

    def test_single_sheet_empty_cell_becomes_null_or_empty_string(self):
        data = _build_excel({
            "Nilai": [
                ["NIM", "Mata Kuliah", "Nilai"],
                ["11111", "PPL", None],
                ["22222", None, "A"],
            ]
        })
        response = self.client.post(self.EXTRACT_URL, {'file': _uploaded("nilai.xlsx", data)})

        self.assertEqual(response.status_code, 200)
        rows = self._get_sheet_rows(response.json(), "Nilai")

        self.assertEqual(len(rows[0]), 3)
        self.assertEqual(len(rows[1]), 3)
        self.assertEqual(rows[1][1], "PPL")
        self.assertEqual(rows[1][2], "null")
        self.assertEqual(len(rows[2]), 3)
        self.assertTrue(rows[2][1] is None or rows[2][1] == "")
        self.assertEqual(rows[2][2], "A")

    def test_single_sheet_only_header_no_data_rows_returns_empty_list_or_only_header(self):
        data = _build_excel({
            "Kosong": [
                ["NIM", "Nama", "Jurusan"],
            ]
        })
        response = self.client.post(self.EXTRACT_URL, {'file': _uploaded("empty.xlsx", data)})

        self.assertEqual(response.status_code, 200)
        rows = self._get_sheet_rows(response.json(), "Kosong")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0], ["NIM", "Nama", "Jurusan"])

    def test_single_sheet_completely_empty_sheet_returns_empty(self):
        data = _build_excel({
            "BenarBenarKosong": []
        })
        response = self.client.post(self.EXTRACT_URL, {'file': _uploaded("all_empty.xlsx", data)})

        self.assertEqual(response.status_code, 200)
        rows = self._get_sheet_rows(response.json(), "BenarBenarKosong")
        self.assertEqual(rows, [])

    def test_multi_sheet_each_sheet_is_separate_object_in_json(self):
        data = _build_excel({
            "Mahasiswa": [
                ["NIM", "Nama"],
                ["001", "Alice"],
            ],
            "Dosen": [
                ["NIP", "Nama"],
                ["D01", "Prof. Budi"],
            ],
        })
        response = self.client.post(self.EXTRACT_URL, {'file': _uploaded("multi.xlsx", data)})

        self.assertEqual(response.status_code, 200)
        body = response.json()

        mahasiswa_rows = self._get_sheet_rows(body, "Mahasiswa")
        dosen_rows = self._get_sheet_rows(body, "Dosen")

        self.assertEqual(len(mahasiswa_rows), 2)
        self.assertEqual(len(dosen_rows), 2)

        self.assertEqual(mahasiswa_rows[1][0], "001")
        self.assertEqual(dosen_rows[1][0], "D01")

    def test_multi_sheet_sheet_name_is_unique_key_in_json(self):
        data = _build_excel({
            "Sheet Alpha": [["A", "B"], [1, 2]],
            "Sheet Beta":  [["X", "Y"], [9, 8]],
        })
        response = self.client.post(self.EXTRACT_URL, {'file': _uploaded("keys.xlsx", data)})

        self.assertEqual(response.status_code, 200)
        body = response.json()

        alpha_rows = self._get_sheet_rows(body, "Sheet Alpha")
        beta_rows  = self._get_sheet_rows(body, "Sheet Beta")

        self.assertIsNotNone(alpha_rows, "Kunci 'Sheet Alpha' harus ada dalam JSON.")
        self.assertIsNotNone(beta_rows,  "Kunci 'Sheet Beta' harus ada dalam JSON.")
        self.assertNotEqual(alpha_rows, beta_rows)

    def test_multi_sheet_empty_cells_handled(self):
        data = _build_excel({
            "Sheet1": [
                ["Kol A", "Kol B"],
                ["val", None],
            ],
            "Sheet2": [
                ["P", "Q", None, None],
                [None, "val2"],
            ],
        })
        response = self.client.post(self.EXTRACT_URL, {'file': _uploaded("multi_empty.xlsx", data)})

        self.assertEqual(response.status_code, 200)
        body = response.json()

        rows1 = self._get_sheet_rows(body, "Sheet1")
        rows2 = self._get_sheet_rows(body, "Sheet2")

        self.assertEqual(len(rows1[1]), 2)
        self.assertEqual(rows1[1][0], "val")
        self.assertEqual(rows1[1][1], "null")
        self.assertEqual(len(rows2[0]), 3)
        self.assertEqual(rows2[0][2], "nullx2")
        self.assertEqual(len(rows2[1]), 3)
        v2 = rows2[1][0]
        self.assertTrue(v2 is None or v2 == "", f"Sheet2 P empty cell: got {v2!r}")
        self.assertEqual(rows2[1][1], "val2")
        self.assertEqual(rows2[1][2], "nullx2")

    def test_multi_sheet_all_sheets_empty_returns_valid_json_structure(self):
        data = _build_excel({
            "EmptyA": [],
            "EmptyB": [],
        })
        response = self.client.post(self.EXTRACT_URL, {'file': _uploaded("all_empty_multi.xlsx", data)})

        self.assertNotEqual(response.status_code, 500)
        if response.status_code == 200:
            body = response.json()
            for sheet_name in ["EmptyA", "EmptyB"]:
                rows = self._get_sheet_rows(body, sheet_name)
                self.assertIsNotNone(rows, f"Kunci '{sheet_name}' harus ada dalam JSON.")

    def test_multi_sheet_partial_empty_sheet_still_extracts_other_sheets(self):
        data = _build_excel({
            "AdaData": [
                ["ID", "Nilai"],
                ["A1", 90],
            ],
            "TanpaData": [],
        })
        response = self.client.post(self.EXTRACT_URL, {'file': _uploaded("partial_empty.xlsx", data)})

        self.assertNotEqual(response.status_code, 500)
        if response.status_code == 200:
            body = response.json()
            ada_rows = self._get_sheet_rows(body, "AdaData")
            self.assertIsNotNone(ada_rows)
            self.assertEqual(len(ada_rows), 2, "Sheet 'AdaData' harus berisi baris data.")

    def _get_sheet_rows(self, body: dict, sheet_name: str):
        if "extracted" in body:
            return body["extracted"].get(sheet_name)

        if "data" in body and isinstance(body["data"], dict):
            return body["data"].get(sheet_name)

        return body.get(sheet_name)
    
    def test_process_uploaded_excel_with_nonexistent_path_returns_error(self):
        success, error, data = process_uploaded_excel("/tmp/file_that_does_not_exist.xlsx")

        self.assertFalse(success)
        self.assertIsNone(data)
        self.assertIn("File not found", error)

    def test_openpyxl_not_installed_raises_runtime_error(self):
        real_import = builtins.__import__

        def fake_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("missing")
            return real_import(name, *args, **kwargs)

        with patch("builtins.__import__", side_effect=fake_import):
            with self.assertRaises(RuntimeError):
                _load_workbook("dummy.xlsx")

    def test_row_with_all_null_cells_is_skipped(self):
        data = _build_excel({
            "Sheet1": [
                ["A", "B", "C"],
                [None, None, None],
                ["1", "2", "3"],
            ]
        })

        response = self.client.post(
            self.EXTRACT_URL,
            {"file": _uploaded("null_row.xlsx", data)},
        )

        self.assertEqual(response.status_code, 200)

        rows = self._get_sheet_rows(response.json(), "Sheet1")

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0], ["A", "B", "C"])
        self.assertEqual(rows[1], ["1", "2", "3"])

    def test_process_uploaded_excel_xls_path_branch(self):
        with patch("file_processing.services.excel_service.parse_xls") as mock_parse:
            mock_parse.return_value = {"Sheet1": [["A", "B"]]}

            success, error, data = process_uploaded_excel("/tmp/test.xls")

            self.assertTrue(success)
            self.assertIsNone(error)
            self.assertEqual(data, {"Sheet1": [["A", "B"]]})

    def test_get_xls_sheet_count_without_release_resources(self):
        from file_processing.services.upload_service import _get_xls_sheet_count
        from unittest.mock import MagicMock
        with patch('xlrd.open_workbook') as mock_open:
            mock_wb = MagicMock()
            mock_wb.nsheets = 5
            del mock_wb.release_resources
            mock_open.return_value = mock_wb
            
            mock_file = MagicMock()
            mock_file.read.return_value = b"dummy"
            
            count = _get_xls_sheet_count(mock_file)
            self.assertEqual(count, 5)

    @patch('file_processing.services.upload_service.validate_file')
    def test_xls_single_sheet_data_extraction(self, mock_validate):
        mock_validate.return_value = (True, None)
        data = _build_xls({
            "Mahasiswa": [
                ["NIM", "Nama"],
                ["12345", "Alice"],
            ]
        })
        f = SimpleUploadedFile("data.xls", data, content_type="application/vnd.ms-excel")
        response = self.client.post(self.EXTRACT_URL, {'file': f})

        self.assertEqual(response.status_code, 200)
        body = response.json()
        rows = self._get_sheet_rows(body, "Mahasiswa")
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0], ["NIM", "Nama"])
        self.assertEqual(rows[1], ["12345", "Alice"])

    @patch('file_processing.services.upload_service.validate_file')
    def test_xls_empty_cells_and_float_conversion(self, mock_validate):
        mock_validate.return_value = (True, None)
        data = _build_xls({
            "Nilai": [
                ["NIM", "Mata Kuliah", "Nilai", "Blank"],
                [11111.0, "PPL", "", ""],
                [22222.0, "", "A", ""],
            ]
        })
        f = SimpleUploadedFile("nilai.xls", data, content_type="application/vnd.ms-excel")
        response = self.client.post(self.EXTRACT_URL, {'file': f})

        self.assertEqual(response.status_code, 200)
        rows = self._get_sheet_rows(response.json(), "Nilai")

        self.assertEqual(rows[0], ["NIM", "Mata Kuliah", "Nilai", "Blank"])
        self.assertEqual(rows[1], ["11111", "PPL", "nullx2"])
        self.assertEqual(rows[2], ["22222", None, "A", "null"])

    @patch('file_processing.services.upload_service.validate_file')
    def test_xls_completely_empty_sheet_returns_empty(self, mock_validate):
        mock_validate.return_value = (True, None)
        data = _build_xls({
            "BenarBenarKosong": []
        })
        f = SimpleUploadedFile("all_empty.xls", data, content_type="application/vnd.ms-excel")
        response = self.client.post(self.EXTRACT_URL, {'file': f})

        self.assertEqual(response.status_code, 200)
        rows = self._get_sheet_rows(response.json(), "BenarBenarKosong")
        self.assertEqual(rows, [])

    def test_xlrd_not_installed_raises_runtime_error(self):
        from file_processing.services.excel_service import _load_xls_workbook
        real_import = builtins.__import__

        def fake_import(name, *args, **kwargs):
            if name == "xlrd":
                raise ImportError("missing")
            return real_import(name, *args, **kwargs)

        with patch("builtins.__import__", side_effect=fake_import):
            with self.assertRaises(RuntimeError):
                _load_xls_workbook("dummy.xls")

    def test_load_xls_workbook_file_not_found(self):
        from file_processing.services.excel_service import _load_xls_workbook
        with self.assertRaises(FileNotFoundError):
            _load_xls_workbook("/nonexistent_fake_path.xls")

    def test_load_xls_workbook_corrupted_file(self):
        import tempfile
        import os
        from file_processing.services.excel_service import _load_xls_workbook

        fd, path = tempfile.mkstemp(suffix=".xls")
        try:
            with os.fdopen(fd, 'wb') as f:
                f.write(b"NOT_A_VALID_XLS_FILE")
            with self.assertRaises(ValueError):
                _load_xls_workbook(path)
        finally:
            os.remove(path)

    def test_process_uploaded_excel_with_file_object(self):
        data = _build_excel({
            "Sheet1": [["A", "B"]]
        })

        f = io.BytesIO(data)

        success, error, result = process_uploaded_excel(f)

        self.assertTrue(success)
        self.assertIsNone(error)
        self.assertIn("Sheet1", result)

    @patch('file_processing.services.upload_service.validate_file')
    def test_xls_row_with_all_null_cells_is_skipped(self, mock_validate):
        mock_validate.return_value = (True, None)

        data = _build_xls({
            "Sheet1": [
                ["A", "B", "C"],
                ["", "", ""],
                ["1", "2", "3"]
            ]
        })

        f = SimpleUploadedFile(
            "null_row.xls",
            data,
            content_type="application/vnd.ms-excel"
        )

        response = self.client.post(self.EXTRACT_URL, {"file": f})

        self.assertEqual(response.status_code, 200)

        rows = self._get_sheet_rows(response.json(), "Sheet1")

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0], ["A", "B", "C"])
        self.assertEqual(rows[1], ["1", "2", "3"])

class FileValidationTests(TestCase):
    UPLOAD_URL = '/upload/'

    def _make_file(self, name: str, content: bytes, content_type: str) -> SimpleUploadedFile:
        return SimpleUploadedFile(name, content, content_type=content_type)

    def _valid_xlsx_bytes(self) -> bytes:

        return _build_excel({"Sheet1": [["A", "B"], [1, 2]]})

    def test_extension_xlsx_is_accepted(self):
        valid_xlsx = self._make_file(
            "data.xlsx",
            self._valid_xlsx_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(self.UPLOAD_URL, {'file': valid_xlsx})
        self.assertEqual(response.status_code, 200)

    def test_magic_number_valid_xlsx_passes(self):
        valid_bytes = self._valid_xlsx_bytes()
        self.assertEqual(valid_bytes[:4], b'\x50\x4B\x03\x04',
                         "Helper _build_excel harus menghasilkan file ZIP valid")

        valid_xlsx = self._make_file(
            "valid.xlsx",
            valid_bytes,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(self.UPLOAD_URL, {'file': valid_xlsx})
        self.assertEqual(response.status_code, 200)

    def test_file_exceeding_10mb_is_rejected(self):
        oversized_content = b'\x50\x4B\x03\x04' + b'A' * (10 * 1024 * 1024)
        oversized_file = self._make_file(
            "besar.xlsx",
            oversized_content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(self.UPLOAD_URL, {'file': oversized_file})
        self.assertIn(response.status_code, [400, 413])
        body = response.json()
        self.assertIn("message", body)

    def test_file_exactly_10mb_is_accepted_or_rejected_gracefully(self):
        exactly_10mb = b'\x50\x4B\x03\x04' + b'B' * (10 * 1024 * 1024 - 4)
        file_10mb = self._make_file(
            "pas10mb.xlsx",
            exactly_10mb,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(self.UPLOAD_URL, {'file': file_10mb})
        self.assertNotEqual(response.status_code, 500,
                            "File tepat 10 MB tidak boleh menyebabkan server error.")

    def test_small_valid_file_is_accepted(self):
        valid_xlsx = self._make_file(
            "kecil.xlsx",
            self._valid_xlsx_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(self.UPLOAD_URL, {'file': valid_xlsx})
        self.assertNotEqual(response.status_code, 400)

    def test_mime_type_xlsx_is_accepted(self):
        file_correct_mime = self._make_file(
            "data.xlsx",
            self._valid_xlsx_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(self.UPLOAD_URL, {'file': file_correct_mime})
        self.assertEqual(response.status_code, 200)

    def test_mime_type_xls_is_accepted(self):
        file_xls_mime = self._make_file(
            "data.xls",
            self._valid_xlsx_bytes(),
            "application/vnd.ms-excel",
        )
        response = self.client.post(self.UPLOAD_URL, {'file': file_xls_mime})
        body = response.json()
        if response.status_code == 400:
            error_msg = body.get("error", "").lower()
            self.assertNotIn("mime", error_msg,
                             "File .xls tidak boleh ditolak karena MIME type.")

    def test_corrupted_xlsx_returns_400_with_error_message(self):
        corrupted_content = b'\x50\x4B\x03\x04' + b'\xFF\xFE\xFD' * 50
        corrupted_file = self._make_file(
            "corrupt.xlsx",
            corrupted_content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(self.UPLOAD_URL, {'file': corrupted_file})
        self.assertEqual(response.status_code, 400,
                         "File corrupted harus mengembalikan 400, bukan 500.")
        body = response.json()
        self.assertIn("message", body)
        self.assertTrue(len(body["message"]) > 0)

    def test_corrupted_xls_returns_400_with_error_message(self):
        corrupted_content = b'\xD0\xCF\x11\xE0' + b'\x00\x01\x02\x03' * 30
        corrupted_xls = self._make_file(
            "corrupt.xls",
            corrupted_content,
            "application/vnd.ms-excel",
        )
        response = self.client.post(self.UPLOAD_URL, {'file': corrupted_xls})
        self.assertEqual(response.status_code, 400)
        if response.status_code == 400:
            body = response.json()
            self.assertIn("message", body)

    def test_empty_file_is_rejected_with_error(self):
        empty_file = self._make_file(
            "empty.xlsx",
            b"",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(self.UPLOAD_URL, {'file': empty_file})
        self.assertEqual(response.status_code, 400)
        body = response.json()
        self.assertIn("message", body)

    def test_corrupted_file_error_message_is_user_friendly(self):
        corrupted_content = b'\x50\x4B\x03\x04' + b'\xDE\xAD\xBE\xEF' * 20
        corrupted_file = self._make_file(
            "corrupt2.xlsx",
            corrupted_content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(self.UPLOAD_URL, {'file': corrupted_file})
        if response.status_code == 400:
            body = response.json()
            error_msg = body.get("message", "")
            self.assertNotIn("Traceback", error_msg)
            self.assertNotIn("raise ", error_msg)
            self.assertTrue(len(error_msg) > 0)

    def test_xlsx_with_more_than_100_sheets_is_rejected(self):
        sheets = {f"Sheet{i}": [["A"], [i]] for i in range(1, 102)}
        too_many_sheets = self._make_file(
            "many_sheets.xlsx",
            _build_excel(sheets),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(self.UPLOAD_URL, {"file": too_many_sheets})

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.json().get("message"),
            "Excel has too many sheets (maximum 100).",
        )

    def test_xls_with_more_than_100_sheets_is_rejected(self):
        sheets = {f"Sheet{i}": [["A"], [i]] for i in range(1, 102)}
        too_many_sheets = self._make_file(
            "many_sheets.xlsx",
            _build_xls(sheets),
            "application/vnd.ms-excel",
        )

        response = self.client.post(self.UPLOAD_URL, {"file": too_many_sheets})

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.json().get("message"),
            "Excel has too many sheets (maximum 100).",
        )


class ExcelServiceDirectUnitTests(SimpleTestCase):
    def test_load_workbook_corrupted_raises_value_error(self):
        with self.assertRaises(ValueError):
            _load_workbook(io.BytesIO(b"corrupt data"))

    def test_empty_xls_sheet_returns_empty_rows(self):
        from file_processing.services.excel_service import parse_xls
        import tempfile
        import os
        data = _build_xls({"Sheet1": []})
        with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as tmp:
            tmp.write(data)
            tmp_path = tmp.name
        try:
            result = parse_xls(tmp_path)
            self.assertEqual(result, {"Sheet1": []})
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    def test_process_uploaded_excel_xlsx_with_ole_signature(self):
        import tempfile
        import os
        data = _build_xls({"Sheet1": [["A", "B"], [1, 2]]})
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(data)
            tmp_path = tmp.name
        try:
            success, error, result = process_uploaded_excel(tmp_path)
            self.assertTrue(success)
            self.assertIsNone(error)
            self.assertIn("Sheet1", result)
            self.assertEqual(result["Sheet1"], [["A", "B"], ["1", "2"]])
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    @patch("file_processing.services.excel_service.parse_excel", side_effect=Exception("mocked exception"))
    def test_process_uploaded_excel_general_exception(self, mock_parse):
        success, error, data = process_uploaded_excel(io.BytesIO(b"some data"))
        self.assertFalse(success)
        self.assertEqual(error, "Invalid or corrupted Excel file.")
        self.assertIsNone(data)

    def test_xls_sheet_to_rows_with_zero_columns(self):
        from unittest.mock import MagicMock
        from file_processing.services.excel_service import _xls_sheet_to_rows
        mock_sheet = MagicMock()
        mock_sheet.nrows = 1
        mock_sheet.ncols = 0
        mock_sheet.cell_value = MagicMock()

        result = _xls_sheet_to_rows(mock_sheet)
        self.assertEqual(result, [])

    def test_has_ole_signature_handles_exception_and_returns_false(self):
        from file_processing.services.excel_service import _has_ole_signature
        result = _has_ole_signature("/nonexistent_path_that_should_fail.xlsx")
        self.assertIs(result, False)

    def test_process_uploaded_excel_xls_with_file_object(self):
        xls_bytes = _build_xls({
            "Sheet1": [["Kolom A", "Kolom B"], ["Nilai 1", "Nilai 2"]]
        })
        f = io.BytesIO(xls_bytes)
        f.name = "test.xls"

        success, error, data = process_uploaded_excel(f)
        self.assertTrue(success, f"Expected success but got error: {error}")
        self.assertIsNone(error)
        self.assertIsNotNone(data)
        self.assertIn("Sheet1", data)
        self.assertEqual(data["Sheet1"], [["Kolom A", "Kolom B"], ["Nilai 1", "Nilai 2"]])

    def test_has_ole_signature_with_file_like_object(self):
        from file_processing.services.excel_service import _has_ole_signature
        from file_processing.services.excel_service import OLE_SIGNATURE
        
        f_correct = io.BytesIO(OLE_SIGNATURE + b"extra data")
        self.assertTrue(_has_ole_signature(f_correct))
        
        f_incorrect = io.BytesIO(b"incorrect signature data")
        self.assertFalse(_has_ole_signature(f_incorrect))

    def test_xls_with_ole_signature_file_object(self):
        xls_bytes = _build_xls({
            "Sheet1": [["Val1"]]
        })
        f = io.BytesIO(xls_bytes)
        f.name = "legacy.xlsx"

        success, error, data = process_uploaded_excel(f)
        self.assertTrue(success, f"Expected success but got error: {error}")
        self.assertIsNone(error)
        self.assertIsNotNone(data)
        self.assertIn("Sheet1", data)


